import sqlite3
import random
from datetime import datetime, timezone, timedelta
from flask import Flask, request, session, redirect, url_for, g, Response, render_template_string
from email.mime.text import MIMEText
from decimal import Decimal
import json
import ssl
import smtplib
from io import BytesIO
import openpyxl
from openpyxl import Workbook
import uuid
import os
from werkzeug.utils import secure_filename

# --- Konfigurasi Aplikasi ---
app = Flask(__name__)
app.config['SECRET_KEY'] = 'kunci-rahasia-yang-sangat-aman-12345'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)
DATABASE = 'accounting.db'

basedir = os.path.abspath(os.path.dirname(__file__)) 

# Lihat Bukti tf QRIS
UPLOAD_FOLDER = os.path.join(basedir, 'static', 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Buat folder jika belum ada
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# =========================
# ‚ø° KONFIGURASI EMAIL (Gunakan App Password BARU Anda)
# =========================
EMAIL_PENGIRIM = "sigrameei@gmail.com"
PASSWORD_EMAIL = "knjw eqmm elga fdce" #

UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# --- Fungsi Database (SQLite3) ---
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

@app.before_request
def make_session_permanent():
    session.permanent = True # Mengaktifkan waktu expired sesi
    app.permanent_session_lifetime = timedelta(minutes=30) # Set 30 menit

@app.before_request
def check_session_timeout():
    """Mengecek apakah admin sudah idle lebih dari 30 menit"""
    # Abaikan pemeriksaan untuk route static atau login
    if request.endpoint in ('static', 'login', 'index', 'logout', 'verify_otp', 'login_page'):
        return

    if 'username' in session:
        # KECUALI CONSUMER (Consumer bebas waktu)
        if session.get('role') == 'consumer':
            return

        # Logika Timeout untuk ADMIN
        now = datetime.now(timezone.utc)
        last_active = session.get('last_active')

        if last_active:
            last_active_dt = datetime.fromisoformat(last_active)
            
            # Hitung selisih waktu (1800 detik = 30 menit)
            if (now - last_active_dt).total_seconds() > 1800: 
                # Hapus sesi
                session.pop('username', None)
                session.pop('role', None)
                session.pop('last_active', None)
                
                # --- PERUBAHAN DI SINI ---
                # Redirect ke login_page dengan status timeout
                return redirect(url_for('login_page', timeout='true'))

        # Update waktu terakhir aktif
        session['last_active'] = now.isoformat()

def generate_journal_code(db, entry_datetime=None):
    """
    Membuat ID Jurnal unik dengan format: SG-YYYYMMDD-HHMMSS-RANDOM5
    entry_datetime: objek datetime (jika manual) atau None (untuk realtime)
    """
    if entry_datetime is None:
        entry_datetime = datetime.now()
    
    date_str = entry_datetime.strftime('%Y%m%d')
    time_str = entry_datetime.strftime('%H%M%S')
    
    # Loop untuk menjamin keunikan (meski kemungkinan tabrakan sangat kecil)
    while True:
        rand_str = f"{random.randint(10000, 99999)}"
        new_code = f"SG{date_str}{time_str}{rand_str}"
        
        # Cek ke database apakah kode ini sudah ada
        exists = db.execute("SELECT 1 FROM journal_entries WHERE journal_code = ?", (new_code,)).fetchone()
        if not exists:
            return new_code

def init_db():
    print("Membuat database...")
    with app.app_context():
        db = get_db()
        db.execute("PRAGMA foreign_keys = ON;")
        
        with db:
            # --- 1. PEMBUATAN STRUKTUR TABEL ---
            db.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('admin', 'consumer')),
                security_answer TEXT NOT NULL,
                phone_number TEXT
            )
            ''')
            db.execute('''
            CREATE TABLE IF NOT EXISTS chart_of_accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_code TEXT UNIQUE NOT NULL,
                account_name TEXT NOT NULL,
                account_type TEXT NOT NULL
            )
            ''')
            db.execute('''
            CREATE TABLE IF NOT EXISTS company_info (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_name TEXT NOT NULL,
                accounting_period TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''')
            db.execute('''
            CREATE TABLE IF NOT EXISTS journal_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                journal_code TEXT UNIQUE NOT NULL,
                entry_timestamp TEXT DEFAULT (strftime('%Y-%m-%d %H:%M:%S', 'now', 'localtime')),
                description TEXT NOT NULL
            )
            ''')
            db.execute('''
            CREATE TABLE IF NOT EXISTS journal_details (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entry_id INTEGER NOT NULL,
                account_code TEXT NOT NULL,
                debit REAL DEFAULT 0,
                credit REAL DEFAULT 0,
                FOREIGN KEY (entry_id) REFERENCES journal_entries (id) ON DELETE CASCADE,
                FOREIGN KEY (account_code) REFERENCES chart_of_accounts (account_code) ON DELETE RESTRICT
            )
            ''')
            db.execute('''
            CREATE TABLE IF NOT EXISTS inventory_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_name TEXT UNIQUE NOT NULL,
                inventory_account TEXT NOT NULL, 
                expense_cogs_account TEXT NOT NULL, 
                sales_account TEXT, 
                sale_price_per_unit REAL DEFAULT 0,
                FOREIGN KEY (inventory_account) REFERENCES chart_of_accounts (account_code),
                FOREIGN KEY (expense_cogs_account) REFERENCES chart_of_accounts (account_code),
                FOREIGN KEY (sales_account) REFERENCES chart_of_accounts (account_code)
            )
            ''')
            db.execute('''
            CREATE TABLE IF NOT EXISTS inventory_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_id INTEGER NOT NULL,
                trx_date DATE NOT NULL,
                trx_type TEXT NOT NULL CHECK(trx_type IN ('purchase', 'sale')),
                description TEXT,
                qty REAL NOT NULL,
                cost_per_unit REAL DEFAULT 0,
                sale_price_per_unit REAL DEFAULT 0,
                FOREIGN KEY (item_id) REFERENCES inventory_items (id)
            )
            ''')
            db.execute('''
            CREATE TABLE IF NOT EXISTS contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                contact_name TEXT NOT NULL,
                contact_code TEXT UNIQUE NOT NULL,
                contact_type TEXT NOT NULL CHECK(contact_type IN ('AR', 'AP')),
                control_account TEXT NOT NULL
            )
            ''')
            db.execute('''
            CREATE TABLE IF NOT EXISTS sub_ledger_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                contact_id INTEGER NOT NULL,
                trx_date DATE NOT NULL,
                description TEXT NOT NULL,
                ref TEXT,
                debit REAL DEFAULT 0,
                credit REAL DEFAULT 0,
                FOREIGN KEY (contact_id) REFERENCES contacts (id)
            )
            ''')
            db.execute('''
            CREATE TABLE IF NOT EXISTS online_payments (
                order_id TEXT PRIMARY KEY,
                amount REAL NOT NULL,
                status TEXT DEFAULT 'pending','verified', 'rejected'
                proof_image TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                item_id INTEGER,
                qty REAL,
                buyer_name TEXT
            )
            ''')
            
            # --- 2. PEMBUATAN DATA AWAL (INSERT OR IGNORE) ---
            
            # A. Users
            db.execute("INSERT OR IGNORE INTO users (username, password, role, security_answer, phone_number) VALUES (?, ?, ?, ?, ?)", 
                       ('admin', 'admin123', 'admin', 'sigrameei@gmail.com', '082241915050'))
            db.execute("INSERT OR IGNORE INTO users (username, password, role, security_answer, phone_number) VALUES (?, ?, ?, ?, ?)", 
                       ('consumer', 'consumer123', 'consumer', 'consumer@mail.com', '081234567890'))
            
            # B. Chart of Accounts
            accounts_data = [
                # Aset Lancar
                ('1101', 'Kas', 'Aset Lancar'), 
                ('1102', 'Piutang Dagang', 'Aset Lancar'), 
                ('1103', 'Perlengkapan', 'Aset Lancar'), 
                ('1105', 'Persediaan Ikan Gurame', 'Aset Lancar'),
                ('1108', 'Persediaan Pakan Ikan', 'Aset Lancar'),
                ('1109', 'Persediaan Obat-obatan', 'Aset Lancar'), # <-- BARU
                # Aset Tetap
                ('1201', 'Peralatan', 'Aset Tetap'), 
                ('1202', 'Akumulasi Penyusutan Peralatan', 'Aset Tetap'), 
                ('1203', 'Kendaraan', 'Aset Tetap'), 
                ('1204', 'Akumulasi Penyusutan Kendaraan', 'Aset Tetap'), 
                ('1205', 'Bangunan', 'Aset Tetap'), 
                ('1206', 'Akumulasi Penyusutan Bangunan', 'Aset Tetap'), 
                ('1301', 'Tanah', 'Aset Tetap'),
                # Liabilitas
                ('2101', 'Utang Dagang', 'Liabilitas'), 
                # Ekuitas
                ('3101', 'Modal', 'Ekuitas'), 
                ('3102', 'Prive', 'Ekuitas'), 
                ('3103', 'Ikhtisar Laba Rugi', 'Ekuitas'),
                # Pendapatan
                ('4101', 'Penjualan Ikan Gurame', 'Pendapatan'), 
                ('4102', 'Pendapatan Lainnya', 'Pendapatan'),
                # Beban
                ('5101', 'HPP Ikan Gurame', 'Beban'), 
                ('6101', 'Beban Akomodasi', 'Beban'), 
                ('6102', 'Beban Listrik dan Air', 'Beban'), 
                ('6103', 'Beban Perlengkapan', 'Beban'), 
                ('6104', 'Beban Gaji', 'Beban'), 
                ('6105', 'Beban Pakan Ikan', 'Beban'),
                ('6106', 'Beban Obat-obatan', 'Beban'), # <-- BARU
                ('6201', 'Biaya Penyusutan Peralatan', 'Beban'), 
                ('6202', 'Biaya Penyusutan Kendaraan', 'Beban'), 
                ('6203', 'Biaya Penyusutan Bangunan', 'Beban')
            ]
            
            for acc in accounts_data:
                db.execute("INSERT OR IGNORE INTO chart_of_accounts (account_code, account_name, account_type) VALUES (?, ?, ?)", acc)

            # C. Inventory Items
            print("Mengisi data Inventory...")
            # Item 1: Ikan Gurame
            db.execute("""
                INSERT OR IGNORE INTO inventory_items 
                (item_name, inventory_account, expense_cogs_account, sales_account, sale_price_per_unit) 
                VALUES (?, ?, ?, ?, ?)""",
                ('Ikan Gurame', '1105', '5101', '4101', '15000')
            )
            # Item 2: Pakan Ikan
            db.execute("""
                INSERT OR IGNORE INTO inventory_items 
                (item_name, inventory_account, expense_cogs_account, sales_account, sale_price_per_unit) 
                VALUES (?, ?, ?, ?, ?)""",
                ('Pakan Ikan', '1108', '6105', None, 0)
            )
            # Item 3: Obat-obatan (BARU)
            db.execute("""
                INSERT OR IGNORE INTO inventory_items 
                (item_name, inventory_account, expense_cogs_account, sales_account, sale_price_per_unit) 
                VALUES (?, ?, ?, ?, ?)""",
                ('Obat-obatan', '1109', '6106', None, 0)
            )
            
            # D. Kontak (Buku Pembantu)
            contacts_data = [
                ('Bu Sri', 'A101', 'AR', '1102'), 
                ('Pak Bagas', 'A102', 'AR', '1102'), 
                ('Bu Fatimah', 'A103', 'AR', '1102'),
                ('Mas Ahmad', 'B201', 'AP', '2101'), 
                ('Bu Annisa', 'B202', 'AP', '2101')
            ]
            for contact in contacts_data:
                db.execute("INSERT OR IGNORE INTO contacts (contact_name, contact_code, contact_type, control_account) VALUES (?, ?, ?, ?)", contact)

            # E. Saldo Awal
            saldo_awal_check = db.execute("SELECT id FROM journal_entries WHERE id = 1").fetchone()
            if not saldo_awal_check:
                print("Membuat Jurnal Saldo Awal...")
                saldo_awal_ts = '2025-01-01 00:00:00'
                j_code = generate_journal_code(db) 
                cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                    (j_code, saldo_awal_ts, "Saldo Awal per 1 Jan 2025"))
                entry_id = cursor.lastrowid

                # Saldo Awal Buku Pembantu
                db.execute("INSERT INTO sub_ledger_log (contact_id, trx_date, description, ref, debit) VALUES (?, ?, ?, ?, ?)", (1, '2025-01-01', 'Saldo', 'NS', 800000))
                db.execute("INSERT INTO sub_ledger_log (contact_id, trx_date, description, ref, debit) VALUES (?, ?, ?, ?, ?)", (2, '2025-01-01', 'Saldo', 'NS', 1000000))
                db.execute("INSERT INTO sub_ledger_log (contact_id, trx_date, description, ref, debit) VALUES (?, ?, ?, ?, ?)", (3, '2025-01-01', 'Saldo', 'NS', 1200000))
                db.execute("INSERT INTO sub_ledger_log (contact_id, trx_date, description, ref, credit) VALUES (?, ?, ?, ?, ?)", (4, '2025-01-01', 'Saldo', 'NS', 3200000))
                db.execute("INSERT INTO sub_ledger_log (contact_id, trx_date, description, ref, credit) VALUES (?, ?, ?, ?, ?)", (5, '2025-01-01', 'Saldo', 'NS', 1800000))

    print("Inisialisasi database selesai.")

# ==================================
#  DATABASE Bukbes pembantu
# ==================================
def init_ledger_separate():
    """Initialize database untuk buku besar pembantu AR & AP"""
    db = get_db()
    
    try:
        # Tabel untuk Piutang (AR)
        db.execute('''
            CREATE TABLE IF NOT EXISTS ledger_ar (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                transaction_date DATE NOT NULL,
                debtor_name TEXT NOT NULL,       -- Nama debitur
                address TEXT,
                phone TEXT,
                description TEXT NOT NULL,
                reference TEXT NOT NULL,         -- No. referensi: INV-001, dll
                debit DECIMAL(15,2) DEFAULT 0,   -- Penambahan piutang
                credit DECIMAL(15,2) DEFAULT 0,  -- Pengurangan piutang
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Tabel untuk Utang (AP)
        db.execute('''
            CREATE TABLE IF NOT EXISTS ledger_ap (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                transaction_date DATE NOT NULL,
                creditor_name TEXT NOT NULL,     -- Nama kreditur
                address TEXT,
                phone TEXT,
                description TEXT NOT NULL,
                reference TEXT NOT NULL,         -- No. referensi: PO-001, dll
                debit DECIMAL(15,2) DEFAULT 0,   -- Pengurangan utang
                credit DECIMAL(15,2) DEFAULT 0,  -- Penambahan utang
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Index untuk performa
        db.execute('CREATE INDEX IF NOT EXISTS idx_ar_date ON ledger_ar(transaction_date)')
        db.execute('CREATE INDEX IF NOT EXISTS idx_ar_name ON ledger_ar(debtor_name)')
        db.execute('CREATE INDEX IF NOT EXISTS idx_ap_date ON ledger_ap(transaction_date)')
        db.execute('CREATE INDEX IF NOT EXISTS idx_ap_name ON ledger_ap(creditor_name)')
        
        db.commit()
        print("‚úÖ Database buku besar pembantu AR & AP berhasil dibuat!")
        return True
        
    except Exception as e:
        print(f"‚ùå Error: {str(e)}")
        return False

@app.route("/admin/init-ledger-separate")
def init_ledger_separate_route():
    """Route untuk inisialisasi database AR & AP"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    success = init_ledger_separate()
    
    if success:
        return """
        <div style="padding: 40px; text-align: center;">
            <h3 style="color: green;">‚úÖ Database AR & AP Berhasil Dibuat!</h3>
            <p>Sekarang Anda bisa mulai menggunakan buku besar pembantu.</p>
            <div style="margin: 30px;">
                <a href="/admin/ledger-ar" 
                   style="background-color: #007bff; color: white; padding: 15px 30px; text-decoration: none; border-radius: 5px; font-size: 16px; margin: 10px;">
                    üìó Buku Piutang (AR)
                </a>
                <a href="/admin/ledger-ap" 
                   style="background-color: #28a745; color: white; padding: 15px 30px; text-decoration: none; border-radius: 5px; font-size: 16px; margin: 10px;">
                    üìò Buku Utang (AP)
                </a>
            </div>
        </div>
        """
    else:
        return """
        <div style="padding: 40px; text-align: center;">
            <h3 style="color: red;">‚ùå Gagal membuat database</h3>
            <p>Silakan cek console untuk detail error.</p>
        </div>
        """
# =========================
# KIRIM OTP 
# =========================
def send_otp_email(email_penerima, kode_otp):
    """Mengirim email berisi kode OTP."""
    
    # Pesan yang akan dikirim
    pesan_html = f"""
    <html>
    <body>
        <p>Halo,</p>
        <p>Gunakan kode verifikasi ini untuk login. Kode akan kedaluwarsa dalam 1 menit.</p>
        <h2 style="color: #333; font-size: 28px; text-align: center;">{kode_otp}</h2>
        <p>Jika Anda tidak meminta kode ini, abaikan email ini.</p>
    </body>
    </html>
    """
    
    msg = MIMEText(pesan_html, 'html')
    msg['Subject'] = f"Kode Verifikasi Login Anda adalah {kode_otp}"
    msg['From'] = EMAIL_PENGIRIM
    msg['To'] = email_penerima
    
    # Buat koneksi aman
    context = ssl.create_default_context()
    
    try:
        # Gunakan server SMTP Gmail dengan SSL
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(EMAIL_PENGIRIM, PASSWORD_EMAIL)
            server.sendmail(EMAIL_PENGIRIM, email_penerima, msg.as_string())
        
        print(f"Email OTP berhasil dikirim ke {email_penerima}")
        return True # Sukses
        
    except Exception as e:
        # Jika gagal (misal: password salah, koneksi gagal)
        print(f"ERROR: Gagal mengirim email ke {email_penerima}. Alasan: {e}")
        return False # Gagal

@app.route('/toggle-sidebar')
def toggle_sidebar():
    """
    Fungsi 100% Python untuk mengubah status sidebar.
    """
    # Ambil status sekarang (default False/Tertutup jika belum ada)
    current_status = session.get('sidebar_open', False)
    
    # Balik statusnya (Jika True jadi False, Jika False jadi True)
    session['sidebar_open'] = not current_status
    
    # Kembali ke halaman sebelumnya (Refresh halaman)
    return redirect(request.referrer or url_for('index'))

def render_page(title, body_content, sidebar_content=None, error_message=None):
    """
    Merender halaman HTML.
    VERSI: 100% PYTHON (Server-Side Logic).
    Status sidebar disimpan di Flask Session.
    """
    
    # 1. Cek Status Sidebar dari Session Python
    # Default False (Tertutup)
    is_open = session.get('sidebar_open', False)
    
    # 2. Tentukan CSS berdasarkan status Python
    if is_open:
        sidebar_css = "transform: translateX(0);"  # Muncul
        content_css = "margin-left: 250px; width: calc(100% - 250px);" # Geser Kanan
    else:
        sidebar_css = "transform: translateX(-100%);" # Sembunyi
        content_css = "margin-left: 0; width: 100%;"    # Full Width

    sidebar_html = ""
    open_sidebar_btn = "" 
    error_html = ""

    if error_message:
        error_html = f'<p style="color: red; border: 1px solid red; padding: 10px; background: #ffe6e6; border-radius: 5px;"><b>ERROR:</b> {error_message}</p>'

    # Hanya render sidebar jika ada kontennya
    if sidebar_content:
        # Tombol Tutup (Link ke Python Route)
        close_link = """
        <p style="margin-top: 30px; border-top: 1px solid rgba(255,255,255,0.3); padding-top:10px;">
            <a href="/toggle-sidebar" style="color: #ff9999; font-weight: 300 !important; text-decoration: none;">‚úñ Tutup Sidebar</a>
        </p>
        """
        
        sidebar_html = f"""
        <div class="sidebar-galaxy" style="{sidebar_css}">
            {sidebar_content}
            {close_link}
        </div>
        """
        
        # Tombol Buka (Link ke Python Route), hanya muncul jika sidebar tertutup
        if not is_open:
            open_sidebar_btn = """
            <a href="/toggle-sidebar" class="btn-auth">
                <span class="icon-arrow" style="margin-right: 8px;">‚ûú</span> BUKA SIDEBAR
            </a>
            """

    # --- HTML & CSS UTAMA ---
    html = f"""
    <!DOCTYPE html>
    <html lang="id">
    <head>
        <meta charset="UTF-8">
        <title>{title} - Sigramee</title>
        <style>
            /* Reset dasar */
            body {{ 
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
                margin: 0; 
                padding: 0; 
                background-color: #f4f6f9;
                overflow-x: hidden;
            }}
            h1, h2 {{ color: #2A4B7C; }} 
            
            /* --- SIDEBAR GALAXY BLUE --- */
            .sidebar-galaxy {{
                position: fixed;
                top: 0; left: 0; bottom: 0;
                width: 250px;
                padding: 15px; 
                box-sizing: border-box;
                overflow-y: auto;
                background-color: rgba(42, 75, 124, 0.95);
                backdrop-filter: blur(10px);
                -webkit-backdrop-filter: blur(10px);
                color: white;
                box-shadow: 2px 0 10px rgba(0,0,0,0.1);
                z-index: 1000;
                transition: transform 0.3s ease-in-out; /* Animasi CSS tetap ada agar halus saat load */
            }}
            
            .sidebar-galaxy::-webkit-scrollbar {{ width: 5px; }}
            .sidebar-galaxy::-webkit-scrollbar-thumb {{ background: rgba(255,255,255,0.3); border-radius: 10px; }}
            
            .sidebar-galaxy ul {{ list-style-type: none !important; padding-left: 0 !important; margin: 0 !important; }}
            .sidebar-galaxy li {{ margin-bottom: 2px; }} 
            
            .sidebar-galaxy h3, .sidebar-galaxy b, .sidebar-galaxy p b {{
                color: #ffffff; 
                border-bottom: 1px solid rgba(255,255,255,0.3);
                display: block; 
                margin-top: 25px; 
                margin-bottom: 8px; 
                padding-bottom: 5px;
                font-size: 1.0em; 
                font-weight: 600; 
                letter-spacing: 0.5px;
                text-transform: uppercase; 
                opacity: 0.9;
            }}
            
            .sidebar-galaxy a {{
                color: rgba(255, 255, 255, 0.85) !important; 
                text-decoration: none; 
                display: block;
                padding: 5px 10px; 
                font-size: 0.95em; 
                font-family: 'Segoe UI', sans-serif;
                border-radius: 4px; 
                transition: 0.2s;
                font-weight: 300 !important; 
                letter-spacing: 0.3px; 
            }}
            
            .sidebar-galaxy a:hover {{ 
                color: #fff !important; 
                background-color: rgba(255, 255, 255, 0.15); 
                padding-left: 15px; 
                opacity: 1; 
                font-weight: 400 !important; 
            }}

            /* --- KONTEN UTAMA --- */
            .main-content {{
                padding: 20px;
                box-sizing: border-box;
                min-height: 100vh;
                transition: margin-left 0.3s ease-in-out, width 0.3s ease-in-out;
            }}

            /* --- NAVBAR --- */
            .navbar {{ 
                background-color: transparent; 
                overflow: hidden; 
                padding: 15px 0; 
                margin-bottom: 20px;
                display: flex; 
                align-items: center;
                width: 100%;
            }}
            
            /* --- STYLE TOMBOL --- */
            .btn-auth {{ 
                background-color: #2A4B7C !important; 
                color: white !important;
                box-shadow: 0 4px 8px rgba(0,0,0,0.25) !important;
                padding: 8px 20px; 
                text-decoration: none; 
                font-weight: bold; 
                font-size: 0.9em;
                display: flex; 
                align-items: center; 
                gap: 8px;
                border: none;
                border-radius: 30px; 
                transition: all 0.3s ease;
                white-space: nowrap; 
                cursor: pointer;
            }}
            
            .btn-auth:hover {{ 
                background-color: #607d8b !important; 
                transform: translateY(-2px);
                box-shadow: 0 6px 12px rgba(0,0,0,0.3) !important;
            }}
            
            .push-right {{ margin-left: auto; }}
            
            .icon-arrow {{
                font-size: 1.1em;
                background: transparent; 
                color: inherit; 
            }}

            /* CSS Global Lainnya (Tabel, Form) */
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; background: white; }}
            th {{ 
                background-color: #e9ecef !important; color: #000000 !important;
                padding: 12px !important; border: 1px solid #ccc !important;
                text-align: center !important; font-weight: bold !important;
            }}
            td {{ border: 1px solid #ddd; padding: 10px; text-align: left; color: #333; }}
            tr:hover {{ background-color: #f8f9fa !important; }}
            
            input[type=text], input[type=password], input[type=number], input[type=date], select {{
                width: 100%; padding: 8px; margin: 5px 0; border: 1px solid #ccc; border-radius: 4px; box-sizing: border-box;
            }}
            form {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.05); }}
            
            .btn-blue {{ background-color: #007BFF; color: white; border:none; padding: 10px 15px; border-radius: 4px; cursor: pointer; }}
            .btn-red {{ background-color: #DC3545; color: white; border:none; padding: 10px 15px; border-radius: 4px; cursor: pointer; }}
        </style>
    </head>
    <body>
        
        {sidebar_html}

        <div class="main-content" style="{content_css}">
            
            <div class="navbar">
                {open_sidebar_btn}
                
                {f'<a href="/logout" class="btn-auth push-right">LOGOUT <span class="icon-arrow">‚ûú</span></a>' if 'username' in session else ''}
            </div>

            {error_html}
            <h1>{title}</h1>
            {body_content}
        </div>

    </body>
    </html>
    """
    return html
# =========================
# HALAMAN KONSUMEN
# =========================
def render_consumer_page(title, content_html, username, initials, display_phone, stats, alert_html, active_menu="Dashboard"):
    """
    Renders the complete consumer dashboard layout (sidebar + content).
    Fix: Icon Setting terpotong (overflow & display issues solved).
    """
    
    # Stats data
    total_trx = stats['total'] or 0
    waiting = stats['waiting'] or 0
    success = stats['success'] or 0
    failed = stats['failed'] or 0
    
    body = f"""
    <!DOCTYPE html>
    <html lang="id">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>{title} | Ivalia Consumer</title>
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700&display=swap');

            /* --- BACKGROUND GRADIENT --- */
            body, html {{
                margin: 0; padding: 0;
                height: 100%;
                font-family: 'Poppins', sans-serif;
                background: linear-gradient(135deg, #408080 0%, #D7F7FF 100%);
                color: white;
                overflow-x: hidden;
            }}

            h2 {{ color: #1a1a1a; }}
            
            .dashboard-container {{ display: flex; min-height: 100vh; }}
            
            /* --- SIDEBAR STYLE --- */
            .sidebar {{
                width: 280px;
                padding: 40px 0;
                display: flex; flex-direction: column; gap: 30px;
                background: transparent; border: none;
                align-items: flex-start;
            }}
            .logo-section {{ margin-left: 30px; margin-bottom: 50px; }}
            .logo-text {{
                font-family: 'Poppins', sans-serif;
                font-size: 2.5em;
                font-weight: 700;
                color: white;
                line-height: 0.8;
                text-shadow: 0 2px 5px rgba(0,0,0,0.2);
            }}

            .menu-item {{
                display: flex; align-items: center; gap: 15px; text-decoration: none;
                color: white; font-size: 1.1em; font-weight: 500; opacity: 0.9; 
                transition: 0.3s; padding: 10px 30px;
                border-radius: 0 25px 25px 0;
                width: calc(100% - 30px);
            }}
            .menu-item:hover {{ opacity: 1; transform: translateX(5px); }}
            .menu-icon {{ width: 28px; height: 28px; fill: none; stroke: currentColor; stroke-width: 2.5; stroke-linecap: round; stroke-linejoin: round; }}

            .menu-item.active {{
                background: rgba(108, 194, 183, 0.4); 
                font-weight: 600;
                opacity: 1;
                box-shadow: 0 4px 15px rgba(0,0,0,0.05);
                border-left: 5px solid white;
            }}
            .menu-item.active .menu-icon {{ stroke-width: 3; filter: drop-shadow(0 0 5px rgba(255,255,255,0.6)); }}

            /* --- CONTENT AREA --- */
            .content-area {{ 
                flex: 1; padding: 40px 50px; 
                background-color: transparent; 
                color: #333;
            }}
            
            /* --- PROFILE CARD (FIX ICON) --- */
            .profile-card {{
                background: rgba(255, 255, 255, 0.15); 
                border-radius: 25px; 
                padding: 30px;
                display: flex; 
                flex-direction: column; 
                box-shadow: 0 4px 15px rgba(0,0,0,0.05);
                margin-bottom: 40px; 
                max-width: 850px;
                color: white;
                position: relative;
                border: 1px solid rgba(255, 255, 255, 0.3);
                /* HAPUS OVERFLOW HIDDEN AGAR ICON TIDAK TERPOTONG */
                /* overflow: hidden; */ 
            }}
            .profile-content {{ display: flex; justify-content: space-between; align-items: center; width: 100%; }}
            .profile-info {{ display: flex; align-items: center; gap: 15px; }}
            .avatar-circle {{ 
                background-color: #81C4E7; 
                width: 60px; height: 60px; 
                border-radius: 50%; 
                display: flex; justify-content: center; align-items: center; 
                font-weight: bold; font-size: 1.4em; color: white;
                border: 2px solid rgba(255,255,255,0.3);
            }}
            .badge-member {{ 
                background: #4C7F7F; 
                color: white; 
                padding: 4px 15px; 
                border-radius: 8px; 
                font-size: 0.8em; 
                display: inline-block; 
                margin-top: 5px;
            }}
            .divider-line {{ 
                background-color: rgba(255,255,255,0.3); 
                height: 1px; width: 100%; margin: 20px 0;
            }}
            .profile-phone {{ display: flex; align-items: center; gap: 10px; font-size: 1.1em; }}
            .phone-icon {{ width: 22px; height: 22px; fill:none; stroke:currentColor; stroke-width:1.5; }}
            
            /* --- SETTINGS BUTTON FIX --- */
            .settings-btn {{ 
                color: white; 
                transition: 0.2s; 
                /* Tambahan agar icon center dan tidak terpotong line-height */
                display: flex; 
                align-items: center; 
                justify-content: center;
                padding: 5px; 
            }}
            .settings-btn:hover {{ transform: rotate(30deg); opacity: 0.8; }}

            /* --- STATS GRID LAYOUT --- */
            .stats-label {{
                color: white; 
                font-size: 1.5em;
                font-weight: 600;
                margin-bottom: 20px;
            }}
            
            .stats-layout {{
                display: grid;
                grid-template-columns: 1fr 1.2fr; 
                gap: 20px;
                max-width: 850px;
            }}
            
            /* Kotak Total */
            .total-container {{
                background: rgba(255, 255, 255, 0.15); 
                border-radius: 15px;
                display: flex;
                flex-direction: column;
                justify-content: center;
                align-items: center;
                text-decoration: none;
                color: white;
                box-shadow: 0 4px 15px rgba(0,0,0,0.05);
                transition: transform 0.2s;
                height: 260px; 
                border: 1px solid rgba(255, 255, 255, 0.3);
            }}
            .total-container:hover {{ transform: translateY(-3px); }}
            
            /* Stack Kanan */
            .status-stack {{
                display: flex;
                flex-direction: column;
                justify-content: space-between;
                gap: 15px;
                height: 260px; 
            }}
            
            .stat-box-row {{
                flex: 1; 
                border-radius: 15px;
                display: flex;
                flex-direction: column;
                justify-content: center;
                align-items: center;
                text-decoration: none;
                color: white;
                box-shadow: 0 4px 10px rgba(0,0,0,0.1);
                transition: transform 0.2s;
                position: relative;
            }}
            .stat-box-row:hover {{ transform: translateX(5px); }}

            /* Warna Status (Soft) */
            .bg-yellow {{ background-color: #FFCA28; color: white; }}
            .bg-green {{ background-color: #66BB6A; color: white; }}
            .bg-red {{ background-color: #EF5350; color: white; }}

            .stat-number {{ font-size: 3.5em; font-weight: 600; line-height: 1; margin-bottom: 5px; }}
            .stat-text {{ font-size: 1em; font-weight: 500; opacity: 0.95; }}
            
            .row-number {{ font-size: 1.8em; font-weight: 700; line-height: 1; }}
            .row-text {{ font-size: 0.85em; font-weight: 600; }}

            /* --- RESPONSIVE --- */
            @media (max-width: 768px) {{
                .sidebar {{ width: 100%; padding: 20px; align-items: center; }}
                .dashboard-container {{ flex-direction: column; }}
                .profile-card, .stats-layout {{ max-width: 100%; }}
                .stats-layout {{ grid-template-columns: 1fr; }} 
                .total-container {{ height: 150px; }}
                .status-stack {{ height: auto; }}
                .stat-box-row {{ padding: 20px; margin-bottom: 10px; }}
            }}
            
            /* CSS Tambahan */
            .status-badge-trx {{ padding: 5px 15px; border-radius: 20px; font-weight: 500; font-size: 0.9em; display: inline-flex; align-items: center; justify-content: center; color: white; text-align: center; height: 30px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }}
            .status-badge-trx.bg-sukses {{ background-color: #A3D89D; white-space: nowrap; }}
            .status-badge-trx.bg-pending {{ background-color: #FFC999; color: #A0522D; white-space: nowrap; }}
            .status-badge-trx.bg-gagal {{ background-color: #D46A6A; white-space: nowrap; }}
        </style>
    </head>
    <body>

        <div class="dashboard-container">
            <div class="sidebar">
                <div class="logo-section">
                    <div class="logo-text"></div>
                </div>

                <a href="/home" class="menu-item {'active' if active_menu == 'Dashboard' else ''}">
                    <svg class="menu-icon" viewBox="0 0 24 24"><path d="M3 9l9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"></path><polyline points="9 22 9 12 15 12 15 22"></polyline></svg>
                    <span class="menu-text">Dashboard</span>
                </a>

                <a href="/consumer/transactions?status=all" class="menu-item {'active' if active_menu == 'Transaksi' else ''}">
                    <svg class="menu-icon" viewBox="0 0 24 24"><circle cx="9" cy="21" r="1"></circle><circle cx="20" cy="21" r="1"></circle><path d="M1 1h4l2.68 13.39a2 2 0 0 0 2 1.61h9.72a2 2 0 0 0 2-1.61L23 6H6"></path></svg>
                    <span class="menu-text">Transaksi</span>
                </a>

                <a href="/shop" class="menu-item {'active' if active_menu == 'Belanja' else ''}">
                    <svg class="menu-icon" viewBox="0 0 24 24"><path d="M6 2L3 6v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2V6l-3-4z"></path><line x1="3" y1="6" x2="21" y2="6"></line><path d="M16 10a4 4 0 0 1-8 0"></path></svg>
                    <span class="menu-text">Belanja!</span>
                </a>

                <a href="/logout" class="menu-item" style="margin-top: auto; opacity: 0.7;">
                    <svg class="menu-icon" viewBox="0 0 24 24"><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"></path><polyline points="16 17 21 12 21 12 16 7"></polyline><line x1="21" y1="12" x2="9" y2="12"></line></svg>
                    <span class="menu-text">Logout</span>
                </a>
            </div>

            <div class="content-area">
                {alert_html}
                
                {' ' if content_html and 'profile-card' not in content_html else ''}
                
                {content_html if 'profile-card' not in content_html else f"""
                
                <div class="profile-card">
                    <div class="profile-content">
                        <div class="profile-info">
                            <div class="avatar-circle">{initials}</div>
                            <div class="name-section">
                                <h2 style="color: white; margin: 0 0 5px 0; font-size: 1.4em;">{username}</h2>
                                <div class="badge-member">MEMBER</div>
                            </div>
                        </div>
                        <a href="/consumer/profile-settings" class="settings-btn">
                            <svg style="width:28px; height:28px; fill:none; stroke:currentColor; stroke-width:2; overflow: visible;" viewBox="0 0 24 24">
                                <circle cx="12" cy="12" r="3"></circle>
                                <path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1 0 2.83 2 2 0 0 1-2.83 0l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-2 2 2 2 0 0 1-2-2v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83 0 2 2 0 0 1 0-2.83l.06-.06a1.65 1.65 0 0 0 .33-1.82 1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1-2-2 2 2 0 0 1 2-2h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 0-2.83 2 2 0 0 1 2.83 0l.06.06a1.65 1.65 0 0 0 1.82.33H9a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 2-2 2 2 0 0 1 2 2v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 0 2 2 0 0 1 0 2.83l.06.06a1.65 1.65 0 0 0-.33 1.82V9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 2 2 2 2 0 0 1-2 2h-.09a1.65 1.65 0 0 0-1.51 1z"></path>
                            </svg>
                        </a>
                    </div>
                    <div class="divider-line"></div>
                    <div class="profile-phone">
                        <svg class="phone-icon" viewBox="0 0 24 24">
                            <path d="M22 16.92v3a2 2 0 0 1-2.18 2 19.79 19.79 0 0 1-8.63-3.07 19.5 19.5 0 0 1-6-6 19.79 19.79 0 0 1-3.07-8.67A2 2 0 0 1 4.11 2h3a2 2 0 0 1 2 1.72 12.84 12.84 0 0 0 .7 2.81 2 2 0 0 1-.45 2.11L8.09 9.91a16 16 0 0 0 6 6l1.27-1.27a2 2 0 0 1 2.11-.45 12.84 12.84 0 0 0 2.81.7A2 2 0 0 1 22 16.92z"></path>
                        </svg> 
                        {display_phone}
                    </div>
                </div>

                <div class="stats-label">Transaksi Hari Ini</div>
                
                <div class="stats-layout">
                    <a href="/consumer/transactions?status=all" class="total-container">
                        <div class="stat-number">{total_trx}</div>
                        <div class="stat-text">Total Transaksi</div>
                    </a>

                    <div class="status-stack">
                        <a href="/consumer/transactions?status=pending" class="stat-box-row bg-yellow">
                            <div class="row-number">{waiting}</div>
                            <div class="row-text">Menunggu Konfirmasi</div>
                        </a>
                        
                        <a href="/consumer/transactions?status=verified" class="stat-box-row bg-green">
                            <div class="row-number">{success}</div>
                            <div class="row-text">Sukses</div>
                        </a>
                        
                        <a href="/consumer/transactions?status=rejected" class="stat-box-row bg-red">
                            <div class="row-number">{failed}</div>
                            <div class="row-text">Gagal</div>
                        </a>
                    </div>
                </div>
                """}
            </div>
        </div>

    </body>
    </html>
    """
    return render_template_string(body)

# ==========================================
# 1. ROUTE HALAMAN UTAMA (LANDING PAGE)
# ==========================================
@app.route("/")
def index():
    """
    Halaman Landing Page (One Page Scroll).
    Update: 
    1. Hapus "Accounting Suite".
    2. Hapus kurung pada SIGRAMEE.
    3. Tambah animasi gerak (Shine Effect) pada teks SIGRAMEE.
    """
    if 'username' in session:
        if session.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))
        elif session.get('role') == 'consumer':
            return redirect(url_for('consumer_home'))
            
    # Cek pesan sukses atau error dari query URL
    success_message = request.args.get('success')
    error_message = request.args.get('error')
    
    # --- HTML NOTIFIKASI CUSTOM ---
    alert_html = ""
    
    if error_message:
        alert_html = f"""
        <div class="custom-alert error-alert" id="autoCloseAlert">
            <div class="alert-icon-box">
                <svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="white" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                    <path d="M10.29 3.86L1.82 18a2 2 0 0 0 1.71 3h16.94a2 2 0 0 0 1.71-3L13.71 3.86a2 2 0 0 0-3.42 0z"></path>
                    <line x1="12" y1="9" x2="12" y2="13"></line>
                    <line x1="12" y1="17" x2="12.01" y2="17"></line>
                </svg>
            </div>
            <div class="alert-text-box">
                {error_message}
            </div>
        </div>
        """
    elif success_message:
        alert_html = f"""
        <div class="custom-alert success-alert" id="autoCloseAlert">
            <div class="alert-icon-box">
                <svg width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="white" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"></polyline></svg>
            </div>
            <div class="alert-text-box">
                {success_message}
            </div>
        </div>
        """

    html = f"""
    <!DOCTYPE html>
    <html lang="id">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>SIGRAMEE - Accounting Suite</title>
        <style>
            /* --- CSS UNTUK ALERT POPUP --- */
            .custom-alert {{
                position: fixed;
                top: 30px; left: 50%; transform: translateX(-50%);
                z-index: 10000; display: flex; align-items: stretch;
                box-shadow: 0 5px 15px rgba(0,0,0,0.2); border-radius: 4px;
                overflow: hidden; animation: slideDown 0.5s cubic-bezier(0.68, -0.55, 0.27, 1.55);
                min-width: 350px; max-width: 90%;
            }}
            @keyframes slideDown {{ 0% {{ top: -100px; opacity: 0; }} 100% {{ top: 30px; opacity: 1; }} }}
            .error-alert .alert-icon-box {{ background-color: #D34545; }}
            .error-alert .alert-text-box {{ background-color: #FF7474; }}
            .success-alert .alert-icon-box {{ background-color: #2E7D32; }}
            .success-alert .alert-text-box {{ background-color: #4CAF50; }}
            .alert-icon-box {{ padding: 12px 15px; display: flex; align-items: center; justify-content: center; }}
            .alert-text-box {{ padding: 12px 20px; color: white; font-family: 'Segoe UI', sans-serif; font-weight: 700; font-size: 1rem; flex-grow: 1; display: flex; align-items: center; }}

            /* --- GLOBAL RESET --- */
            body, html {{
                margin: 0; padding: 0;
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                scroll-behavior: smooth;
                overflow-x: hidden;
                background-color: #1a1a1a;
            }}

            /* --- NAVBAR --- */
            .navbar {{
                position: fixed; top: 0; left: 0; width: 100%; padding: 30px 50px;
                display: flex; justify-content: space-between; align-items: center;
                z-index: 9999; box-sizing: border-box; 
                background: transparent !important; 
                transition: padding 0.4s ease; 
            }}
            .navbar.scrolled {{ padding: 15px 50px; }}
            
            .brand {{ color: white; font-size: 1.2em; font-weight: bold; letter-spacing: 2px; text-transform: uppercase; text-shadow: 0 2px 10px rgba(0,0,0,0.5); }}
            .nav-links {{ display: flex; gap: 40px; }}
            .nav-links a {{
                color: rgba(255,255,255,0.8); text-decoration: none; font-weight: 500; 
                font-size: 0.95em; transition: color 0.3s ease; position: relative; 
                text-transform: uppercase; letter-spacing: 1px; padding-bottom: 5px;
                text-shadow: 0 2px 4px rgba(0,0,0,0.5);
            }}
            .nav-links a::after {{
                content: ''; position: absolute; width: 0; height: 2px; bottom: 0; left: 0;
                background-color: white; transition: width 0.3s ease-in-out;
                box-shadow: 0 2px 4px rgba(0,0,0,0.5);
            }}
            .nav-links a:hover {{ color: white; }}
            .nav-links a.active {{ color: white; font-weight: 700; }}
            .nav-links a:hover::after, .nav-links a.active::after {{ width: 100%; }}

            .btn-login {{
                background: transparent; color: white; border: 2px solid white; padding: 8px 25px; border-radius: 30px;
                text-decoration: none; font-weight: bold; display: flex; align-items: center; gap: 10px; transition: 0.3s;
                text-shadow: 0 2px 5px rgba(0,0,0,0.5); box-shadow: 0 2px 5px rgba(0,0,0,0.3);
            }}
            .btn-login:hover {{ background: white; color: #1a1a1a; text-shadow: none; }}

            /* --- SECTION GLOBAL --- */
            section {{
                min-height: 100vh; width: 100%; position: relative; overflow: hidden;
                background-size: cover; background-position: center; background-attachment: fixed; 
                display: flex; align-items: center; justify-content: center;
            }}
            .overlay {{
                position: absolute; top: 0; left: 0; width: 100%; height: 100%;
                background: rgba(0,0,0,0.25); z-index: 1;
            }}
            .content {{ position: relative; z-index: 2; width: 85%; margin: 0 auto; }}

            /* --- ANIMASI REVEAL --- */
            .reveal {{ opacity: 0; transition: all 1.2s cubic-bezier(0.25, 0.46, 0.45, 0.94); }}
            .reveal.active {{ opacity: 1; transform: translate(0, 0) !important; }}
            .from-left {{ transform: translateX(-100px); }}
            .from-right {{ transform: translateX(100px); }}
            .from-top {{ transform: translateY(-100px); }}
            .from-bottom {{ transform: translateY(100px); }}

            /* --- HOME SECTION A (ANIMASI TEKS) --- */
            #home-a {{ background-image: url('/static/landing_bg.jpeg'); }}
            #home-a .content {{ display: flex; flex-direction: column; align-items: flex-end; }}
            .home-text-wrapper {{ text-align: right; max-width: 900px; }} /* Rata kanan */

            /* ANIMASI TEKS BERGERAK (SHINE EFFECT) */
            .text-reveal-main {{ 
                font-size: 9vw; /* Ukuran diperbesar */
                font-weight: 900; 
                line-height: 0.9; 
                text-transform: uppercase; 
                letter-spacing: -2px;
                
                /* Gradient untuk efek mengkilap/gerak */
                background: linear-gradient(
                    to right, 
                    #ffffff 20%, 
                    #a4c639 50%, 
                    #ffffff 80%
                );
                background-size: 200% auto;
                
                color: #fff; /* Fallback */
                background-clip: text;
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
                
                animation: shine 3s linear infinite;
            }}

            @keyframes shine {{
                to {{
                    background-position: 200% center;
                }}
            }}

            /* HOME B */
            #home-b {{ background-image: url('/static/home_slide3.jpeg'); }}
            #home-b .content {{ text-align: center; }}
            .village-text {{ font-size: 5vw; font-weight: 700; color: white; line-height: 1; text-shadow: 0 10px 30px rgba(0,0,0,0.8); }}

            /* ABOUT */
            #about {{ background-color: #35504B; background-image: none; padding: 80px 50px; }}
            .about-grid-exact {{ display: grid; grid-template-columns: 1fr 1.8fr 1fr; gap: 40px; max-width: 1300px; width: 100%; position: relative; z-index: 2; color: white; align-items: stretch; }}
            .about-left {{ display: flex; flex-direction: column; justify-content: space-between; height: 100%; }}
            .about-title-huge {{ font-size: 4.5em; line-height: 0.9; margin: 0 0 20px 0; text-transform: uppercase; font-weight: 800; }}
            .card-base {{ background-color: #263634; background-size: cover; background-position: center; border-radius: 20px; position: relative; overflow: hidden; box-shadow: 0 15px 30px rgba(0,0,0,0.2); }}
            .badge-pill {{ position: absolute; top: 50%; left: 50%; transform: translate(-50%, -50%); background-color: rgba(20, 30, 20, 0.9); color: white; padding: 10px 25px; border-radius: 30px; font-weight: bold; font-size: 0.95em; border: 1px solid #555; text-transform: capitalize; white-space: nowrap; z-index: 10; }}
            .img-tall {{ height: 380px; margin-top: auto; }} 
            .img-medium {{ height: 350px; width: 100%; }}
            .img-small {{ height: 250px; width: 100%; }}
            .phone-link {{ margin-top: 20px; font-size: 1.1em; font-weight: 500; display: flex; align-items: center; gap: 10px; color: white; text-decoration: none; transition: transform 0.3s, color 0.3s; width: fit-content; }}
            .phone-link:hover {{ color: #ff6b6b; transform: translateX(5px); }}
            .about-center {{ display: flex; flex-direction: column; justify-content: flex-start; gap: 30px; padding-top: 0; }}
            .about-desc-text {{ font-size: 1em; line-height: 1.6; text-align: justify; color: #e0e0e0; margin: 0; }}
            .about-right {{ display: flex; flex-direction: column; padding-top: 0; gap: 30px; }}
            .card-map-container {{ position: relative; border-radius: 20px; overflow: hidden; box-shadow: 0 15px 30px rgba(0,0,0,0.2); background-color: #263634; min-height: 180px; flex-grow: 1; display: flex; flex-direction: column; }}
            .card-map-container iframe {{ width: 100%; height: 100%; border: 0; position: absolute; top: 0; left: 0; filter: opacity(0.9); }}
            .map-overlay-text {{ position: absolute; bottom: 0; left: 0; width: 100%; background: rgba(38, 54, 52, 0.95); color: white; padding: 15px 20px; text-decoration: none; display: flex; align-items: center; gap: 15px; box-sizing: border-box; border-top: 1px solid rgba(255,255,255,0.1); transition: background 0.3s; z-index: 5; }}
            .map-overlay-text:hover {{ background: #263634; }}
            .map-icon {{ font-size: 1.8em; color: #ff6b6b; }}
            .map-detail p {{ margin: 0; font-size: 0.9em; line-height: 1.3; font-weight: 500; }}

            /* LAPORAN */
            #laporan {{ background-image: url('/static/home_slide1.jpeg'); padding: 50px; }}
            .laporan-grid {{ display: grid; grid-template-columns: 0.7fr 1.3fr; gap: 60px; max-width: 1400px; width: 100%; position: relative; z-index: 2; align-items: center; }}
            .laporan-left {{ position: relative; padding: 20px; }}
            .doc-img {{ width: 90%; border-radius: 5px; box-shadow: 0 20px 50px rgba(0,0,0,0.6); transform: rotate(-2deg); background: white; }}
            .download-card {{ position: absolute; bottom: 0; right: 0; background-color: #051008; padding: 25px 30px; border-radius: 15px; border: 1px solid #333; box-shadow: 0 10px 40px rgba(0,0,0,0.8); min-width: 220px; z-index: 3; }}
            .download-card p {{ margin: 0 0 10px 0; color: white; font-size: 1.2em; font-weight: 700; line-height: 1.2; }}
            .download-card a {{ color: white; font-weight: 400; text-decoration: underline; font-size: 1.1em; }}
            .laporan-right {{ text-align: left; color: white; }}
            .title-biz {{ font-size: 3.5vw; font-weight: 300; line-height: 1; text-transform: capitalize; margin: 0; letter-spacing: -1px; }}
            .title-sig {{ font-size: 6.5vw; font-weight: 900; line-height: 0.8; text-transform: uppercase; color: #a4c639; margin: 0 0 50px 0; letter-spacing: -2px; }}
            .biz-content-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 40px; align-items: start; }}
            .biz-sub-1 p {{ font-size: 1.4em; line-height: 1.2; font-weight: 500; margin-top: 0; margin-bottom: 20px; }}
            .biz-hr {{ width: 100%; height: 1px; background: #555; margin-bottom: 20px; }}
            .activity-img {{ width: 100%; height: 180px; object-fit: cover; border-radius: 15px; box-shadow: 0 5px 15px rgba(0,0,0,0.3); }}
            .biz-sub-2 p.lead {{ font-size: 1.4em; line-height: 1.2; font-weight: 500; margin-top: 0; margin-bottom: 20px; }}
            .green-dash {{ width: 40px; height: 6px; background-color: #a4c639; border-radius: 4px; margin-bottom: 20px; }}
            .biz-sub-2 p.desc {{ font-size: 1em; line-height: 1.6; color: #ccc; text-align: justify; }}

            /* GALLERY A */
            #gallery-a {{ background-image: url('/static/gallery_bg.jpeg'); }}
            #gallery-a .overlay {{ background: rgba(20, 40, 30, 0.85); }}
            .gallery-collage {{ display: grid; grid-template-columns: 1fr 1fr; gap: 50px; width: 100%; max-width: 1300px; position: relative; z-index: 2; align-items: center; }}
            .collage-left {{ position: relative; height: 600px; }}
            .gal-title-layer {{ position: absolute; top: 0; left: 0; font-size: 4em; font-weight: 300; line-height: 0.8; color: white; z-index: 5; text-transform: uppercase; }}
            .gal-title-layer span {{ display: block; font-size: 2.5em; font-weight: 800; color: rgba(255,255,255,0.5); -webkit-text-stroke: 2px white; }}
            .photo-stack-1 {{ position: absolute; bottom: 0; left: 0; width: 250px; height: 250px; background-image: url('/static/gallery_img1.jpg'); background-size: cover; border-radius: 20px; box-shadow: 0 10px 30px rgba(0,0,0,0.5); z-index: 3; }}
            .photo-stack-2 {{ position: absolute; top: 150px; left: 180px; width: 200px; height: 350px; background-image: url('/static/gallery_img2.jpeg'); background-size: cover; border-radius: 20px; transform: rotate(-5deg); border: 5px solid white; z-index: 4; }}
            .photo-stack-3 {{ position: absolute; bottom: -50px; right: 50px; width: 220px; height: 220px; background-image: url('/static/gallery_img3.png'); background-size: cover; border-radius: 20px; z-index: 2; }}
            .collage-right {{ display: flex; flex-direction: column; gap: 50px; }}
            .photo-wide {{ width: 100%; height: 300px; background-image: url('/static/gallery_img4.png'); background-size: cover; border-radius: 30px; box-shadow: 0 20px 50px rgba(0,0,0,0.4); }}
            .collage-desc {{ color: white; font-size: 3em; font-weight: 600; line-height: 1; letter-spacing: -1px; position: relative; padding-left: 40px; }}
            .collage-desc::before {{ content: ''; position: absolute; left: 0; top: 15px; width: 20px; height: 20px; background-color: #7daea6; border-radius: 50%; }}
            .collage-hr {{ width: 100%; height: 1px; background: rgba(255,255,255,0.3); margin-top: 20px; }}

            /* GALLERY B */
            #gallery-b {{ background-image: url('/static/home_slide2.jpeg'); }}
            #gallery-b .overlay {{ background: rgba(0,0,0,0.3); }}

            .clarity-container {{
                position: relative; z-index: 2; width: 85%; height: 80vh;
                display: flex; flex-direction: column; justify-content: center;
            }}
            .clarity-title {{ color: white; font-size: 6.5vw; font-weight: 600; line-height: 0.9; letter-spacing: -2px; }}
            .thin-line {{ width: 100%; height: 1px; background-color: rgba(255,255,255,0.4); margin: 30px 0; }}
            .plus-sign {{ position: absolute; right: 0; top: 60%; font-size: 3em; color: white; font-weight: 300; }}
            .arrow-green-btn {{ position: absolute; bottom: 0; left: 0; width: 60px; height: 30px; border: 2px solid #a4c639; border-radius: 20px; display: flex; align-items: center; justify-content: center; color: #a4c639; font-size: 1.5em; cursor: pointer; }}
            .meet-team {{ position: absolute; bottom: 0; right: 0; text-align: left; color: white; }}
            .meet-team h4 {{ font-size: 1.5em; font-weight: 600; margin: 0 0 10px 0; }}
            .meet-team ul {{ list-style: none; padding: 0; margin: 0; font-size: 1em; line-height: 1.4; font-weight: 500; }}
            .meet-team li::before {{ content: "‚Ä¢ "; margin-right: 5px; }}

        </style>
    </head>
    <body>

        {alert_html}

        <nav class="navbar" id="navbar">
            <div class="brand">WELCOME</div>
            <div class="nav-links">
                <a href="#home-a" id="link-home" class="active">Home</a>
                <a href="#about" id="link-about">About</a>
                <a href="#laporan" id="link-laporan">Laporan Keuangan</a>
                <a href="#gallery-a" id="link-gallery">Gallery</a>
            </div>
            <a href="/login-page" class="btn-login">LOGIN ‚ûú</a>
        </nav>

        <section id="home-a">
            <div class="overlay"></div>
            <div class="content">
                <div class="home-text-wrapper reveal from-left active"> 
                    <div class="text-reveal-main">SIGRAMEE</div>
                </div>
            </div>
        </section>

        <section id="home-b">
            <div class="overlay"></div>
            <div class="content">
                <div class="village-text reveal from-bottom">
                    SIGRAMEE keeps<br>your records calm<br>and clear ‚Äî village<br>style.
                </div>
                <div style="margin-top: 20px; font-size: 1em; letter-spacing: 5px; color:white;" class="reveal from-bottom" style="transition-delay: 0.2s;">
                    SIGRAMEE
                </div>
            </div>
        </section>
        
        <section id="about">
            <div class="about-grid-exact">
                <div class="about-left">
                    <h2 class="about-title-huge reveal from-left">ABOUT<br>SIGRAMEE</h2>
                    <div class="card-base img-tall reveal from-bottom" style="background-image: url('/static/about_since2019.png'); transition-delay: 0.2s;">
                        <div class="badge-pill">Since 2020</div>
                    </div>
                    <a href="tel:081328078581" class="phone-link reveal from-left" style="transition-delay: 0.4s;">
                        <span style="color: #ff6b6b;">üìû</span> 0813 2807 8581
                    </a>
                </div>
                <div class="about-center">
                    <div class="card-base img-medium reveal from-top" style="background-image: url('/static/about_premium.png'); transition-delay: 0.1s;">
                        <div class="badge-pill">Premium</div>
                    </div>
                    <p class="about-desc-text reveal from-bottom" style="transition-delay: 0.3s;">
                        SIGRAMEE merupakan sistem pencatatan keuangan yang mendukung usaha budidaya gurame, 
                        mulai dari pengelolaan kolam, pemberian pakan, hingga proses panen.
                        Budidaya gurame dilakukan secara teratur dengan pengawasan kualitas air, 
                        pakan pelet dan daun-daunan, serta perawatan kolam yang rutin.
                    </p>
                </div>
                <div class="about-right">
                    <div class="card-map-container reveal from-right" style="transition-delay: 0.3s;">
                        <iframe src="https://maps.google.com/maps?q=Desa+Kertodeso,+Kec.+Mirit,+Kab.+Kebumen&t=&z=13&ie=UTF8&iwloc=&output=embed" frameborder="0" scrolling="no" marginheight="0" marginwidth="0"></iframe>
                        <a href="https://maps.google.com/maps?q=Desa+Kertodeso,+Kec.+Mirit,+Kab.+Kebumen" target="_blank" class="map-overlay-text">
                            <div class="map-icon">üìç</div>
                            <div class="map-detail"><p>Desa Kertodeso, Kec.<br>Mirit, Kab. Kebumen</p></div>
                        </a>
                    </div>
                    <div class="card-base img-small reveal from-bottom" style="background-image: url('/static/about_wellsafe.png'); transition-delay: 0.5s;">
                        <div class="badge-pill">Wellsafe</div>
                    </div>
                </div>
            </div>
        </section>

        <section id="laporan">
            <div class="overlay"></div>
            <div class="laporan-grid">
                <div class="laporan-left reveal from-left">
                    <img src="/static/report_doc.jpg" alt="Document" class="doc-img">
                    <div class="download-card reveal from-bottom" style="transition-delay: 0.3s;">
                        <p>Unduh laporan<br>terbaru:</p>
                        <a href="/download-income-statement">Download Here.</a>
                    </div>
                </div>
                <div class="laporan-right">
                    <h2 class="title-biz reveal from-top">Business Overview</h2>
                    <h1 class="title-sig reveal from-right" style="transition-delay: 0.1s;">SIGRAMEE</h1>
                    <div class="biz-content-grid">
                        <div class="biz-sub-1 reveal from-bottom" style="transition-delay: 0.2s;">
                            <p>Lihat perkembangan keuangan usaha secara langsung melalui dashboard.</p>
                            <div class="biz-hr"></div>
                            <img src="/static/report_activity.png" alt="Activity" class="activity-img">
                        </div>
                        <div class="biz-sub-2 reveal from-bottom" style="transition-delay: 0.4s;">
                            <p class="lead">Data tersimpan rapi, mudah ditelusuri, dan nyaman untuk dianalisis.</p>
                            <div class="green-dash"></div>
                            <p class="desc">
                                SIGRAMEE menyediakan ringkasan keuangan yang dapat diunduh kapan saja. 
                                Pelanggan dapat memantau perkembangan usaha dan melihat riwayat aktivitas 
                                secara jelas untuk memastikan transparansi dan kemudahan dalam mengawasi kondisi finansial.
                            </p>
                        </div>
                    </div>
                </div>
            </div>
        </section>

        <section id="gallery-a">
            <div class="overlay"></div>
            <div class="gallery-collage">
                <div class="collage-left">
                    <div class="gal-title-layer reveal from-left">
                        Gallery<br><span>SIGRAMEE</span>
                    </div>
                    <div class="photo-stack-1 reveal from-bottom" style="transition-delay: 0.2s;"></div>
                    <div class="photo-stack-2 reveal from-top" style="transition-delay: 0.4s;"></div>
                    <div class="photo-stack-3 reveal from-right" style="transition-delay: 0.6s;"></div>
                </div>
                <div class="collage-right">
                    <div class="photo-wide reveal from-right" style="transition-delay: 0.3s;"></div>
                    <div class="collage-desc reveal from-bottom" style="transition-delay: 0.5s;">
                        A simple glimpse<br>into life around the<br>ponds.
                        <div class="collage-hr"></div>
                    </div>
                </div>
            </div>
        </section>

        <section id="gallery-b">
            <div class="overlay"></div>
            
            <div class="clarity-container reveal from-bottom">
                <div class="clarity-title">
                    Let‚Äôs create clarity and<br>
                    progress ‚Äî in every<br>
                    process.
                </div>
                
                <div class="thin-line"></div>
                
                <div class="plus-sign">+</div>
                
                <div class="arrow-green-btn">
                    ‚ûú
                </div>
                
                <div class="meet-team">
                    <h4>Meet our team:</h4>
                    <ul>
                        <li>Arvanessa Fauzya Sirin</li>
                        <li>Sahda Ardelia Artanti</li>
                        <li>Syifaurrahmah</li>
                    </ul>
                </div>
            </div>
        </section>

        <script>
            // Auto Close Alert Script
            document.addEventListener('DOMContentLoaded', function() {{
                const alertBox = document.getElementById('autoCloseAlert');
                if (alertBox) {{
                    setTimeout(function() {{
                        alertBox.style.opacity = '0';
                        setTimeout(function() {{
                            alertBox.remove();
                        }}, 500);
                    }}, 3000); // Hilang setelah 3 detik
                }}
            }});
            
            // Navbar Background Change on Scroll
            window.addEventListener('scroll', () => {{
                const navbar = document.getElementById('navbar');
                if (window.scrollY > 50) {{
                    navbar.classList.add('scrolled');
                }} else {{
                    navbar.classList.remove('scrolled');
                }}
            }});

            // Intersection Observer untuk Animasi Reveal Elemen
            const observerOptions = {{
                threshold: 0.1,
                rootMargin: "0px 0px -15% 0px" 
            }};

            const revealObserver = new IntersectionObserver((entries) => {{
                entries.forEach(entry => {{
                    if (entry.isIntersecting) {{
                        entry.target.classList.add('active');
                    }}
                }});
            }}, observerOptions);

            document.querySelectorAll('.reveal').forEach(el => revealObserver.observe(el));

            // LOGIKA UPDATE NAVIGASI (Underline Animation)
            const sections = document.querySelectorAll('section');
            const navLinks = document.querySelectorAll('.nav-links a');

            window.addEventListener('scroll', () => {{
                let current = '';
                
                sections.forEach(section => {{
                    const sectionTop = section.offsetTop;
                    const sectionHeight = section.clientHeight;
                    if (pageYOffset >= (sectionTop - sectionHeight / 3)) {{
                        current = section.getAttribute('id');
                    }}
                }});

                navLinks.forEach(a => {{
                    a.classList.remove('active');
                    
                    if (current === 'home-a' || current === 'home-b') {{
                        if (a.getAttribute('href') === '#home-a') a.classList.add('active');
                    }} 
                    else if (current === 'gallery-a' || current === 'gallery-b') {{
                        if (a.getAttribute('href') === '#gallery-a') a.classList.add('active');
                    }}
                    else if (a.getAttribute('href') === '#' + current) {{
                        a.classList.add('active');
                    }}
                }});
            }});
        </script>
    </body>
    </html>
    """
    return render_template_string(html)

# ==========================================
# 2. ROUTE HALAMAN LOGIN (YANG TADI DI INDEX)
# ==========================================
@app.route("/login-page")
def login_page():
    # 1. Cek jika user sudah login (dan sesi masih valid)
    if 'username' in session:
        if session.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))
        elif session.get('role') == 'consumer':
            return redirect(url_for('consumer_home'))
            
    # 2. Ambil parameter
    success_message = request.args.get('success')
    error_message = request.args.get('error')
    is_timeout = request.args.get('timeout') == 'true'
    
    # Format pesan notifikasi biasa
    msg_div = ""
    if success_message:
        msg_div = f'<div style="background:rgba(255,255,255,0.2); color:white; padding:10px; border-radius:5px; margin-bottom:20px; border:1px solid white; text-align:center;">{success_message}</div>'
    if error_message:
        msg_div = f'<div style="background:#ffcccc; color:#cc0000; padding:10px; border-radius:5px; margin-bottom:20px; text-align:center; font-weight:bold;">{error_message}</div>'

    # --- HTML MODAL TIMEOUT (SESUAI GAMBAR) ---
    timeout_modal = ""
    if is_timeout:
        timeout_modal = """
        <style>
            /* Overlay Gelap (Backdrop) */
            .timeout-overlay {
                position: fixed; top: 0; left: 0; width: 100%; height: 100%;
                background-color: rgba(0, 0, 0, 0.6); /* Hitam transparan */
                backdrop-filter: blur(3px);
                z-index: 99999;
                display: flex; justify-content: center; align-items: center;
            }

            /* Kotak Putih */
            .timeout-box {
                background-color: white;
                width: 500px; max-width: 90%;
                padding: 40px 30px;
                border-radius: 20px;
                text-align: center;
                position: relative;
                border: 2px solid #4285F4; /* Border Biru sesuai gambar */
                box-shadow: 0 20px 50px rgba(0,0,0,0.3);
                
                /* Animasi POP */
                animation: popIn 0.4s cubic-bezier(0.175, 0.885, 0.32, 1.275);
            }

            @keyframes popIn {
                0% { transform: scale(0.5); opacity: 0; }
                100% { transform: scale(1); opacity: 1; }
            }

            /* Ikon Jam Pasir Merah */
            .icon-wrapper {
                margin-bottom: 20px;
            }
            .timeout-svg {
                width: 100px; 
                height: 100px;
            }

            /* Teks */
            .timeout-title {
                font-family: 'Segoe UI', sans-serif;
                font-size: 1.3em;
                font-weight: 600;
                color: #000;
                margin-bottom: 10px;
                line-height: 1.4;
            }

            /* Tombol Biru */
            .btn-back-login {
                display: inline-block;
                background-color: #4285F4; /* Warna Biru */
                color: white;
                text-decoration: none;
                padding: 12px 30px;
                border-radius: 8px;
                font-weight: 600;
                font-size: 1em;
                margin-top: 25px;
                transition: background 0.3s;
                box-shadow: 0 4px 10px rgba(66, 133, 244, 0.3);
            }
            .btn-back-login:hover {
                background-color: #3367D6;
            }
        </style>

        <div class="timeout-overlay">
            <div class="timeout-box">
                <div class="icon-wrapper">
                    <svg class="timeout-svg" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
                        <path d="M12 22C17.5228 22 22 17.5228 22 12C22 6.47715 17.5228 2 12 2C6.47715 2 2 6.47715 2 12C2 17.5228 6.47715 22 12 22Z" stroke="#FF3B30" stroke-width="0" fill="transparent"/> 
                        <path d="M12 4V2C6.48 2 2 6.48 2 12H4C4 7.58 7.58 4 12 4ZM20 12C20 16.42 16.42 20 12 20V22C17.52 22 22 17.52 22 12H20ZM12 20C7.58 20 4 16.42 4 12H2C2 17.52 6.48 22 12 22V20ZM22 12C22 7.58 18.42 4 12 4V2C17.52 2 22 6.48 22 12H22Z" fill="#FF3B30"/>
                        <path d="M17.65 6.35C16.2 4.9 14.21 4 12 4V1L7 5L12 9V6C14.21 6 16.2 6.9 17.65 8.35L19.07 6.93C19.07 6.93 17.65 6.35 17.65 6.35Z" fill="#FF3B30"/>
                        <path d="M6.35 17.65C7.8 19.1 9.79 20 12 20V23L17 19L12 15V18C9.79 18 7.8 17.1 6.35 15.65L4.93 17.07C4.93 17.07 6.35 17.65 6.35 17.65Z" fill="#FF3B30"/>
                        
                        <path d="M16 8H8V9.5C8 10.5 8.5 11.5 9.5 12C8.5 12.5 8 13.5 8 14.5V16H16V14.5C16 13.5 15.5 12.5 14.5 12C15.5 11.5 16 10.5 16 9.5V8ZM14 14.5V15H10V14.5C10 13.5 11 13 12 13C13 13 14 13.5 14 14.5ZM12 11C11 11 10 10.5 10 9.5V9H14V9.5C14 10.5 13 11 12 11Z" fill="#FF3B30"/>
                    </svg>
                </div>

                <div class="timeout-title">
                    Maaf, sesi login anda telah habis.<br>
                    Silahkan login kembali.
                </div>

                <a href="/login-page" class="btn-back-login">Kembali ke laman login</a>
            </div>
        </div>
        """

    html = f"""
    <!DOCTYPE html>
    <html lang="id">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Login - Sistem Akuntansi Gurame</title>
        <style>
            body, html {{ margin: 0; padding: 0; height: 100%; font-family: 'Segoe UI', sans-serif; }}
            .login-container {{ display: flex; height: 100vh; width: 100%; }}
            .left-pane {{ flex: 2.5; background-image: url('/static/login_bg.jpg'); background-size: cover; background-position: center; position: relative; }}
            .right-pane {{ flex: 1; background-color: #3E8EAE; display: flex; flex-direction: column; justify-content: center; padding: 0 5%; color: white; box-shadow: -5px 0 20px rgba(0,0,0,0.2); min-width: 350px; }}
            h1.login-title {{ font-size: 3rem; font-weight: 700; margin-bottom: 5px; text-align: center; letter-spacing: 2px; text-transform: uppercase; margin-top: 0; }}
            p.login-subtitle {{ text-align: center; font-size: 1rem; font-weight: 400; margin-top: 0; margin-bottom: 30px; opacity: 0.9; }}
            label {{ display: block; font-size: 1rem; margin-bottom: 6px; font-weight: 500; }}
            input[type="text"], input[type="password"] {{ width: 100%; padding: 10px 15px; border: none; border-radius: 4px; font-size: 1rem; margin-bottom: 20px; box-sizing: border-box; background-color: white; }}
            .form-footer {{ display: flex; justify-content: space-between; align-items: center; margin-top: 10px; }}
            .forgot-pass {{ color: white; text-decoration: underline; font-size: 0.9rem; }}
            .btn-login {{ background-color: #004AAD; color: white; border: none; padding: 10px 35px; font-size: 1rem; font-weight: bold; border-radius: 5px; cursor: pointer; transition: background 0.3s; text-transform: uppercase; box-shadow: 0 4px 6px rgba(0,0,0,0.2); }}
            .btn-login:hover {{ background-color: #003380; }}
            .register-link {{ margin-top: 50px; text-align: center; font-size: 0.9rem; opacity: 0.9; }}
            .register-link a {{ color: white; text-decoration: underline; font-weight: bold; }}
            @media (max-width: 900px) {{ .login-container {{ flex-direction: column; }} .left-pane {{ flex: 1; min-height: 30vh; }} .right-pane {{ flex: 2; padding: 40px 30px; }} }}
        </style>
    </head>
    <body>
        {timeout_modal}
        
        <div class="login-container">
            <div class="left-pane"></div>
            <div class="right-pane">
                {msg_div}
                <h1 class="login-title">LOGIN</h1>
                <p class="login-subtitle">Silahkan login untuk masuk</p>
                <form action="/login" method="POST">
                    <label for="email">Email</label>
                    <input type="text" id="email" name="email" required autocomplete="email">
                    <label for="password">Password</label>
                    <input type="password" id="password" name="password" required autocomplete="current-password">
                    <div class="form-footer">
                        <a href="/forgot-password" class="forgot-pass">Lupa Sandi?</a>
                        <input type="submit" value="LOGIN" class="btn-login">
                    </div>
                </form>
                <div class="register-link">Belum Punya Akun? <a href="/register">Daftar disini</a></div>
            </div>
        </div>
    </body>
    </html>
    """
    return render_template_string(html)

# =========================
# LUPA PASSWORD
# =========================
def render_forgot_password_page(error=None):
    error_html = f'<div style="color: red; font-size: 0.9em; margin-bottom: 15px; text-align: left; background: #ffe6e6; padding: 10px; border-radius: 5px; border-left: 3px solid red;">{error}</div>' if error else ""

    html = f"""
    <!DOCTYPE html>
    <html lang="id">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Lupa Sandi - SIGRAMEE</title>
        <style>
            body, html {{ margin: 0; padding: 0; height: 100%; font-family: 'Segoe UI', sans-serif; background-color: #5896A8; display: flex; justify-content: center; align-items: center; }}
            .main-card {{ display: flex; width: 900px; height: 550px; background-color: rgba(255, 255, 255, 0.15); border-radius: 20px; overflow: hidden; box-shadow: 0 20px 50px rgba(0,0,0,0.2); backdrop-filter: blur(5px); }}
            
            /* KIRI: Gambar Ilustrasi */
            .left-section {{ flex: 1.2; display: flex; justify-content: center; align-items: center; padding: 20px; }}
            .left-section img {{ max-width: 100%; max-height: 90%; object-fit: contain; filter: drop-shadow(10px 10px 20px rgba(0,0,0,0.2)); }}
            
            /* KANAN: Form Putih */
            .right-section {{ flex: 0.8; background-color: white; padding: 50px; display: flex; flex-direction: column; justify-content: center; margin: 25px 25px 25px 0; border-radius: 15px; box-shadow: 0 10px 30px rgba(0,0,0,0.1); }}
            
            h2 {{ font-size: 2.2rem; font-weight: 800; margin: 0 0 10px 0; text-transform: uppercase; letter-spacing: 1px; color: #000; }}
            p.subtitle {{ font-size: 1rem; color: #333; margin: 0 0 30px 0; font-weight: 500; }}
            
            .form-group {{ margin-bottom: 20px; }}
            label {{ display: block; font-size: 1rem; margin-bottom: 8px; color: #333; font-weight: 500; }}
            
            input[type="text"] {{ 
                width: 100%; padding: 15px; border: none; background-color: #DDF3FF; 
                border-radius: 5px; font-size: 1rem; box-sizing: border-box; outline: none; color: #333; 
            }}
            input:focus {{ background-color: #cceeff; box-shadow: 0 0 0 2px #0047AB; }}
            
            .btn-action {{ 
                background-color: #0047AB; color: white; border: none; border-radius: 8px; padding: 15px 0; 
                font-size: 1.1rem; font-weight: 700; cursor: pointer; width: 100%; transition: 0.3s; 
                box-shadow: 0 4px 10px rgba(0, 71, 171, 0.3); margin-top: 20px; 
            }}
            .btn-action:hover {{ background-color: #003380; transform: translateY(-2px); }}
            
            /* Link Kembali */
            .back-link {{ margin-top: 20px; text-align: center; }}
            .back-link a {{ text-decoration: none; color: #666; font-size: 0.9rem; font-weight: 600; }}
            .back-link a:hover {{ color: #0047AB; }}
        </style>
    </head>
    <body>
        <div class="main-card">
            <div class="left-section">
                <img src="/static/forgot_illustration.png" alt="Ilustrasi Lupa Sandi" onerror="this.src='https://img.freepik.com/free-vector/forgot-password-concept-illustration_114360-1123.jpg'">
            </div>
            <div class="right-section">
                <h2>LUPA SANDI</h2>
                <p class="subtitle">Masukkan email yang terdaftar</p>
                {error_html}
                <form action="/forgot-password" method="POST">
                    <div class="form-group">
                        <label>Email:</label>
                        <input type="text" name="security_answer" placeholder="contoh@email.com" required>
                    </div>
                    <button type="submit" class="btn-action">Verifikasi</button>
                </form>
                <div class="back-link"><a href="/login-page">‚Üê Kembali ke Login</a></div>
            </div>
        </div>
    </body>
    </html>
    """
    return render_template_string(html)

# --- HELPER TAMPILAN: RESET PASSWORD ---
def render_reset_password_page(username, error=None):
    error_html = f'<div style="color: red; font-size: 0.9em; margin-bottom: 15px; text-align: left; background: #ffe6e6; padding: 10px; border-radius: 5px; border-left: 3px solid red;">{error}</div>' if error else ""

    html = f"""
    <!DOCTYPE html>
    <html lang="id">
    <head>
        <meta charset="UTF-8">
        <title>Reset Password - SIGRAMEE</title>
        <style>
            body, html {{ margin: 0; padding: 0; height: 100%; font-family: 'Segoe UI', sans-serif; background-color: #5896A8; display: flex; justify-content: center; align-items: center; }}
            .main-card {{ display: flex; width: 900px; height: 550px; background-color: rgba(255, 255, 255, 0.15); border-radius: 20px; overflow: hidden; box-shadow: 0 20px 50px rgba(0,0,0,0.2); backdrop-filter: blur(5px); }}
            
            .left-section {{ flex: 1.2; display: flex; justify-content: center; align-items: center; padding: 20px; }}
            .left-section img {{ max-width: 100%; max-height: 90%; object-fit: contain; filter: drop-shadow(10px 10px 20px rgba(0,0,0,0.2)); }}
            
            .right-section {{ flex: 0.8; background-color: white; padding: 50px; display: flex; flex-direction: column; justify-content: center; margin: 25px 25px 25px 0; border-radius: 15px; box-shadow: 0 10px 30px rgba(0,0,0,0.1); }}
            
            h2 {{ font-size: 1.8rem; font-weight: 800; margin: 0 0 10px 0; text-transform: uppercase; letter-spacing: 1px; color: #000; }}
            p.subtitle {{ font-size: 0.95rem; color: #333; margin: 0 0 30px 0; font-weight: 500; }}
            
            .form-group {{ margin-bottom: 15px; }}
            label {{ display: block; font-size: 0.95rem; margin-bottom: 8px; color: #333; font-weight: 500; }}
            
            input[type="password"] {{ 
                width: 100%; padding: 12px 15px; border: none; background-color: #DDF3FF; 
                border-radius: 5px; font-size: 1rem; box-sizing: border-box; outline: none; color: #333; 
            }}
            input:focus {{ background-color: #cceeff; box-shadow: 0 0 0 2px #0047AB; }}
            
            .btn-action {{ 
                background-color: #0047AB; color: white; border: none; border-radius: 8px; padding: 15px 0; 
                font-size: 1.1rem; font-weight: 700; cursor: pointer; width: 100%; transition: 0.3s; 
                box-shadow: 0 4px 10px rgba(0, 71, 171, 0.3); margin-top: 10px; 
            }}
            .btn-action:hover {{ background-color: #003380; transform: translateY(-2px); }}
        </style>
    </head>
    <body>
        <div class="main-card">
            <div class="left-section">
                <img src="/static/forgot_illustration.png" alt="Ilustrasi Reset" onerror="this.src='https://img.freepik.com/free-vector/reset-password-concept-illustration_114360-7966.jpg'">
            </div>
            <div class="right-section">
                <h2>RESET PASSWORD</h2>
                <p class="subtitle">Silahkan masukkan password baru anda</p>
                {error_html}
                <form action="/reset-password" method="POST">
                    <div class="form-group">
                        <label>Password Baru:</label>
                        <input type="password" name="new_password" required>
                    </div>
                    <div class="form-group">
                        <label>Konfirmasi Password Baru:</label>
                        <input type="password" name="confirm_password" required>
                    </div>
                    <button type="submit" class="btn-action">Simpan</button>
                </form>
            </div>
        </div>
    </body>
    </html>
    """
    return render_template_string(html)

@app.route("/forgot-password", methods=['GET', 'POST'])
def forgot_password():
    """
    Halaman 1: Verifikasi Email.
    Jika email ada di DB -> Redirect ke Reset Password.
    """
    
    if request.method == 'POST':
        # Ambil input email dari form
        email_input = request.form.get('security_answer')
        
        db = get_db()
        # Cek apakah email ada di kolom security_answer (atau kolom email jika ada)
        user = db.execute("SELECT * FROM users WHERE security_answer = ?", (email_input,)).fetchone()
        
        if user:
            # --- KUNCI KEBERHASILAN ---
            # 1. Simpan username target di session agar halaman selanjutnya tahu siapa yang mau direset
            session['user_to_reset'] = user['username']
            
            # 2. Redirect LANGSUNG ke route reset_password
            return redirect(url_for('reset_password'))
        else:
            # Jika email tidak ditemukan
            return render_forgot_password_page(error="Email tidak terdaftar dalam sistem.")

    # Jika method GET (baru buka halaman)
    return render_forgot_password_page()


# ==========================================
# ROUTE: RESET PASSWORD
# ==========================================
@app.route("/reset-password", methods=['GET', 'POST'])
def reset_password():
    """
    Halaman 2: Input Password Baru.
    Hanya bisa diakses jika sudah lolos verifikasi email.
    """
    
    # KEAMANAN: Cek apakah user sudah melewati tahap verifikasi email
    if 'user_to_reset' not in session:
        return redirect(url_for('forgot_password', error="Silakan verifikasi email terlebih dahulu."))
    
    username = session['user_to_reset']
    
    if request.method == 'POST':
        new_pass = request.form.get('new_password')
        confirm_pass = request.form.get('confirm_password')
        
        if not new_pass or not confirm_pass:
             return render_reset_password_page(username, error="Password tidak boleh kosong.")
        
        if new_pass != confirm_pass:
            return render_reset_password_page(username, error="Konfirmasi password tidak cocok.")
        
        # Proses Update Password di Database
        try:
            db = get_db()
            with db:
                db.execute("UPDATE users SET password = ? WHERE username = ?", (new_pass, username))
            
            # BERHASIL: Hapus sesi reset dan lempar ke login
            session.pop('user_to_reset', None)
            return redirect(url_for('login_page', success="Password berhasil direset! Silakan login dengan sandi baru."))
            
        except Exception as e:
            return render_reset_password_page(username, error=f"Terjadi kesalahan database: {e}")

    # Tampilkan form reset password
    return render_reset_password_page(username)

# =========================
# LOGIN
# =========================
@app.route("/login", methods=['POST'])
def login():
    email = request.form['email']
    password = request.form['password']
    
    db = get_db()
    # Query ke 'security_answer' BUKAN 'username'
    user = db.execute("SELECT * FROM users WHERE security_answer = ?", (email,)).fetchone()
    
    if user and user['password'] == password:
        # User terotentikasi, sekarang kirim OTP
        user_email = user['security_answer'] # Ambil email penerima
        
        # 1. Buat Kode OTP 6 digit
        otp_code = str(random.randint(100000, 999999))
        
        # 2. Tetapkan waktu kedaluwarsa (1 menit dari sekarang, gunakan UTC)
        session['otp_expiry'] = datetime.now(timezone.utc) + timedelta(minutes=1)
        
        # 3. Simpan kode dan info user sementara di session
        session['otp_code'] = otp_code
        session['otp_user_info'] = {'username': user['username'], 'role': user['role']}
        
        # 4. KIRIM EMAIL (Panggilan ke fungsi pengirim email yang sesungguhnya)
        sukses_kirim = send_otp_email(user_email, otp_code) # <-- KODE DIPERBAIKI!
        
        if sukses_kirim:
            # 5. Redirect ke halaman verifikasi HANYA JIKA email sukses terkirim
            return redirect(url_for('verify_otp'))
        else:
            # 6. Jika email gagal terkirim
            # Pesan error akan muncul di terminal (contoh: App Password salah)
            return redirect(url_for('index', error="Gagal mengirim email verifikasi. Cek kembali App Password Anda (di terminal akan ada detail error)."))
    else:
        # Jika data tidak ada atau password salah, kirim pesan error
        return redirect(url_for('index', error="Email atau Password salah."))
        return redirect(url_for('index', error="Email atau Password salah.")) 

# --- GANTI BAGIAN INI DI DALAM FILE PYTHON ANDA ---

def render_verify_otp_page(email_tujuan, error=None):
    """
    Merender halaman OTP dengan desain khusus sesuai gambar referensi.
    Layout: Full Screen, Card Centered, Split View (Image Left, Form Right).
    """
    
    error_html = ""
    if error:
        error_html = f'<div style="color: red; font-size: 0.9em; margin-bottom: 15px; text-align: center; background: #ffe6e6; padding: 8px; border-radius: 5px;">{error}</div>'

    html = f"""
    <!DOCTYPE html>
    <html lang="id">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Verifikasi OTP</title>
        <style>
            body, html {{
                margin: 0; padding: 0;
                height: 100%;
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                background-color: #E3F2FD; /* Latar belakang biru muda */
                display: flex;
                justify-content: center;
                align-items: center;
            }}

            .otp-container {{
                background-color: white;
                width: 900px;
                height: 550px;
                border-radius: 20px;
                box-shadow: 0 10px 25px rgba(0,0,0,0.1);
                display: flex;
                overflow: hidden;
            }}

            /* --- SISI KIRI (GAMBAR) --- */
            .otp-left {{
                flex: 1;
                background-color: #F0F8FF;
                display: flex;
                justify-content: center;
                align-items: center;
                padding: 40px;
                position: relative;
            }}
            
            .otp-left img {{
                max-width: 100%;
                height: auto;
                object-fit: contain;
            }}

            /* --- SISI KANAN (FORM) --- */
            .otp-right {{
                flex: 1;
                padding: 50px;
                display: flex;
                flex-direction: column;
                justify-content: center;
                text-align: center;
            }}

            h2 {{
                margin: 0 0 15px 0;
                color: #333;
                font-size: 2rem;
            }}

            p.desc {{
                color: #666;
                font-size: 1rem;
                margin-bottom: 30px;
                line-height: 1.5;
            }}

            .email-highlight {{
                color: #000;
                font-weight: bold;
            }}

            /* --- INPUT KOTAK-KOTAK --- */
            .otp-inputs {{
                display: flex;
                justify-content: center;
                gap: 10px;
                margin-bottom: 25px;
            }}

            .otp-box {{
                width: 50px;
                height: 55px;
                border: 1px solid #ddd;
                background-color: #E1F5FE; /* Biru sangat muda sesuai gambar */
                border-radius: 8px;
                font-size: 1.5rem;
                font-weight: bold;
                text-align: center;
                color: #333;
                outline: none;
                transition: all 0.3s;
            }}

            .otp-box:focus {{
                border-color: #42A5F5;
                background-color: #fff;
                box-shadow: 0 0 5px rgba(66, 165, 245, 0.5);
            }}

            /* --- TOMBOL --- */
            .btn-verify {{
                background-color: #4285F4; /* Biru Google/Standar */
                color: white;
                border: none;
                border-radius: 8px;
                padding: 15px;
                font-size: 1.1rem;
                font-weight: bold;
                cursor: pointer;
                width: 100%;
                transition: background 0.3s;
                box-shadow: 0 4px 6px rgba(66, 133, 244, 0.3);
            }}

            .btn-verify:hover {{
                background-color: #3367D6;
                transform: translateY(-1px);
            }}

            .resend-link {{
                margin-top: 20px;
                font-size: 0.9rem;
                color: #666;
            }}

            .resend-link a {{
                color: #000;
                font-weight: bold;
                text-decoration: none;
                border-bottom: 1px solid #000;
            }}
            
            .resend-link a:hover {{
                color: #4285F4;
                border-color: #4285F4;
            }}

            /* Responsif untuk layar kecil */
            @media (max-width: 768px) {{
                .otp-container {{
                    flex-direction: column;
                    width: 90%;
                    height: auto;
                }}
                .otp-left {{
                    display: none; /* Sembunyikan gambar di HP agar rapi */
                }}
                .otp-right {{
                    padding: 30px;
                }}
            }}
        </style>
    </head>
    <body>

        <div class="otp-container">
            <div class="otp-left">
                <img src="/static/otp_image.png" alt="OTP Illustration" onerror="this.src='https://img.freepik.com/free-vector/two-factor-authentication-concept-illustration_114360-5288.jpg'"> 
            </div>
            
            <div class="otp-right">
                <h2>Verifikasi OTP</h2>
                <p class="desc">
                    Kode telah dikirimkan ke email<br>
                    <span class="email-highlight">{email_tujuan}</span>
                </p>

                {error_html}

                <form id="otpForm" action="/verify-otp" method="POST">
                    <input type="hidden" name="otp_code" id="real_otp_code">
                    
                    <div class="otp-inputs">
                        <input type="text" class="otp-box" maxlength="1" oninput="handleInput(this, 0)" onkeydown="handleBackspace(event, 0)">
                        <input type="text" class="otp-box" maxlength="1" oninput="handleInput(this, 1)" onkeydown="handleBackspace(event, 1)">
                        <input type="text" class="otp-box" maxlength="1" oninput="handleInput(this, 2)" onkeydown="handleBackspace(event, 2)">
                        <input type="text" class="otp-box" maxlength="1" oninput="handleInput(this, 3)" onkeydown="handleBackspace(event, 3)">
                        <input type="text" class="otp-box" maxlength="1" oninput="handleInput(this, 4)" onkeydown="handleBackspace(event, 4)">
                        <input type="text" class="otp-box" maxlength="1" oninput="handleInput(this, 5)" onkeydown="handleBackspace(event, 5)">
                    </div>

                    <div class="resend-link">
                        Tidak menerima kode? <br>
                        <a href="/login-page">Kirim ulang kode</a> </div>

                    <br>
                    <button type="submit" class="btn-verify" onclick="combineOtp()">Verifikasi</button>
                </form>
            </div>
        </div>

        <script>
            // Script untuk Auto-Focus pindah kotak
            function handleInput(elm, index) {{
                // Jika user mengetik angka
                if (elm.value.length === 1) {{
                    // Pindah ke kotak berikutnya jika ada
                    const next = document.querySelectorAll('.otp-box')[index + 1];
                    if (next) {{
                        next.focus();
                    }}
                }}
                // Gabungkan kode setiap kali mengetik
                combineOtp();
            }}

            function handleBackspace(e, index) {{
                // Jika tombol Backspace ditekan dan kotak kosong
                if (e.key === 'Backspace' && e.target.value === '') {{
                    // Pindah ke kotak sebelumnya
                    const prev = document.querySelectorAll('.otp-box')[index - 1];
                    if (prev) {{
                        prev.focus();
                    }}
                }}
            }}

            function combineOtp() {{
                // Mengambil semua value dari kotak dan menggabungkannya ke input hidden
                let inputs = document.querySelectorAll('.otp-box');
                let combined = '';
                inputs.forEach(input => {{
                    combined += input.value;
                }});
                document.getElementById('real_otp_code').value = combined;
            }}
        </script>
    </body>
    </html>
    """
    return render_template_string(html)

@app.route("/verify-otp", methods=['GET', 'POST'])
def verify_otp():
    # Cek apakah pengguna punya data OTP di session
    if 'otp_code' not in session or 'otp_expiry' not in session:
        return redirect(url_for('index', error="Sesi verifikasi tidak ditemukan. Silakan login kembali."))

    # Ambil email user dari session info (untuk ditampilkan di layar)
    user_info = session.get('otp_user_info', {})
    
    # Ambil security_answer (email) dari database untuk ditampilkan
    # Ini penting agar tampilan "Kode dikirim ke ..." akurat
    db = get_db()
    u = db.execute("SELECT security_answer FROM users WHERE username = ?", (user_info.get('username'),)).fetchone()
    email_to_show = u['security_answer'] if u else "Email Anda"

    try:
        expiry_time_raw = session.get('otp_expiry')
        if isinstance(expiry_time_raw, str):
            expiry_time = datetime.fromisoformat(expiry_time_raw).replace(tzinfo=timezone.utc)
        elif isinstance(expiry_time_raw, datetime):
            expiry_time = expiry_time_raw.replace(tzinfo=timezone.utc)
        else:
             return redirect(url_for('index', error="Tipe data sesi OTP tidak valid. Silakan login kembali."))

    except Exception as e:
        print(f"ERROR Parsing OTP Expiry: {e}")
        return redirect(url_for('index', error="Terjadi kesalahan waktu sesi. Silakan login kembali."))

    # Cek Kedaluwarsa
    if datetime.now(timezone.utc) > expiry_time:
        session.pop('otp_code', None)
        session.pop('otp_expiry', None)
        session.pop('otp_user_info', None)
        return redirect(url_for('index', error="Kode OTP telah kedaluwarsa. Silakan login kembali."))

    if request.method == 'POST':
        # Proses verifikasi kode
        user_code = request.form.get('otp_code') # Ambil dari input hidden
        
        if user_code == session['otp_code']:
            # BERHASIL!
            session['username'] = user_info['username']
            session['role'] = user_info['role']
            
            # Hapus data OTP
            session.pop('otp_code', None)
            session.pop('otp_expiry', None)
            session.pop('otp_user_info', None)
            
            # Arahkan ke dashboard
            if session['role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('consumer_home'))
        else:
            # Kode salah - Render ulang halaman OTP dengan error
            # PERHATIKAN: Di sini kita memanggil render_verify_otp_page (bukan form)
            return render_verify_otp_page(email_to_show, error="Kode yang Anda masukkan salah.")

    # --- JIKA GET REQUEST (Baru masuk halaman) ---
    # PERHATIKAN: Di sini kita memanggil render_verify_otp_page (bukan form)
    return render_verify_otp_page(email_to_show)

@app.route("/register", methods=['GET', 'POST'])
def register():
    """Menangani halaman registrasi pengguna baru. Hanya boleh mendaftar sebagai consumer."""
    
    # Email admin yang hanya diperbolehkan
    ADMIN_EMAIL_HARDCODED = 'sigrameei@gmail.com'
    
    if request.method == 'POST':
        # 1. Ambil data dari form
        username = request.form['username']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        security_answer = request.form['security_answer'] # Ini "email"
        
        # 2. Validasi
        if password != confirm_password:
            return render_register_form(error="Password tidak cocok.")
        
        if not username or not password or not security_answer:
            return render_register_form(error="Semua kolom wajib diisi.")
            
        # --- BATASAN KEAMANAN BARU: HANYA SATU EMAIL ADMIN YANG DIIZINKAN ---
        if security_answer.lower() == ADMIN_EMAIL_HARDCODED:
             return render_register_form(error=f"Email '{security_answer}' sudah terdaftar sebagai akun Admin utama dan tidak dapat didaftarkan ulang.")

        # 3. Simpan ke database
        db = get_db()
        try:
            with db:
                # Otomatis daftarkan pengguna baru sebagai 'consumer'
                db.execute(
                    "INSERT INTO users (username, password, role, security_answer) VALUES (?, ?, ?, ?)",
                    (username, password, 'consumer', security_answer) # Role DITETAPKAN SEBAGAI 'consumer'
                )
            # 4. Jika sukses, kembali ke login
            return redirect(url_for('index', success="Registrasi berhasil! Silakan login."))
            
        except sqlite3.IntegrityError:
            # Ini akan error jika username sudah ada (karena UNIQUE)
            return render_register_form(error="Username ini sudah terdaftar. Silakan pilih nama lain.")

    # Jika method GET, tampilkan form
    return render_register_form()

def render_register_form(error=None):
    """
    Helper untuk merender form registrasi dengan layout Split Screen khusus.
    Kiri: Form (Background Biru/Tosca).
    Kanan: Gambar Ikan Gurame.
    """
    
    # Menampilkan pesan error jika ada
    error_html = ""
    if error:
        error_html = f"""
        <div style="background-color: #ffcccc; color: #cc0000; padding: 10px; border-radius: 5px; margin-bottom: 20px; font-weight: bold; text-align: center;">
            {error}
        </div>
        """

    html = f"""
    <!DOCTYPE html>
    <html lang="id">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Register - SIGRAMEE</title>
        <style>
            /* RESET & FONT */
            body, html {{
                margin: 0; padding: 0;
                height: 100%;
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            }}

            /* CONTAINER UTAMA (SPLIT SCREEN) */
            .register-container {{
                display: flex;
                height: 100vh;
                width: 100%;
            }}

            /* --- SISI KIRI: FORM --- */
            .left-pane {{
                flex: 1;
                background-color: #3E8EAE; /* Warna Biru/Tosca sesuai gambar */
                display: flex;
                flex-direction: column;
                justify-content: center;
                padding: 40px 60px;
                color: white;
                min-width: 400px;
                position: relative;
            }}

            /* TYPOGRAPHY */
            h1.reg-title {{
                font-size: 3.5rem;
                font-weight: 800;
                margin: 0;
                text-transform: uppercase;
                letter-spacing: 1px;
                line-height: 1;
            }}

            p.reg-subtitle {{
                font-size: 1.1rem;
                margin-top: 10px;
                margin-bottom: 30px;
                font-weight: 400;
                opacity: 0.9;
            }}

            /* FORM ELEMENTS */
            label {{
                display: block;
                font-size: 1.1rem;
                margin-bottom: 8px;
                font-weight: 500;
            }}

            input[type="text"], input[type="password"] {{
                width: 100%;
                padding: 12px 15px;
                border: none;
                border-radius: 4px;
                font-size: 1rem;
                margin-bottom: 20px;
                box-sizing: border-box;
                outline: none;
            }}

            input[type="text"]:focus, input[type="password"]:focus {{
                box-shadow: 0 0 0 3px rgba(255,255,255,0.3);
            }}

            /* BUTTON DAFTAR */
            .btn-daftar {{
                background-color: #0047AB; /* Biru Tua */
                color: white;
                padding: 12px 40px;
                border: none;
                border-radius: 5px;
                font-size: 1.2rem;
                font-weight: 800;
                text-transform: uppercase;
                cursor: pointer;
                transition: background 0.3s;
                float: right; /* Tombol di kanan */
                box-shadow: 0 4px 6px rgba(0,0,0,0.2);
            }}

            .btn-daftar:hover {{
                background-color: #003380;
            }}
            
            /* FOOTER LINK */
            .login-link {{
                margin-top: 60px;
                font-size: 1rem;
            }}
            .login-link a {{
                color: white;
                text-decoration: underline;
                font-weight: bold;
            }}

            /* --- SISI KANAN: GAMBAR --- */
            .right-pane {{
                flex: 1.5;
                /* Menggunakan gambar placeholder Gurame, ganti url ini jika punya file lokal di /static/ */
                background-image: url('/static/register_bg.png');                 
                background-size: cover;
                background-position: center;
                position: relative;
            }}

            /* RESPONSIVE (HP) */
            @media (max-width: 900px) {{
                .register-container {{ flex-direction: column; }}
                .right-pane {{ display: none; }} /* Sembunyikan gambar di layar kecil */
                .left-pane {{ width: 100%; padding: 30px; min-width: auto; }}
                h1.reg-title {{ font-size: 2.5rem; }}
            }}
        </style>
    </head>
    <body>
        <div class="register-container">
            
            <div class="left-pane">
                <h1 class="reg-title">REGISTER</h1>
                <p class="reg-subtitle">Silahkan buat akun baru:</p>

                {error_html}

                <form action="/register" method="POST">
                    <label>Username</label>
                    <input type="text" name="username" required autocomplete="username">
                    
                    <label>Password</label>
                    <input type="password" name="password" required autocomplete="new-password">
                    
                    <label>Konfirmasi Password</label>
                    <input type="password" name="confirm_password" required autocomplete="new-password">
                    
                    <label>Email</label>
                    <input type="text" name="security_answer" required>

                    <div style="overflow: hidden; margin-top: 10px;">
                        <input type="submit" value="DAFTAR" class="btn-daftar">
                    </div>
                </form>

                <div class="login-link">
                    Sudah Punya Akun? <a href="/login-page">Login disini</a>
                </div>
            </div>

            <div class="right-pane"></div>
            
        </div>
    </body>
    </html>
    """
    return render_template_string(html)

def get_initials(name):
    """Mengubah 'Ika Alia' menjadi 'IA'"""
    if not name: return "U"
    parts = name.strip().split()
    if len(parts) >= 2:
        return (parts[0][0] + parts[1][0]).upper()
    return parts[0][:2].upper()

def format_phone_indo(phone):
    """Mengubah 08xx menjadi +628xx"""
    if not phone: return "08XX-XXXX-XXXX"
    clean_phone = phone.strip()
    if clean_phone.startswith("0"):
        return "+62" + clean_phone[1:]
    return clean_phone

@app.route("/logout")
def logout():
    # Bersihkan session login utama
    session.pop('username', None)
    session.pop('role', None)
    
    # Bersihkan juga session OTP jika ada
    session.pop('otp_code', None)
    session.pop('otp_expiry', None)
    session.pop('otp_user_info', None)
    
    return redirect(url_for('login_page'))

# =========================
# ROUTE KONSUMEN
# =========================
@app.route("/home")
def consumer_home():
    if session.get('role') != 'consumer':
        return redirect(url_for('index'))
    
    username = session.get('username', 'Customer')
    
    success_msg = request.args.get('success')
    error_msg = request.args.get('error')
    
    db = get_db()
    
    user_data = db.execute("SELECT phone_number FROM users WHERE username = ?", (username,)).fetchone()
    phone = user_data['phone_number'] if user_data else ""
    
    initials = get_initials(username)
    display_phone = format_phone_indo(phone)
    
    stats = db.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as waiting,
            SUM(CASE WHEN status = 'verified' THEN 1 ELSE 0 END) as success,
            SUM(CASE WHEN status = 'rejected' THEN 1 ELSE 0 END) as failed
        FROM online_payments 
        WHERE buyer_name = ?
    """, (username,)).fetchone()
    
    total_trx = stats['total'] or 0
    waiting = stats['waiting'] or 0
    success = stats['success'] or 0
    failed = stats['failed'] or 0
    in_process = 0 

    alert_html = ""
    if success_msg:
        alert_html = f"""
        <div style="background: rgba(40, 167, 69, 0.9); color: white; padding: 15px; border-radius: 10px; margin-bottom: 20px; box-shadow: 0 5px 15px rgba(0,0,0,0.1); display: flex; align-items: center; gap: 10px;">
            <span>‚úÖ</span> {success_msg}
        </div>
        """
    elif error_msg:
        alert_html = f"""
        <div style="background: rgba(220, 53, 69, 0.9); color: white; padding: 15px; border-radius: 10px; margin-bottom: 20px; box-shadow: 0 5px 15px rgba(0,0,0,0.1); display: flex; align-items: center; gap: 10px;">
            <span>‚ö†Ô∏è</span> {error_msg}
        </div>
        """

    content_html = f"""
        <div class="profile-card">
            <div class="profile-content">
                <div class="profile-info">
                    <div class="avatar-circle">{initials}</div>
                    <div class="name-section">
                        <h2 style="color: white; margin-bottom: 5px;">{username}</h2> 
                        <div class="badge-member">MEMBER</div>
                    </div>
                </div>
                <a href="/consumer/profile-settings" class="settings-btn">
                    <svg style="width:28px; height:28px; fill:none; stroke:currentColor; stroke-width:2;" viewBox="0 0 24 24">
                        <circle cx="12" cy="12" r="3"></circle>
                        <path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1 0 2.83 2 2 0 0 1-2.83 0l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-2 2 2 2 0 0 1-2-2v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83 0 2 2 0 0 1 0-2.83l.06-.06a1.65 1.65 0 0 0 .33-1.82 1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1-2-2 2 2 0 0 1 2-2h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 0-2.83 2 2 0 0 1 2.83 0l.06.06a1.65 1.65 0 0 0 1.82.33H9a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 2-2 2 2 0 0 1 2 2v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 0 2 2 0 0 1 0 2.83l.06.06a1.65 1.65 0 0 0-.33 1.82V9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 2 2 2 2 0 0 1-2 2h-.09a1.65 1.65 0 0 0-1.51 1z"></path>
                    </svg>
                </a>
            </div>
            
            <div class="divider-line"></div>
            
            <div class="profile-phone">
                <svg class="phone-icon" viewBox="0 0 24 24">
                    <path d="M22 16.92v3a2 2 0 0 1-2.18 2 19.79 19.79 0 0 1-8.63-3.07 19.5 19.5 0 0 1-6-6 19.79 19.79 0 0 1-3.07-8.67A2 2 0 0 1 4.11 2h3a2 2 0 0 1 2 1.72 12.84 12.84 0 0 0 .7 2.81 2 2 0 0 1-.45 2.11L8.09 9.91a16 16 0 0 0 6 6l1.27-1.27a2 2 0 0 1 2.11-.45 12.84 12.84 0 0 0 2.81.7A2 2 0 0 1 22 16.92z"></path>
                </svg> 
                {display_phone}
            </div>
        </div>

        <div class="stats-label">Transaksi Hari Ini</div>
        
        <div class="stats-grid">
            <a href="/consumer/transactions?status=all" class="stat-link-box total-box">
                <div class="total-number">{total_trx}</div>
                <div class="total-text">Total Transaksi</div>
            </a>

            <div class="status-grid">
                <a href="/consumer/transactions?status=pending" class="stat-link-box status-box bg-orange">
                    <div class="status-num">{waiting}</div>
                    <div class="status-text">Menunggu Konfirmasi</div>
                </a>
                <a href="/consumer/transactions?status=verified" class="stat-link-box status-box bg-green">
                    <div class="status-num">{success}</div>
                    <div class="status-text">Sukses</div>
                </a>
                <a href="/consumer/transactions?status=pending" class="stat-link-box status-box bg-blue">
                    <div class="status-num">{in_process}</div>
                    <div class="status-text">Dalam Proses</div>
                </a>
                <a href="/consumer/transactions?status=rejected" class="stat-link-box status-box bg-red">
                    <div class="status-num">{failed}</div>
                    <div class="status-text">Gagal</div>
                </a>
            </div>
        </div>
    """
    
    return render_consumer_page("Dashboard", content_html, username, initials, display_phone, stats, alert_html, active_menu="Dashboard")

@app.route("/consumer/transactions")
def consumer_transactions_list():
    if session.get('role') != 'consumer':
        return redirect(url_for('index'))

    username = session.get('username')
    db = get_db()
    
    # --- Ambil Filter dari URL ---
    status_filter = request.args.get('status', 'all')
    invoice_no_filter = request.args.get('invoice_no', '').strip()
    start_date_filter = request.args.get('start_date', '')
    end_date_filter = request.args.get('end_date', '')
    
    # 1. Query Data Awal
    query = """
        SELECT op.*, 
               ii.item_name, 
               ii.sale_price_per_unit
        FROM online_payments op
        JOIN inventory_items ii ON op.item_id = ii.id
        WHERE op.buyer_name = ?
    """
    params = [username]
    status_display = "Semua Transaksi"
    
    # 2. Logika Filter (Menambah klausa WHERE)
    if status_filter != 'all':
        query += " AND op.status = ?"
        params.append(status_filter)
        
        if status_filter == 'pending':
            status_display = "Menunggu Konfirmasi"
        elif status_filter == 'verified':
            status_display = "Sukses"
        elif status_filter == 'rejected':
            status_display = "Gagal"
    
    if invoice_no_filter:
        query += " AND op.order_id LIKE ?"
        params.append(f"%{invoice_no_filter}%")
        status_display = f"Hasil Pencarian: {invoice_no_filter}"
        
    if start_date_filter:
        query += " AND op.created_at >= ?"
        params.append(f"{start_date_filter} 00:00:00")

    if end_date_filter:
        query += " AND op.created_at <= ?"
        params.append(f"{end_date_filter} 23:59:59")
        
    query += " ORDER BY op.created_at DESC"
    
    # 3. Eksekusi Query
    transactions = db.execute(query, params).fetchall()
    
    # --- START: HTML/CSS LAYOUT BARU ---
    style_css = f"""
    <style>
        /* Kontainer Utama */
        .single-trx-container {{
            width: 100%;
            background: transparent !important; 
            box-shadow: none !important;
            padding: 0;
            margin: 0;
            color: white; 
        }}
        
        /* Tombol Dashboard */
        .dashboard-link-box {{
            margin-bottom: 25px;
        }}
        .btn-dashboard {{
            background: rgba(255, 255, 255, 0.2); 
            color: white; 
            padding: 10px 20px; 
            border-radius: 30px; 
            text-decoration: none;
            font-weight: 500; 
            display: inline-flex; 
            align-items: center; 
            gap: 5px; 
            border: 1px solid rgba(255, 255, 255, 0.5);
            transition: 0.3s;
        }}
        .btn-dashboard:hover {{ background: rgba(255, 255, 255, 0.3); }}
        .svg-back {{ width: 20px; height: 20px; fill: none; stroke: currentColor; stroke-width: 2; }}

        /* --- FILTER BOX STYLE (SESUAI GAMBAR) --- */
        .filter-box {{
            background: rgba(255, 255, 255, 0.15); /* Transparan putih dikit */
            border: 1px solid rgba(255, 255, 255, 0.3);
            border-radius: 15px;
            padding: 30px;
            margin-bottom: 30px;
            color: white;
            box-shadow: 0 4px 15px rgba(0,0,0,0.05);
        }}
        
        /* Grid Layout untuk Filter: Kiri (Status/Cari) & Kanan (Tanggal) */
        .filter-main-grid {{
            display: grid;
            grid-template-columns: 1.5fr 1fr; /* Kiri lebih lebar dikit */
            gap: 40px;
            align-items: start;
        }}
        
        .filter-left-col {{
            display: flex;
            flex-direction: column;
            gap: 20px;
        }}
        
        .filter-right-col {{
            display: flex;
            flex-direction: column;
            gap: 20px;
        }}
        
        /* Baris tanggal (Mulai & Akhir sejajar) */
        .date-row {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 20px;
        }}

        .filter-group {{
            display: flex;
            flex-direction: column;
            gap: 8px;
        }}
        
        .filter-label {{
            font-weight: 600;
            font-size: 1.1em;
            color: white;
        }}
        
        /* Style Input Glassmorphism (Hijau Tosca Transparan) */
        .filter-input, .filter-select {{
            width: 100%;
            padding: 12px 15px;
            border: none;
            border-radius: 10px;
            background: rgba(108, 194, 183, 0.5); /* Warna Input Sesuai Gambar */
            color: white;
            box-sizing: border-box;
            font-weight: 500;
            font-size: 1em;
            outline: none;
        }}
        
        .filter-input::placeholder {{
            color: rgba(255, 255, 255, 0.7);
        }}
        
        .filter-select option {{
            background: #5DA090; /* Background saat dropdown dibuka */
            color: white;
        }}
        
        /* Tombol Terapkan Filter */
        .btn-filter-submit {{
            background-color: rgba(255, 255, 255, 0.2);
            color: white;
            padding: 10px;
            border: 1px solid rgba(255, 255, 255, 0.5);
            border-radius: 10px;
            font-weight: bold;
            cursor: pointer;
            transition: 0.3s;
            margin-top: auto; /* Dorong ke bawah jika ada space */
        }}
        .btn-filter-submit:hover {{
            background-color: rgba(255, 255, 255, 0.3);
        }}

        /* --- TABEL DATA STYLE --- */
        .trx-data-wrapper {{
            background-color: transparent; 
            border: 1px solid rgba(255, 255, 255, 0.5);
            border-radius: 15px; 
            padding: 20px;
            overflow-x: auto; 
            box-shadow: 0 4px 15px rgba(0,0,0,0.05);
        }}
        .trx-table-content {{ min-width: 750px; }}
        .trx-row-item {{
            display: flex;
            padding: 15px 0;
            gap: 20px; 
            background: transparent;
            color: white;
            justify-content: flex-start;
            border-top: 1px solid rgba(255, 255, 255, 0.2); 
            align-items: center;
        }}
        .trx-header-row {{ border-top: none; padding-bottom: 10px; }}
        
        .trx-line-group {{ display: flex; flex-direction: column; flex-shrink: 0; }}
        
        /* Lebar Kolom */
        .trx-line-group:nth-child(1) {{ width: 25%; }} /* Invoice */
        .trx-line-group:nth-child(2) {{ width: 25%; }} /* Item */
        .trx-line-group:nth-child(3) {{ width: 20%; }} /* Harga */
        .trx-line-group:nth-child(4) {{ width: 15%; }} /* Tanggal */
        .trx-line-group:nth-child(5) {{ width: 15%; }} /* Status */
        
        .label-field {{
            font-size: 1em; color: white; font-weight: 600; 
            margin-bottom: 5px; text-transform: capitalize; 
        }}
        .data-field {{
            font-size: 1em; font-weight: 500; color: white; white-space: nowrap; 
        }}
        
        /* Badge Status */
        .status-badge-trx {{ 
            padding: 5px 15px; 
            border-radius: 20px; 
            font-weight: 500; 
            font-size: 0.9em; 
            display: inline-flex; 
            align-items: center; 
            justify-content: center; 
            color: white; 
            text-align: center;
            min-width: 100px;
            height: 30px; 
            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        }}
        .bg-sukses {{ background-color: #A3D89D; }}
        .bg-pending {{ background-color: #FFC999; color: #A0522D; }}
        .bg-gagal {{ background-color: #D46A6A; }}
        
        @media (max-width: 768px) {{
            .filter-main-grid {{ grid-template-columns: 1fr; gap: 20px; }}
            .date-row {{ grid-template-columns: 1fr; }}
        }}
    </style>
    """
    
    # --- Filter Options ---
    status_options = [
        ('all', 'Semua Status'),
        ('pending', 'Menunggu Konfirmasi'),
        ('verified', 'Sukses'),
        ('rejected', 'Gagal')
    ]

    # --- HTML Form Filter (LAYOUT DIPERBAIKI) ---
    filter_form_html = f"""
    <h1 style="color: white; margin-bottom: 5px;">Riwayat Transaksi</h1>
    <p style="color: rgba(255, 255, 255, 0.9); margin-top: 0; margin-bottom: 30px;">Menampilkan data riwayat transaksi yang telah anda lakukan</p>
    
    <form method="GET" action="/consumer/transactions" class="filter-box">
        <div class="filter-main-grid">
            
            <div class="filter-left-col">
                <div class="filter-group">
                    <label class="filter-label">Status</label>
                    <select name="status" class="filter-select">
                        {''.join([
                            f'<option value="{val}" {"selected" if status_filter == val else ""}>{label}</option>' 
                            for val, label in status_options
                        ])}
                    </select>
                </div>
                
                <div class="filter-group">
                    <label class="filter-label">Cari</label>
                    <input type="text" name="invoice_no" value="{invoice_no_filter}" placeholder="Nomor Invoice" class="filter-input">
                </div>
            </div>
            
            <div class="filter-right-col">
                <div class="date-row">
                    <div class="filter-group">
                        <label class="filter-label">Mulai</label>
                        <input type="date" name="start_date" value="{start_date_filter}" class="filter-input">
                    </div>
                    <div class="filter-group">
                        <label class="filter-label">Akhir</label>
                        <input type="date" name="end_date" value="{end_date_filter}" class="filter-input">
                    </div>
                </div>
                
                <button type="submit" class="btn-filter-submit">Terapkan Filter</button>
            </div>
            
        </div>
    </form>
    """
    
    # 4. Render Data Rows
    transaction_rows = ""
    
    if transactions:
        # Header Tabel
        transaction_rows += f"""
        <div class="trx-header-row trx-row-item">
            <div class="trx-line-group"><span class="label-field">Nomor Invoice</span></div>
            <div class="trx-line-group"><span class="label-field">Item</span></div>
            <div class="trx-line-group"><span class="label-field">Harga</span></div>
            <div class="trx-line-group"><span class="label-field">Tanggal</span></div>
            <div class="trx-line-group"><span class="label-field">Status</span></div>
        </div>
        """
        
        # Loop Data
        for trx in transactions:
            try:
                dt_obj = datetime.fromisoformat(trx['created_at']).replace(tzinfo=timezone.utc).astimezone(timezone(timedelta(hours=7)))
                date_display = dt_obj.strftime('%d-%m-%Y')
                time_display = dt_obj.strftime('%H.%M')
                tanggal_lengkap = f"{date_display} {time_display}"
            except:
                tanggal_lengkap = trx['created_at'][:16].replace('T', ' ').replace('-', '-')

            item_qty = Decimal(str(trx['qty'] or 0))
            price_per_unit = Decimal(str(trx['sale_price_per_unit'] or 0))
            total_harga = price_per_unit * item_qty
            
            status_class = "bg-pending"
            status_text = "Menunggu Konfir"
            if trx['status'] == 'verified':
                status_class = "bg-sukses"
                status_text = "Sukses"
            elif trx['status'] == 'rejected':
                status_class = "bg-gagal"
                status_text = "Gagal"
                
            item_display = f"{trx['item_name']} - {item_qty:,.0f} unit"
            
            transaction_rows += f"""
            <div class="trx-row-item">
                <div class="trx-line-group"><span class="data-field">{trx['order_id']}</span></div>
                <div class="trx-line-group"><span class="data-field">{item_display}</span></div>
                <div class="trx-line-group"><span class="data-field">{format_currency(total_harga)}</span></div>
                <div class="trx-line-group"><span class="data-field">{tanggal_lengkap}</span></div>
                <div class="trx-line-group"><span class="status-badge-trx {status_class}">{status_text}</span></div>
            </div>
            """
    else:
        transaction_rows = f"""
        <div style="text-align:center; padding: 60px 20px; color: white; font-size: 1.2em; font-weight: 600;">
            Informasi tidak ditemukan
        </div>
        """
        
    # 5. HTML Layout Akhir
    content_html = f"""
    {style_css}

    <div class="single-trx-container">
        <div class="dashboard-link-box">
            <a href="/home" class="btn-dashboard">
                <svg class="svg-back" viewBox="0 0 24 24"><polyline points="15 18 9 12 15 6"></polyline></svg>
                Dashboard
            </a>
        </div>
        
        {filter_form_html}

        <div class="trx-data-wrapper">
            <div class="trx-table-content">
                {transaction_rows}
            </div>
        </div>
    </div>
    """

    # Ambil data user dan stats untuk render_consumer_page
    user_data = db.execute("SELECT phone_number FROM users WHERE username = ?", (username,)).fetchone()
    initials = get_initials(username)
    display_phone = format_phone_indo(user_data['phone_number'] if user_data else "")
    
    stats_for_sidebar = db.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as waiting,
            SUM(CASE WHEN status = 'verified' THEN 1 ELSE 0 END) as success,
            SUM(CASE WHEN status = 'rejected' THEN 1 ELSE 0 END) as failed
        FROM online_payments 
        WHERE buyer_name = ?
    """, (username,)).fetchone()
    alert_html = "" 

    return render_consumer_page(status_display, content_html, username, initials, display_phone, stats_for_sidebar, alert_html, active_menu="Transaksi")

@app.route("/consumer/profile-settings", methods=['GET', 'POST'])
def consumer_profile_settings():
    if session.get('role') != 'consumer':
        return redirect(url_for('index'))
    
    username = session.get('username')
    db = get_db()
    
    # Ambil data pengguna saat ini
    user_data = db.execute("SELECT username, phone_number, security_answer FROM users WHERE username = ?", (username,)).fetchone()
    
    # Ambil alert messages
    success_msg = request.args.get('success')
    error_msg = request.args.get('error')
    
    alert_html = ""
    if success_msg:
        alert_html = f"""
        <div style="background: rgba(40, 167, 69, 0.9); color: white; padding: 15px; border-radius: 10px; margin-bottom: 20px; box-shadow: 0 5px 15px rgba(0,0,0,0.1); display: flex; align-items: center; gap: 10px;">
            <span>‚úÖ</span> {success_msg}
        </div>
        """
    elif error_msg:
        alert_html = f"""
        <div style="background: rgba(220, 53, 69, 0.9); color: white; padding: 15px; border-radius: 10px; margin-bottom: 20px; box-shadow: 0 5px 15px rgba(0,0,0,0.1); display: flex; align-items: center; gap: 10px;">
            <span>‚ö†Ô∏è</span> {error_msg}
        </div>
        """

    # --- Logika POST (Update Profil) ---
    if request.method == 'POST':
        new_username = request.form['new_username'].strip()
        new_phone = request.form['new_phone'].strip()
        
        # Validasi (hanya update jika ada perubahan)
        if not new_username:
            return redirect(url_for('consumer_profile_settings', error="Nama tidak boleh kosong."))

        try:
            with db:
                # Update Username dan Phone (Security Answer/Email tidak diubah di sini)
                db.execute("""
                    UPDATE users 
                    SET username = ?, phone_number = ? 
                    WHERE username = ?
                """, (new_username, new_phone, username))
                
                # Update Session jika username berubah
                if new_username != username:
                    session['username'] = new_username

            return redirect(url_for('consumer_profile_settings', success="Profil berhasil diperbarui."))
        
        except Exception as e:
            # Catch error duplikasi username
            return redirect(url_for('consumer_profile_settings', error=f"Gagal update: Username '{new_username}' mungkin sudah digunakan."))
    
    # --- Logika GET (Tampilan Formulir) ---
    
    # Dapatkan data user untuk helper render_consumer_page
    current_username = session.get('username')
    initials = get_initials(current_username)
    display_phone = format_phone_indo(user_data['phone_number'] if user_data else "")
    
    stats = db.execute("SELECT COUNT(*) as total FROM online_payments WHERE buyer_name = ?", (current_username,)).fetchone()
    stats_for_sidebar = {'total': stats['total'], 'waiting': 0, 'success': 0, 'failed': 0}
    
    # HTML Formulir (SUDAH DIBERSIHKAN DARI KODE BELANJA)
    content_html = f"""
    <style>
        .profile-form-container {{
            background: rgba(255, 255, 255, 0.15);
            border: 1px solid rgba(255, 255, 255, 0.4);
            border-radius: 15px;
            padding: 40px;
            color: white;
            max-width: 850px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.05);
        }}
        .form-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 25px 40px;
            margin-bottom: 30px;
        }}
        .form-group {{
            display: flex;
            flex-direction: column;
        }}
        .form-label {{
            font-weight: 600;
            font-size: 1.1em;
            margin-bottom: 8px;
            color: white;
        }}
        .form-input {{
            width: 100%;
            padding: 12px 15px;
            border: none;
            border-radius: 10px;
            background: rgba(108, 194, 183, 0.6); 
            color: white;
            box-sizing: border-box;
            font-weight: 400;
            font-size: 1em;
            outline: none;
        }}
        .form-input::placeholder {{
            color: rgba(255, 255, 255, 0.8);
        }}
        .btn-update {{
            background-color: #4AA0C6; 
            color: white; 
            padding: 12px 30px; 
            border: none;
            border-radius: 10px;
            font-weight: bold;
            font-size: 1em;
            cursor: pointer;
            width: auto;
            min-width: 200px;
            transition: 0.3s;
            box-shadow: 0 4px 10px rgba(0,0,0,0.1);
        }}
        .btn-update:hover {{
            background-color: #388ba8;
            transform: translateY(-2px);
        }}
        .dashboard-btn-container {{
            margin-bottom: 30px;
        }}
        .dashboard-btn-container .btn-dashboard {{
            background: rgba(255, 255, 255, 0.2); 
            color: white; 
            padding: 10px 20px; 
            border-radius: 30px; 
            text-decoration: none;
            font-weight: 500; 
            display: inline-flex; 
            align-items: center; 
            gap: 5px; 
            border: 1px solid rgba(255, 255, 255, 0.5);
            transition: 0.3s;
        }}
        .dashboard-btn-container .btn-dashboard:hover {{
            background: rgba(255, 255, 255, 0.3);
        }}
        .dashboard-btn-container .svg-back {{ width: 20px; height: 20px; fill: none; stroke: currentColor; stroke-width: 2; }}
        
        @media (max-width: 768px) {{
            .form-grid {{ grid-template-columns: 1fr; gap: 20px; }}
        }}
    </style>
    
    <div class="dashboard-btn-container">
        <a href="/home" class="btn-dashboard">
            <svg class="svg-back" viewBox="0 0 24 24"><polyline points="15 18 9 12 15 6"></polyline></svg>
            Dashboard
        </a>
    </div>

    <div class="profile-form-container">
        <h2 style="color: white; margin-top: 0; margin-bottom: 5px;">Profil</h2>
        <p style="color: rgba(255, 255, 255, 0.9); margin-bottom: 30px; margin-top: 0;">Silahkan isi update data profil anda.</p>
        
        <form method="POST">
            <div class="form-grid">
                
                <div class="form-group">
                    <label class="form-label">Nama</label>
                    <input type="text" name="new_username" class="form-input" 
                           value="{current_username}" 
                           placeholder="Isi nama baru anda" required>
                </div>

                <div class="form-group">
                    <label class="form-label">Username</label>
                    <input type="text" name="new_username_confirm" class="form-input" 
                           value="{current_username}" 
                           placeholder="Isi username baru anda" disabled> 
                    </div>
                
                <div class="form-group">
                    <label class="form-label">Email</label>
                    <input type="text" name="email" class="form-input" 
                           value="{user_data['security_answer'] if user_data else ''}" 
                           placeholder="Contoh: nama@gmail.com" disabled>
                </div>
                
                <div class="form-group">
                    <label class="form-label">No. Handphone</label>
                    <input type="text" name="new_phone" class="form-input" 
                           value="{user_data['phone_number'] if user_data and user_data['phone_number'] else ''}" 
                           placeholder="+628XX-XXXX-XXXX">
                </div>
                
            </div>
            
            <button type="submit" class="btn-update">Update Profile</button>
        </form>
    </div>
    """
    
    # Panggil render_consumer_page
    return render_consumer_page("Update Profil", content_html, current_username, initials, display_phone, stats_for_sidebar, alert_html, active_menu="Dashboard")

@app.route("/shop")
def consumer_shop():
    # 1. Cek Akses Consumer
    if session.get('role') != 'consumer':
        return redirect(url_for('index'))

    username = session.get('username', 'Tamu')
    db = get_db()
    
    # 2. Ambil Item yang Dijual & Hitung Stok
    items_for_sale = db.execute("""
        SELECT id, item_name, sale_price_per_unit 
        FROM inventory_items 
        WHERE sales_account IS NOT NULL
    """).fetchall()
    
    available_items = []
    for item in items_for_sale:
        logs = db.execute("SELECT * FROM inventory_log WHERE item_id = ? ORDER BY trx_date, id", (item['id'],)).fetchall()
        available_stock = 0
        for log in logs:
            if log['trx_type'] == 'purchase': available_stock += log['qty']
            elif log['trx_type'] == 'sale': available_stock -= log['qty']
        
        if available_stock > 0:
            available_items.append({
                'id': item['id'],
                'item_name': item['item_name'],
                'price': item['sale_price_per_unit'],
                'stock': available_stock
            })

    # 3. Buat Opsi Dropdown HTML (Didefinisikan SEBELUM HTML Body)
    item_options = ""
    for item in available_items:
        price_fmt = format_currency(item['price']) 
        item_options += f'<option value="{item["id"]}" data-stock="{item["stock"]}" data-price="{item["price"]}" data-name="{item["item_name"]}">{item["item_name"]} - Stok: {item["stock"]} ({price_fmt})</option>'

    # 4. HTML Content (Layout Glassmorphism + Grid Split)
    content_html = f"""
    <style>
        /* --- CSS KHUSUS HALAMAN BELANJA --- */
        
        /* Kontainer Transparan (Glass Effect) */
        .shop-card {{
            background: rgba(255, 255, 255, 0.15);
            border: 1px solid rgba(255, 255, 255, 0.4);
            border-radius: 20px;
            padding: 40px;
            color: white;
            max-width: 950px;
            margin-top: 20px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.05);
        }}
        
        .shop-header-text {{
            font-size: 1.2em;
            margin-bottom: 20px;
            font-weight: 500;
            color: rgba(255,255,255,0.9);
        }}

        /* Grid Layout: Kiri Form (2 Bagian), Kanan Total (1 Bagian) */
        .shop-grid {{
            display: grid;
            grid-template-columns: 2fr 1fr; /* Pembagian kolom */
            gap: 50px;
            align-items: center; /* Rata tengah secara vertikal */
        }}
        
        /* Styling Form Input */
        .form-group {{ margin-bottom: 20px; }}
        
        .form-label {{
            font-weight: 600;
            margin-bottom: 8px;
            display: block;
            font-size: 1em;
            color: white;
        }}
        
        .custom-input, .custom-select {{
            width: 100%;
            padding: 12px 15px;
            border: none;
            border-radius: 10px;
            /* Warna input hijau transparan */
            background: rgba(163, 216, 207, 0.6); 
            color: white;
            font-size: 1em;
            box-sizing: border-box;
            outline: none;
        }}
        .custom-input::placeholder {{ color: rgba(255,255,255,0.7); }}
        .custom-select option {{ background: #5DA090; color: white; }}

        /* Bagian Kanan: Total & Tombol */
        .total-area {{
            text-align: right;
            display: flex;
            flex-direction: column;
            align-items: flex-end;
            justify-content: center;
        }}
        
        .total-label {{
            font-size: 2.2em;
            font-weight: 700;
            color: #00bf63; /* WARNA BARU: Hijau Tajam */
            margin-bottom: 5px;
            line-height: 1;
        }}
        .total-value {{
            font-size: 2.2em;
            font-weight: 700;
            color: #00bf63; /* WARNA BARU: Hijau Tajam */
            margin-bottom: 25px;
        }}
        
        .btn-beli-now {{
            background-color: #4AA0C6; /* Biru */
            color: white;
            padding: 12px 0;
            border-radius: 10px;
            border: none;
            font-weight: bold;
            font-size: 1em;
            cursor: pointer;
            transition: 0.3s;
            width: 100%; /* Tombol memenuhi lebar kolom kanan */
            box-shadow: 0 4px 10px rgba(0,0,0,0.1);
        }}
        .btn-beli-now:hover {{ 
            background-color: #388ba8; 
            transform: translateY(-2px);
        }}

        /* Tombol Dashboard (Back) */
        .btn-back-dashboard {{
            display: inline-flex;
            align-items: center;
            gap: 5px;
            background: rgba(255,255,255,0.2);
            border: 1px solid rgba(255,255,255,0.4);
            padding: 8px 20px;
            border-radius: 30px;
            color: white;
            text-decoration: none;
            margin-bottom: 20px;
            font-weight: 500;
            transition: 0.3s;
        }}
        .btn-back-dashboard:hover {{ background: rgba(255,255,255,0.3); }}
        .btn-back-dashboard svg {{ width: 20px; height: 20px; fill: none; stroke: currentColor; stroke-width: 2; }}
        
        /* Responsive Mobile */
        @media (max-width: 768px) {{
            .shop-grid {{ grid-template-columns: 1fr; gap: 30px; }}
            .total-area {{ align-items: center; text-align: center; margin-top: 20px; }}
            .total-label, .total-value {{ font-size: 1.8em; }}
        }}
        
        /* --- MODAL CSS (Untuk Konfirmasi) --- */
        .modal-overlay {{
            display: none;
            position: fixed; z-index: 9999; left: 0; top: 0;
            width: 100%; height: 100%;
            background-color: rgba(0,0,0,0.6);
            backdrop-filter: blur(4px);
            align-items: center; justify-content: center;
            opacity: 0; transition: opacity 0.3s;
        }}
        .modal-overlay.show {{ display: flex; opacity: 1; }}
        
        .modal-box {{
            background-color: #5DA090;
            width: 400px; max-width: 90%;
            padding: 30px; border-radius: 15px;
            text-align: center; color: white;
            box-shadow: 0 15px 30px rgba(0,0,0,0.3);
            border: 2px solid rgba(255,255,255,0.8);
            transform: scale(0.9); transition: transform 0.3s;
        }}
        .modal-overlay.show .modal-box {{ transform: scale(1); }}
        
        .modal-details {{
            background: rgba(0,0,0,0.1);
            padding: 15px; border-radius: 10px;
            margin: 20px 0; text-align: left; font-size: 0.95em;
        }}
        .detail-row {{ display: flex; justify-content: space-between; margin-bottom: 8px; }}
        .total-row-modal {{ 
            display: flex; justify-content: space-between; 
            margin-top: 10px; padding-top: 10px; 
            border-top: 1px solid rgba(255,255,255,0.3); 
            font-weight: bold; font-size: 1.1em;
        }}
        
        .btn-modal {{
            width: 100%; padding: 12px; border: none; border-radius: 8px;
            font-weight: bold; cursor: pointer; margin-top: 10px;
        }}
        .btn-confirm {{ background-color: #4AA0C6; color: white; }}
        .btn-cancel {{ background-color: #FF6B6B; color: white; }}
        
    </style>

    <div>
        <a href="/home" class="btn-back-dashboard">
            <svg viewBox="0 0 24 24"><polyline points="15 18 9 12 15 6"></polyline></svg>
            Dashboard
        </a>
        
        <div class="shop-header-text">Silahkan input data untuk melakukan pembelian</div>

        <div class="shop-card">
            <form id="purchaseForm" action="/purchase" method="POST">
                <div class="shop-grid">
                    
                    <div class="form-section">
                        <div class="form-group">
                            <label class="form-label">Pilih Barang</label>
                            <select id="item_id" name="item_id" onchange="updateForm()" class="custom-select">
                                <option value="" data-price="0" data-stock="0">Silahkan pilih item yang akan dibeli</option>
                                {item_options}
                            </select>
                        </div>
                        
                        <div class="form-group">
                            <label class="form-label">Kuantitas (max <span id="maxQty">0</span>)</label>
                            <input type="number" id="qty" name="qty" value="1" min="1" 
                                   oninput="calculateTotal()" class="custom-input" placeholder="Jumlah beli">
                        </div>
                        
                        <div class="form-group">
                            <label class="form-label">Metode Pembayaran</label>
                            <select id="payment_method" name="payment_method" onchange="updateForm()" class="custom-select">
                                <option value="" disabled selected>Silahkan isi dengan apa anda membayar</option>
                                
                                <option value="cash">Tunai</option>
                                
                                <option value="credit">Kredit</option>
                                <option value="qris">QRIS (Scan Barcode)</option>
                            </select>
                        </div>
                    </div>
                    
                    <div class="total-area">
                        <div class="total-label">Total:</div>
                        <div id="totalPriceDisplay" class="total-value">Rp 0,00</div>
                        
                        <button type="button" onclick="showModal()" class="btn-beli-now">Beli Sekarang</button>
                    </div>
                </div>
            </form>
        </div>
    </div>

    <div id="confirmModal" class="modal-overlay">
        <div class="modal-box">
            <div style="font-size: 50px; margin-bottom: 10px;">‚úì</div>
            <h3 style="margin: 0;">KONFIRMASI PESANAN</h3>
            <p style="margin: 5px 0 20px; opacity: 0.9;">Pastikan data pesanan kamu valid</p>
            
            <div class="modal-details">
                <div class="detail-row"><span>Username</span> <span>{username}</span></div>
                <div class="detail-row"><span>Item</span> <span id="modalItem">-</span></div>
                <div class="detail-row"><span>Qty</span> <span id="modalQty">0 unit</span></div>
                <div class="detail-row"><span>Pembayaran</span> <span id="modalPayment">-</span></div>
                
                <div class="total-row-modal">
                    <span>Total Bayar</span> <span id="modalTotal">Rp 0,00</span>
                </div>
            </div>
            
            <button onclick="submitRealForm()" class="btn-modal btn-confirm">Beli Sekarang</button>
            <button onclick="closeModal()" class="btn-modal btn-cancel">Batal</button>
        </div>
    </div>

    <script>
        // Format Rupiah
        function formatRupiah(angka) {{
            return new Intl.NumberFormat('id-ID', {{ style: 'currency', currency: 'IDR' }}).format(angka);
        }}

        function updateForm() {{
            const itemSelect = document.getElementById('item_id');
            const selectedOption = itemSelect.options[itemSelect.selectedIndex];
            const stock = parseInt(selectedOption.getAttribute('data-stock')) || 0;
            
            document.getElementById('maxQty').innerText = stock;
            document.getElementById('qty').max = stock;
            calculateTotal();
        }}

        function calculateTotal() {{
            const itemSelect = document.getElementById('item_id');
            const selectedOption = itemSelect.options[itemSelect.selectedIndex];
            const price = parseFloat(selectedOption.getAttribute('data-price')) || 0;
            const qty = parseFloat(document.getElementById('qty').value) || 0;
            
            const total = price * qty;
            document.getElementById('totalPriceDisplay').innerText = formatRupiah(total);
            return total;
        }}

        function showModal() {{
            const itemSelect = document.getElementById('item_id');
            const selectedOption = itemSelect.options[itemSelect.selectedIndex];
            const qty = document.getElementById('qty').value;
            const paySelect = document.getElementById('payment_method');
            const payText = paySelect.options[paySelect.selectedIndex].text;
            
            if (!itemSelect.value || qty <= 0) {{
                alert("Harap pilih barang dan kuantitas yang valid!");
                return;
            }}
            
            // Isi data modal
            document.getElementById('modalItem').innerText = selectedOption.getAttribute('data-name');
            document.getElementById('modalQty').innerText = qty + " unit";
            document.getElementById('modalPayment').innerText = payText;
            document.getElementById('modalTotal').innerText = document.getElementById('totalPriceDisplay').innerText;
            
            // Tampilkan modal dengan animasi
            const modal = document.getElementById('confirmModal');
            modal.classList.add('show');
        }}

        function closeModal() {{
            document.getElementById('confirmModal').classList.remove('show');
        }}
        
        function submitRealForm() {{
            const paymentMethod = document.getElementById('payment_method').value;
            const form = document.getElementById('purchaseForm');
            
            if (paymentMethod === 'qris') form.action = "/purchase-manual-qris";
            else form.action = "/purchase";
            
            form.submit();
        }}
        
        // Tutup modal jika klik di luar box
        window.onclick = function(event) {{
            const modal = document.getElementById('confirmModal');
            if (event.target == modal) closeModal();
        }}
        
        // Init
        updateForm();
    </script>
    """

    # 5. Persiapan Data Sidebar (Query Statistik Lengkap)
    user_data = db.execute("SELECT phone_number FROM users WHERE username = ?", (username,)).fetchone()
    initials = get_initials(username)
    display_phone = format_phone_indo(user_data['phone_number'] if user_data else "")
    
    # Query Statistik lengkap (agar tidak IndexError)
    stats_for_sidebar = db.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as waiting,
            SUM(CASE WHEN status = 'verified' THEN 1 ELSE 0 END) as success,
            SUM(CASE WHEN status = 'rejected' THEN 1 ELSE 0 END) as failed
        FROM online_payments 
        WHERE buyer_name = ?
    """, (username,)).fetchone()
    
    alert_html = ""
    
    # 6. Render Menggunakan 'render_consumer_page' (Bukan render_page!)
    return render_consumer_page("Belanja", content_html, username, initials, display_phone, stats_for_sidebar, alert_html, active_menu="Belanja")

@app.route("/purchase", methods=['POST'])
def purchase():
    #1. Cek Hak Akses
    if session.get('role') != 'consumer':
        return redirect(url_for('index'))

    db = get_db()
    
    #2. Ambil Data dari Form
    item_id = request.form.get('item_id')
    qty_str = request.form.get('qty')
    payment_method = request.form.get('payment_method')
    username = session.get('username')
    
    #Validasi Input Kosong
    if not item_id or not qty_str:
        return redirect(url_for('consumer_shop', error="Data tidak lengkap."))
        
    qty = float(qty_str)
    
     #3. Validasi Item & Stok
    item = db.execute("SELECT * FROM inventory_items WHERE id = ?", (item_id,)).fetchone()
    
    # [PENYESUAIAN] Jika item salah, kembali ke SHOP (bukan home)
    if not item: 
        return redirect(url_for('consumer_shop', error="Barang tidak ditemukan."))
    
    logs = db.execute("SELECT * FROM inventory_log WHERE item_id = ?", (item_id,)).fetchall()
    stock = sum(l['qty'] if l['trx_type'] == 'purchase' else -l['qty'] for l in logs)
    
    # [PENYESUAIAN] Jika stok kurang, kembali ke SHOP dengan pesan error
    if qty > stock:
        return redirect(url_for('consumer_shop', error=f"Stok tidak cukup! Sisa: {stock}"))

    total_price = qty * item['sale_price_per_unit']

    # 4. Proses Transaksi
    # (QRIS Manual sudah diarahkan JS ke /purchase-manual-qris, jadi di sini hanya Tunai/Kredit)
    if payment_method in ['cash', 'credit']:
        success = record_sale_transaction(db, item, qty, username, payment_method, total_price)
        if success:
            #[PENYESUAIAN] Redirect ke DASHBOARD (home) dengan pesan sukses
            return redirect(url_for('consumer_home', success="Pembelian berhasil! Data transaksi telah diperbarui."))
    
    # Jika gagal
    return redirect(url_for('consumer_shop', error="Metode pembayaran tidak valid atau gagal diproses."))

# --- Helper Function Baru (Letakkan di luar route) ---
def record_sale_transaction(db, item, qty, username, payment_method, total_revenue):
    """Fungsi ini mencatat Stok Keluar dan Jurnal Akuntansi"""
    
    description = f"Penjualan {item['item_name']} kepada {username} ({payment_method})"
    asset_account = '1101' if (payment_method == 'cash' or payment_method == 'qris') else '1102'
    
    trx_datetime = datetime.now()
    trx_date = trx_datetime.strftime('%Y-%m-%d')
    trx_timestamp_str = trx_datetime.strftime('%Y-%m-%d %H:%M:%S')

    # A. Hitung HPP (FIFO/Average logic yang sudah ada)
    logs = db.execute("SELECT * FROM inventory_log WHERE item_id = ? ORDER BY trx_date, id", (item['id'],)).fetchall()
    average_stack = []
    for log in logs:
        if log['trx_type'] == 'purchase':
            average_stack.append([log['qty'], log['cost_per_unit']])
        elif log['trx_type'] == 'sale':
            qty_to_sell_hist = log['qty']
            while qty_to_sell_hist > 0 and average_stack:
                if average_stack[0][0] <= qty_to_sell_hist:
                    qty_to_sell_hist -= average_stack.pop(0)[0]
                else:
                    average_stack[0][0] -= qty_to_sell_hist
                    qty_to_sell_hist = 0
    
    # Hitung HPP transaksi ini
    temp_qty = qty
    total_cogs = 0.0
    while temp_qty > 0 and average_stack:
        if average_stack[0][0] <= temp_qty:
            qty_taken = average_stack[0][0]
            cost_taken = average_stack[0][1]
            total_cogs += (qty_taken * cost_taken)
            temp_qty -= qty_taken
            average_stack.pop(0)
        else:
            total_cogs += (temp_qty * average_stack[0][1])
            average_stack[0][0] -= temp_qty
            temp_qty = 0

    with db:
        # 1. Catat Stok Keluar
        db.execute(
            "INSERT INTO inventory_log (item_id, trx_date, trx_type, description, qty, sale_price_per_unit) VALUES (?, ?, ?, ?, ?, ?)",
            (item['id'], trx_date, 'sale', description, qty, item['sale_price_per_unit'])
        )
        
        # 2. Jurnal Pendapatan
        j_code_1 = generate_journal_code(db, trx_datetime)
        cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)", (j_code_1, trx_timestamp_str, description))
        entry_id_1 = cursor.lastrowid
        db.execute("INSERT INTO journal_details (entry_id, account_code, debit, credit) VALUES (?, ?, ?, ?)", (entry_id_1, asset_account, total_revenue, 0))
        db.execute("INSERT INTO journal_details (entry_id, account_code, debit, credit) VALUES (?, ?, ?, ?)", (entry_id_1, item['sales_account'], 0, total_revenue))
        
        # 3. Jurnal HPP
        j_code_2 = generate_journal_code(db, trx_datetime)
        cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)", (j_code_2, trx_timestamp_str, description + " (HPP)"))
        entry_id_2 = cursor.lastrowid
        db.execute("INSERT INTO journal_details (entry_id, account_code, debit, credit) VALUES (?, ?, ?, ?)", (entry_id_2, item['expense_cogs_account'], total_cogs, 0))
        db.execute("INSERT INTO journal_details (entry_id, account_code, debit, credit) VALUES (?, ?, ?, ?)", (entry_id_2, item['inventory_account'], 0, total_cogs))

    return True

# =========================
# ROUTE ADMIN
# =========================
def get_admin_sidebar_html():
    """
    Menu Sidebar Admin.
    Ditambahkan link 'Main Dashboard' agar user bisa kembali ke Home.
    """
    return """
    <h3>Home</h3>
    <ul>
        <li><a href="/admin" style="color: #ffdd57; font-weight: bold !important;">üè† Main Dashboard</a></li>
    </ul>
    
    <p><b>Pembayaran QRIS</b></p> 
    <ul>
        <li><a href="/admin/verify-payments" style="color: #ff6b6b;">Cek Transfer</a></li>
    </ul>

    <p><b>Data Awal</b></p>
    <ul>
        <li><a href="/admin/chart-of-accounts">Daftar Akun</a></li>
        <li><a href="/admin/opening-balance">Neraca Saldo Awal</a></li>
        <li><a href="/admin/transactions">Riwayat Transaksi</a></li>
    </ul>
    
    <p><b>Entri Jurnal</b></p>
    <ul>
        <li><a href="/admin/general-journal">Jurnal Umum</a></li>
        <li><a href="/admin/adjusting-entries">Jurnal Penyesuaian</a></li>
        <li><a href="/admin/closing-entries">Jurnal Penutup</a></li>
    </ul>
    
    <p><b>Persediaan</b></p>
    <ul>
        <li><a href="/admin/inventory-journal">Inventory Card</a></li>
    </ul>

    <p><b>Buku Besar</b></p>
    <ul>
        <li><a href="/admin/ledger">Buku Besar</a></li>
        <li><a href="/admin/ledger-ar">Buku Pembantu Piutang</a></li>
        <li><a href="/admin/ledger-ap">Buku Pembantu Utang</a></li>
    </ul>
    
    <p><b>Neraca</b></p>
    <ul>
        <li><a href="/admin/trial-balance">Neraca Setelah Buku Besar</a></li>
        <li><a href="/admin/work-sheet">Neraca Lajur</a></li>
    </ul>

    <p><b>Laporan Keuangan</b></p>
    <ul>
        <li><a href="/admin/income-statement">Laporan Laba Rugi</a></li>
        <li><a href="/admin/equity-change">Laporan Perubahan Modal</a></li>
        <li><a href="/admin/financial-position">Laporan Posisi Keuangan</a></li>
    </ul>
    """
# =========================
# DASHBOARD ADMIN
# =========================
@app.route("/admin")
def admin_dashboard():
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    sidebar = get_admin_sidebar_html()
    
    body = f"""
    <style>
        /* HILANGKAN JUDUL BAWAAN */
        h1 {{ display: none !important; }}

        /* --- 1. BACKGROUND IMAGE (FIXED) --- */
        .hero-bg {{
            position: fixed;
            top: 0;
            left: 0;
            width: 100vw;
            height: 100vh;
            background-image: url('/static/dashboard_bg.jpeg'); 
            background-size: cover;
            background-position: center;
            z-index: -5; 
        }}

        /* --- 2. OVERLAY GELAP (FIXED) --- */
        .hero-overlay {{
            position: fixed;
            top: 0;
            left: 0;
            width: 100vw;
            height: 100vh;
            background: rgba(0, 0, 0, 0.3);
            z-index: -4;
        }}

        /* --- 3. CONTAINER TEKS (DINAMIS) --- */
        .hero-text-container {{
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            height: 100vh;
            margin-top: -80px; 
            margin-left: -20px;
            margin-right: -20px;
            width: calc(100% + 40px);
        }}

        /* --- 4. MODIFIKASI NAVBAR LOGOUT --- */
        .navbar {{
            position: relative;
            z-index: 10;
        }}
        .btn-auth {{
            color: white !important;
            text-shadow: 0 1px 3px rgba(0,0,0,0.5);
        }}
        /* SAYA MENGHAPUS CSS .icon-arrow DI SINI AGAR MENGIKUTI GLOBAL STYLE */

        /* --- 5. STYLE TYPOGRAPHY --- */
        .hero-title {{
            font-family: 'Segoe UI', sans-serif;
            font-size: 6em; 
            font-weight: 800; 
            color: white;
            margin: 0;
            letter-spacing: 5px;
            text-shadow: 2px 2px 15px rgba(0,0,0,0.5);
            line-height: 1;
            text-align: center;
            opacity: 0;
            transform: translateY(-80px);
            animation: dropDown 1.2s cubic-bezier(0.2, 0.8, 0.2, 1) forwards;
        }}

        .hero-subtitle {{
            font-family: 'Segoe UI', sans-serif;
            font-size: 2em;
            font-weight: 300;
            color: white;
            margin-top: 15px;
            letter-spacing: 4px;
            text-transform: uppercase;
            text-align: center;
            opacity: 0;
            transform: translateY(-50px);
            animation: dropDown 1.2s cubic-bezier(0.2, 0.8, 0.2, 1) forwards;
            animation-delay: 0.3s; 
        }}

        @keyframes dropDown {{
            0% {{ opacity: 0; transform: translateY(-80px); }}
            100% {{ opacity: 1; transform: translateY(0); }}
        }}
        
        @media (max-width: 768px) {{
            .hero-title {{ font-size: 3em; }}
            .hero-subtitle {{ font-size: 1.2em; }}
        }}
    </style>

    <div class="hero-bg"></div>
    <div class="hero-overlay"></div>

    <div class="hero-text-container">
        <div class="hero-title">SIGRAMEE</div>
        <div class="hero-subtitle">DASHBOARD KEUANGAN</div>
    </div>
    """
    
    return render_page("", body, sidebar_content=sidebar)

# ==================================
# DAFTAR AKUN
# ==================================
@app.route("/admin/chart-of-accounts")
def chart_of_accounts():
    """Menampilkan daftar akun dengan checkbox dan tombol."""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
        
    db = get_db()
    accounts = db.execute("SELECT id, account_code, account_name, account_type FROM chart_of_accounts ORDER BY account_code").fetchall()
    
    # Dapatkan pesan error jika ada (dari redirect)
    error = request.args.get('error')
    
    table_rows = ""
    for acc in accounts:
        # Checkbox menggunakan 'account_code' sebagai nilainya
        table_rows += f"""
        <tr>
            <td class="col-pilih"><input type="checkbox" name="selected_codes" value="{acc['account_code']}"></td>
            <td>{acc['account_code']}</td>
            <td>{acc['account_name']}</td>
            <td>{acc['account_type']}</td>
        </tr>
        """
        
    body = f"""
    <div style="background: #f0f8ff; padding: 15px; border-left: 4px solid #007bff; margin-bottom: 20px; font-size: 14px; border-radius: 4px;">
        <strong>üìã Panduan Penomoran Akun Aset Tetap:</strong>
        <ul style="margin: 10px 0; padding-left: 20px;">
            <li><strong>Aset Tetap</strong>: 1 (Aset) + 2 (Tetap) + XX Ganjil (contoh: <code>1201</code> Peralatan, <code>1203</code> Kendaraan)</li>
            <li><strong>Akumulasi Penyusutan</strong>: 1 + 2 + XX Genap (contoh: <code>1202</code> Akumulasi Penyusutan Peralatan)</li>
            <li><strong>Beban Penyusutan</strong>: 5 (Beban) + 2 (Penyusutan) + XX (contoh: <code>5201</code> Biaya Penyusutan Peralatan)</li>
        </ul>
    </div>

    <form action="/admin/delete-accounts" method="POST">
        <table class="table-condensed">
            <thead>
                <tr>
                    <th class="col-pilih">Pilih</th>
                    <th>Kode Akun</th>
                    <th>Nama Akun</th>
                    <th>Tipe Akun</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
        <br>
        <input type="submit" value="Hapus Akun Terpilih" class="btn-red" 
               onclick="return confirm('PERINGATAN: Akun yang sudah digunakan dalam jurnal tidak akan bisa dihapus. Lanjutkan?');">
        <a href="/admin/add-account">
            <input type="button" value="Input Akun Manual" class="btn-blue">
        </a>
    </form>
    """

    # Kirim pesan error ke render_page jika ada
    return render_page("Daftar Akun", body, sidebar_content=get_admin_sidebar_html(), error_message=error)

@app.route("/admin/transactions")
def transactions_list():
    """Menampilkan daftar transaksi dengan 'Nomor' dan 'ID Jurnal' baru."""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
        
    db = get_db()
    
    # --- PERUBAHAN DI SINI: Tambahkan 'WHERE id != 1' ---
    # Ini untuk menyembunyikan Saldo Awal dari daftar transaksi
    entries = db.execute(
        "SELECT id, journal_code, entry_timestamp, description FROM journal_entries WHERE id != 1 ORDER BY id ASC"
    ).fetchall()
    # --- AKHIR PERUBAHAN ---
    
    table_rows = ""
    for entry in entries:
        entry_date, entry_time = entry['entry_timestamp'].split(' ', 1)
        
        table_rows += f"""
        <tr>
            <td class="col-pilih"><input type="checkbox" name="selected_ids" value="{entry['id']}"></td>
            <td class="col-nomor">{entry['id']}</td>
            <td>{entry['journal_code']}</td>
            <td>{entry_date}</td>      
            <td>{entry_time}</td>      
            <td>{entry['description']}</td>
        </tr>
        """
        
    body = f"""
    
    <form action="/admin/delete-transactions" method="POST">
        <table class="table-condensed">
            <thead>
                <tr>
                    <th class="col-pilih">Pilih</th>
                    <th class="col-nomor">Nomor</th>
                    <th>ID Jurnal</th>
                    <th>Tanggal</th>
                    <th>Jam</th>
                    <th>Deskripsi</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>
        <br>
        <input type="submit" value="Hapus Transaksi Terpilih" class="btn-red"
               onclick="return confirm('Anda yakin ingin menghapus transaksi terpilih? Ini akan menghapus header dan semua detail jurnalnya.');">
    </form>
    """
    return render_page("Daftar Transaksi", body, sidebar_content=get_admin_sidebar_html())

@app.route("/admin/add-account", methods=['GET', 'POST'])
def add_account():
    """Formulir untuk menambah akun baru secara manual."""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        # Proses form
        code = request.form['account_code']
        name = request.form['account_name']
        tipe = request.form['account_type']
        
        try:
            db = get_db()
            with db:
                db.execute("INSERT INTO chart_of_accounts (account_code, account_name, account_type) VALUES (?, ?, ?)",
                           (code, name, tipe))
            # Jika sukses, kembali ke daftar akun
            return redirect(url_for('chart_of_accounts'))
        except sqlite3.IntegrityError:
            # Jika gagal (misal: kode duplikat)
            error = f"Gagal menambah akun. Kode Akun '{code}' mungkin sudah ada."
            # Tampilkan form lagi dengan pesan error
            return render_add_account_form(error)
        
    # Jika method GET, tampilkan form
    return render_add_account_form()

@app.route("/admin/delete-accounts", methods=['POST'])
def delete_accounts():
    """Memproses penghapusan akun yang dipilih."""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
        
    codes_to_delete = request.form.getlist('selected_codes')
    db = get_db()
    error_msg = None
    
    # Kita gunakan blok try-except di luar loop untuk keamanan transaksi,
    # atau di dalam loop jika ingin menghapus yang bisa dihapus saja.
    # Di sini saya gunakan pendekatan: Coba hapus satu per satu.
    
    try:
        with db:
            for code in codes_to_delete:
                try:
                    db.execute("DELETE FROM chart_of_accounts WHERE account_code = ?", (code,))
                except sqlite3.IntegrityError:
                    # Tangkap error jika akun sedang digunakan di jurnal
                    # Kita simpan pesan error, tapi loop tetap lanjut (biar akun lain tetap terhapus jika dipilih banyak)
                    error_msg = f"Gagal menghapus akun {code}: Akun ini sudah digunakan dalam jurnal transaksi."
    except Exception as e:
        # Error umum lain (misal koneksi db)
        error_msg = f"Terjadi kesalahan sistem: {str(e)}"

    # === BAGIAN PENTING YANG SEBELUMNYA HILANG ===
    
    if error_msg:
        # Jika ada error, kembali ke halaman akun dengan membawa pesan error di URL
        return redirect(url_for('chart_of_accounts', error=error_msg))
    else:
        # Jika sukses tanpa error, kembali bersih
        return redirect(url_for('chart_of_accounts'))
    
@app.route("/admin/opening-balance", methods=['GET', 'POST'])
def opening_balance():
    """Halaman untuk menginput neraca saldo awal & Info Perusahaan secara manual"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    error_message = None
    success_message = None
    
    # 1. Ambil data perusahaan saat ini untuk ditampilkan
    try:
        company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    except:
        company_info = None # Handle jika tabel belum ada

    # ==========================================
    # LOGIKA PENYIMPANAN (POST)
    # ==========================================
    if request.method == 'POST':
        try:
            with db:
                # A. SIMPAN INFORMASI PERUSAHAAN
                nama_perusahaan = request.form.get('company_name', '').strip()
                periode_akuntansi = request.form.get('accounting_period', '').strip()
                
                # Cek apakah tabel company_info ada, jika tidak buat (opsional/safety)
                db.execute("CREATE TABLE IF NOT EXISTS company_info (id INTEGER PRIMARY KEY, company_name TEXT, accounting_period TEXT)")
                
                existing_company = db.execute("SELECT id FROM company_info LIMIT 1").fetchone()
                if existing_company:
                    db.execute("UPDATE company_info SET company_name = ?, accounting_period = ? WHERE id = ?", 
                               (nama_perusahaan, periode_akuntansi, existing_company['id']))
                else:
                    db.execute("INSERT INTO company_info (company_name, accounting_period) VALUES (?, ?)", 
                               (nama_perusahaan, periode_akuntansi))

                # B. SIMPAN JURNAL SALDO AWAL
                # 1. Hapus Saldo Awal lama jika ada
                saldo_awal_entry = db.execute("SELECT id FROM journal_entries WHERE description LIKE '%Saldo Awal%' LIMIT 1").fetchone()
                entry_id = None
                
                if saldo_awal_entry:
                    db.execute("DELETE FROM journal_details WHERE entry_id = ?", (saldo_awal_entry['id'],))
                    entry_id = saldo_awal_entry['id']
                else:
                    # Buat Header Jurnal Baru
                    journal_code = generate_journal_code(db, datetime.strptime('2025-01-01', '%Y-%m-%d'))
                    cursor = db.execute(
                        "INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                        (journal_code, '2025-01-01 00:00:00', "Saldo Awal per 1 Jan 2025")
                    )
                    entry_id = cursor.lastrowid
                
                total_debit = Decimal('0')
                total_credit = Decimal('0')
                
                # Tampung nilai aset tetap dulu untuk perhitungan server-side (backup jika JS gagal)
                aset_tetap_values = {}
                for key, value in request.form.items():
                    if key.startswith('balance_'):
                        code = key.replace('balance_', '')
                        val = Decimal(str(float(value))) if value and value.strip() else Decimal('0')
                        if code in ['1201', '1203', '1205', '1207']: # Kode Aset Induk
                            aset_tetap_values[code] = val

                # Loop semua input balance
                for key, value in request.form.items():
                    if key.startswith('balance_'):
                        account_code = key.replace('balance_', '')
                        balance_value = Decimal(str(float(value))) if value and value.strip() else Decimal('0')
                        
                        # Cek tipe akun untuk normal balance
                        account_info = db.execute("SELECT account_type, account_name FROM chart_of_accounts WHERE account_code = ?", (account_code,)).fetchone()
                        
                        if account_info and balance_value != 0:
                            account_type = account_info['account_type']
                            account_name = account_info['account_name'].lower()
                            
                            debit_val = 0.0
                            credit_val = 0.0

                            # Logika Debit/Kredit Berdasarkan Tipe Akun
                            is_contra_asset = 'akumulasi' in account_name and 'penyusutan' in account_name
                            
                            # Kelompok DEBIT (Normal)
                            if account_type in ('Aset Lancar', 'Aset Tetap', 'Beban') and not is_contra_asset:
                                if balance_value > 0: # Positif masuk Debit
                                    debit_val = float(balance_value)
                                else: # Negatif masuk Kredit
                                    credit_val = float(abs(balance_value))
                            
                            # Kelompok KREDIT (Normal)
                            else: 
                                if balance_value > 0: # Positif masuk Kredit
                                    credit_val = float(balance_value)
                                else: # Negatif masuk Debit
                                    debit_val = float(abs(balance_value))

                            # Insert Detail
                            if debit_val > 0 or credit_val > 0:
                                db.execute("INSERT INTO journal_details (entry_id, account_code, debit, credit) VALUES (?, ?, ?, ?)",
                                           (entry_id, account_code, debit_val, credit_val))
                                total_debit += Decimal(str(debit_val))
                                total_credit += Decimal(str(credit_val))
                
                if total_debit != total_credit:
                    error_message = f"Neraca Saldo Awal tidak balance! Debit: {format_currency(total_debit)}, Kredit: {format_currency(total_credit)}"
                    db.rollback()
                else:
                    success_message = f"Data Perusahaan & Saldo Awal berhasil disimpan! Total: {format_currency(total_debit)}"
                    # Refresh data perusahaan untuk tampilan
                    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
                    
        except Exception as e:
            error_message = f"Terjadi kesalahan: {str(e)}"
            import traceback
            print(traceback.format_exc())
            try: db.rollback()
            except: pass

    # ==========================================
    # LOGIKA TAMPILAN (GET)
    # ==========================================
    try:
        accounts = db.execute("""
            SELECT coa.account_code, coa.account_name, coa.account_type
            FROM chart_of_accounts coa
            ORDER BY 
                CASE coa.account_type
                    WHEN 'Aset Lancar' THEN 1
                    WHEN 'Aset Tetap' THEN 2
                    WHEN 'Liabilitas' THEN 3
                    WHEN 'Ekuitas' THEN 4
                    WHEN 'Pendapatan' THEN 5
                    WHEN 'Beban' THEN 6
                    ELSE 7
                END, coa.account_code
        """).fetchall()
    except:
        accounts = []

    # --- 1. FORM DATA PERUSAHAAN ---
    company_name_val = company_info['company_name'] if company_info else ''
    company_period_val = company_info['accounting_period'] if company_info else ''
    
    company_form_html = f"""
    <div style="background-color: #e7f3ff; padding: 20px; border-radius: 5px; margin-bottom: 30px; border-left: 4px solid #007bff;">
        <h3 style="margin-top: 0; color: #0056b3;">üè¢ Informasi Perusahaan</h3>
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
            <div>
                <label style="display: block; margin-bottom: 5px; font-weight: bold;">Nama Perusahaan:</label>
                <input type="text" name="company_name" value="{company_name_val}" 
                       placeholder="Contoh: PT Maju Jaya" required
                       style="width: 100%; padding: 10px; border: 1px solid #ccc; border-radius: 4px;">
            </div>
            <div>
                <label style="display: block; margin-bottom: 5px; font-weight: bold;">Periode Akuntansi:</label>
                <input type="text" name="accounting_period" value="{company_period_val}" 
                       placeholder="Contoh: 2025" required
                       style="width: 100%; padding: 10px; border: 1px solid #ccc; border-radius: 4px;">
            </div>
        </div>
        <small style="color: #666;">*Data ini akan ditampilkan pada kop laporan keuangan.</small>
    </div>
    """

    # --- 2. LOOP AKUN UNTUK FORM ---
    account_inputs = ""
    current_type = ""
    
    # Mapping Aset ke Akumulasi (Code Aset: Data Config)
    aset_penyusutan_map = {
        '1201': {'nama': 'Peralatan', 'akumulasi': '1202', 'default_masa_manfaat': 5},
        '1203': {'nama': 'Kendaraan', 'akumulasi': '1204', 'default_masa_manfaat': 8},
        '1205': {'nama': 'Bangunan', 'akumulasi': '1206', 'default_masa_manfaat': 20},
        '1301': {'nama': 'Tanah', 'akumulasi': None, 'default_masa_manfaat': 0}
    }
    
    for account in accounts:
        if account['account_type'] != current_type:
            if current_type != "":
                account_inputs += "</tbody></table><br>"
            
            # Judul Kategori
            account_inputs += f"""
            <h4 style="border-bottom: 2px solid #ddd; padding-bottom: 5px; margin-top: 20px;">{account['account_type']}</h4>
            <table style="width: 100%; border-collapse: collapse; margin-bottom: 10px;">
                <thead>
                    <tr style="background-color: #f8f9fa; color: #495057;">
                        <th style="padding: 10px; border: 1px solid #dee2e6; width: 15%;">Kode</th>
                        <th style="padding: 10px; border: 1px solid #dee2e6; width: 35%;">Nama Akun</th>
                        <th style="padding: 10px; border: 1px solid #dee2e6; width: 25%;">Saldo Awal</th>
                        <th style="padding: 10px; border: 1px solid #dee2e6; width: 25%;">Keterangan</th>
                    </tr>
                </thead>
                <tbody>
            """
            current_type = account['account_type']

        # --- MULAI PENGGANTIAN ---
        
        # 1. Tentukan Default Saldo Normal (Debit vs Kredit)
        if account['account_type'] in ('Aset Lancar', 'Aset Tetap', 'Beban'):
            # Grup Debit (Warna Hijau)
            note_text = "Normal: <span style='color: #28a745; font-weight: bold;'>Debit</span>"
        else:
            # Grup Kredit: Liabilitas, Ekuitas, Pendapatan (Warna Merah)
            note_text = "Normal: <span style='color: #dc3545; font-weight: bold;'>Kredit</span>"

        # 2. Styling Khusus & Pengecualian (Kontra Aset)
        row_style = ""
        
        if 'akumulasi' in account['account_name'].lower() and 'penyusutan' in account['account_name'].lower():
            row_style = "background-color: #fff3cd;" # Background Kuning
            # Akumulasi adalah Aset tapi saldonya Kredit (Kontra)
            note_text = "‚ö†Ô∏è Kontra Aset (<span style='color: #dc3545; font-weight: bold;'>Kredit</span>)"
            
        elif account['account_type'] == 'Beban':
            row_style = "background-color: #fbfcfd;" # Background Putih Abu sedikit biar beda

        # 3. Cek fitur hitung otomatis untuk Javascript
        code = account['account_code']
        is_asset_parent = code in aset_penyusutan_map
        has_calc_feature = is_asset_parent and aset_penyusutan_map[code]['akumulasi'] is not None
        
        # --- AKHIR PENGGANTIAN ---
            
        # Cek apakah ini Aset Tetap yang punya fitur hitung otomatis
        code = account['account_code']
        is_asset_parent = code in aset_penyusutan_map
        has_calc_feature = is_asset_parent and aset_penyusutan_map[code]['akumulasi'] is not None

        # Input Field Akun Utama
        account_inputs += f"""
        <tr style="{row_style}">
            <td style="padding: 8px; border: 1px solid #dee2e6; text-align: center;">{code}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6;">{account['account_name']}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6;">
                <input type="number" step="0.01" 
                       name="balance_{code}" 
                       id="balance_{code}"
                       placeholder="0"
                       style="width: 100%; padding: 6px; border: 1px solid #ccc; border-radius: 4px;"
                       {'onkeyup="hitungPenyusutan(\'' + code + '\')"' if has_calc_feature else ''}>
            </td>
            <td style="padding: 8px; border: 1px solid #dee2e6; font-size: 0.85em; color: #666;">
                {note_text}
            </td>
        </tr>
        """
        
        # Input Field Tambahan (Masa Manfaat & Tahun Pakai) -> Muncul di bawah Aset Induk
        if has_calc_feature:
            default_life = aset_penyusutan_map[code]['default_masa_manfaat']
            account_inputs += f"""
            <tr style="background-color: #fdfdfe;">
                <td colspan="2" style="text-align: right; padding: 8px; border-bottom: 1px solid #dee2e6; color: #007bff;">
                    <small><strong>‚öôÔ∏è Kalkulator Penyusutan:</strong></small>
                </td>
                <td colspan="2" style="padding: 8px; border-bottom: 1px solid #dee2e6;">
                    <div style="display: flex; gap: 10px; align-items: center;">
                        <div style="flex: 1;">
                            <small>Masa Manfaat (Thn):</small>
                            <input type="number" name="masa_manfaat_{code}" id="masa_manfaat_{code}" 
                                   value="{default_life}" min="1" 
                                   style="width: 100%; padding: 4px; border: 1px solid #ddd;"
                                   onkeyup="hitungPenyusutan('{code}')" onchange="hitungPenyusutan('{code}')">
                        </div>
                        <div style="flex: 1;">
                            <small>Tahun Pakai:</small>
                            <input type="number" name="tahun_pakai_{code}" id="tahun_pakai_{code}" 
                                   value="0" min="0" 
                                   style="width: 100%; padding: 4px; border: 1px solid #ddd;"
                                   onkeyup="hitungPenyusutan('{code}')" onchange="hitungPenyusutan('{code}')">
                        </div>
                    </div>
                    <div id="info_text_{code}" style="margin-top: 5px; font-size: 0.85em; color: #28a745; font-weight: bold;"></div>
                </td>
            </tr>
            """
    
    account_inputs += "</tbody></table>"

    # --- 3. TABLE DATA TERSIMPAN (PREVIEW) ---
    # (Opsional: Ambil data existing untuk preview di bawah form, kode sama seperti sebelumnya)
    saldo_table_html = "<tr><td colspan='4' style='text-align:center; padding:10px;'>Belum ada data</td></tr>"
    # ... (Kode preview saldo Anda yang lama bisa ditaruh di sini jika mau, saya sederhanakan agar fokus ke fungsi utama)

    # --- 4. SUSUN HALAMAN UTAMA ---
    body = f"""
    <style>
        .btn-save {{ background-color: #28a745; color: white; border: none; padding: 12px 25px; border-radius: 5px; cursor: pointer; font-size: 16px; }}
        .btn-save:hover {{ background-color: #218838; }}
        .btn-reset {{ background-color: #6c757d; color: white; border: none; padding: 12px 25px; border-radius: 5px; cursor: pointer; font-size: 16px; margin-left: 10px; }}
    </style>

    {f'<div style="background: #f8d7da; color: #721c24; padding: 15px; border-radius: 5px; margin-bottom: 20px;">{error_message}</div>' if error_message else ''}
    {f'<div style="background: #d4edda; color: #155724; padding: 15px; border-radius: 5px; margin-bottom: 20px;">{success_message}</div>' if success_message else ''}

    <form method="POST" action="/admin/opening-balance" onsubmit="return confirm('Simpan Neraca Saldo Awal? Data lama akan ditimpa.');">
        
        {company_form_html}
        
        <div style="background: white; padding: 20px; border: 1px solid #ddd; border-radius: 5px;">
            <h3 style="margin-top: 0; border-bottom: 2px solid #007bff; padding-bottom: 10px;">üí∞ Input Saldo Awal Akun</h3>
            <div style="background: #fff3cd; padding: 10px; margin-bottom: 15px; border-radius: 4px; font-size: 0.9em;">
                 <strong>Petunjuk:</strong>
                 <ul style="margin: 5px 0; padding-left: 20px;">
                    <li>Masukkan saldo positif saja. Sistem otomatis menentukan Debit/Kredit.</li>
                    <li>Untuk <strong>Aset Tetap</strong>, isi "Nilai Aset", "Masa Manfaat", & "Tahun Pakai".</li>
                    <li>Kolom <strong>Akumulasi Penyusutan</strong> akan terisi otomatis!</li>
                 </ul>
            </div>

            {account_inputs}
            
            <div style="text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee;">
                <button type="submit" class="btn-save">üíæ Simpan Semua Data</button>
                <button type="button" class="btn-reset" onclick="if(confirm('Reset form?')) document.querySelector('form').reset();">üîÑ Reset</button>
            </div>
        </div>
    </form>

    <script>
    // Map Kode Aset -> Kode Akumulasi
    const assetToAccumMap = {{
        '1201': '1202', // Peralatan -> Akum. Peralatan
        '1203': '1204', // Kendaraan -> Akum. Kendaraan
        '1205': '1206'  // Bangunan -> Akum. Bangunan
    }};

    function formatRupiah(angka) {{
        return new Intl.NumberFormat('id-ID', {{ style: 'currency', currency: 'IDR', maximumFractionDigits: 0 }}).format(angka);
    }}

    function hitungPenyusutan(assetCode) {{
        // 1. Tentukan Kode Akumulasi Pasangannya
        const accumCode = assetToAccumMap[assetCode];
        if (!accumCode) return;

        // 2. Ambil Elemen Input
        const elAsset = document.getElementById('balance_' + assetCode);
        const elLife  = document.getElementById('masa_manfaat_' + assetCode);
        const elYears = document.getElementById('tahun_pakai_' + assetCode);
        
        // Input Target (Akumulasi)
        const elAccumInput = document.getElementById('balance_' + accumCode);
        // Teks Info
        const elInfoText = document.getElementById('info_text_' + assetCode);

        // 3. Ambil Nilai (Konversi ke Float, default 0)
        let valAsset = parseFloat(elAsset.value) || 0;
        let valLife  = parseFloat(elLife.value)  || 0;
        let valYears = parseFloat(elYears.value) || 0;

        // 4. Rumus Perhitungan
        let depPerYear = 0;
        let totalAccum = 0;

        if (valLife > 0) {{
            depPerYear = valAsset / valLife;
            totalAccum = depPerYear * valYears;
        }}

        // 5. UPDATE FIELD INPUT AKUMULASI (Ini yang sebelumnya kurang)
        if (elAccumInput) {{
            // Masukkan hasil hitungan ke input box akumulasi
            // Kita bulatkan agar rapi, atau biarkan desimal
            elAccumInput.value = Math.round(totalAccum); 
        }}

        // 6. Update Teks Info untuk User
        if (elInfoText) {{
            if (totalAccum > 0) {{
                elInfoText.innerHTML = `Penyusutan/Thn: ${{formatRupiah(depPerYear)}} | Total Akumulasi: ${{formatRupiah(totalAccum)}} (Auto-filled)`;
            }} else {{
                elInfoText.innerHTML = "";
            }}
        }}
    }}
    </script>
    """
    # ==========================================
    # B. PERSIAPAN DATA TAMPILAN (GET METHOD)
    # ==========================================
    
    # 1. Ambil Daftar Akun Master
    try:
        accounts = db.execute("""
            SELECT coa.account_code, coa.account_name, coa.account_type
            FROM chart_of_accounts coa
            ORDER BY 
                CASE coa.account_type
                    WHEN 'Aset Lancar' THEN 1 WHEN 'Aset Tetap' THEN 2 WHEN 'Liabilitas' THEN 3
                    WHEN 'Ekuitas' THEN 4 WHEN 'Pendapatan' THEN 5 WHEN 'Beban' THEN 6 ELSE 7
                END, coa.account_code
        """).fetchall()
    except:
        accounts = []

    # 2. Generate HTML Form Input
    company_name_val = company_info['company_name'] if company_info else ''
    company_period_val = company_info['accounting_period'] if company_info else ''
    
    # -- Form Bagian Atas (Perusahaan) --
    company_form_html = f"""
    <div style="background-color: #e7f3ff; padding: 20px; border-radius: 5px; margin-bottom: 20px; border-left: 4px solid #007bff;">
        <h3 style="margin-top: 0; color: #0056b3;">üè¢ Informasi Perusahaan</h3>
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
            <div>
                <label style="font-weight: bold;">Nama Perusahaan:</label>
                <input type="text" name="company_name" value="{company_name_val}" placeholder="Contoh: PT Maju Jaya" required style="width: 100%; padding: 8px;">
            </div>
            <div>
                <label style="font-weight: bold;">Periode Akuntansi:</label>
                <input type="text" name="accounting_period" value="{company_period_val}" placeholder="Contoh: 2025" required style="width: 100%; padding: 8px;">
            </div>
        </div>
    </div>
    """

    # -- Loop Form Akun --
    account_inputs = ""
    current_type = ""
    aset_penyusutan_map = {
        '1201': {'akumulasi': '1202', 'default': 5},
        '1203': {'akumulasi': '1204', 'default': 8},
        '1205': {'akumulasi': '1206', 'default': 20}
    }
    
    for account in accounts:
        if account['account_type'] != current_type:
            if current_type != "": account_inputs += "</tbody></table><br>"
            account_inputs += f"""
            <h4 style="border-bottom: 2px solid #ddd; margin-top: 15px;">{account['account_type']}</h4>
            <table style="width: 100%; border-collapse: collapse;">
                <thead>
                    <tr style="background: #f8f9fa;">
                        <th style="padding:8px; border:1px solid #ddd; width:15%">Kode</th>
                        <th style="padding:8px; border:1px solid #ddd; width:35%">Nama Akun</th>
                        <th style="padding:8px; border:1px solid #ddd; width:25%">Saldo Awal</th>
                        <th style="padding:8px; border:1px solid #ddd; width:25%">Ket</th>
                    </tr>
                </thead><tbody>
            """
            current_type = account['account_type']

        # Styling Logic
        row_bg = ""
        # Tentukan Normal Balance Text
        if account['account_type'] in ('Aset Lancar', 'Aset Tetap', 'Beban'):
            ket_html = "<span style='color:#28a745; font-weight:bold;'>Debit</span>"
        else:
            ket_html = "<span style='color:#dc3545; font-weight:bold;'>Kredit</span>"

        # Override untuk Akumulasi
        if 'akumulasi' in account['account_name'].lower() and 'penyusutan' in account['account_name'].lower():
            row_bg = "background-color: #fff3cd;"
            ket_html = "‚ö†Ô∏è Kontra Aset (<span style='color:#dc3545; font-weight:bold;'>Kredit</span>)"
        elif account['account_type'] == 'Beban':
            row_bg = "background-color: #fbfcfd;"

        code = account['account_code']
        has_calc = code in aset_penyusutan_map
        
        # Baris Akun Utama
        account_inputs += f"""
        <tr style="{row_bg}">
            <td style="padding:8px; border:1px solid #ddd; text-align:center;">{code}</td>
            <td style="padding:8px; border:1px solid #ddd;">{account['account_name']}</td>
            <td style="padding:8px; border:1px solid #ddd;">
                <input type="number" step="0.01" name="balance_{code}" id="balance_{code}" placeholder="0"
                       style="width:100%; padding:5px;" {'onkeyup="hitungPenyusutan(\''+code+'\')"' if has_calc else ''}>
            </td>
            <td style="padding:8px; border:1px solid #ddd; font-size:0.85em; color:#555;">Normal: {ket_html}</td>
        </tr>
        """
        
        # Baris Kalkulator (Hidden by default unless asset)
        if has_calc:
            default_life = aset_penyusutan_map[code]['default']
            account_inputs += f"""
            <tr style="background:#fdfdfe;">
                <td colspan="2" style="text-align:right; padding:5px; border-bottom:1px solid #ddd; color:#007bff;"><small>‚öôÔ∏è Kalkulator:</small></td>
                <td colspan="2" style="padding:5px; border-bottom:1px solid #ddd;">
                    <div style="display:flex; gap:5px;">
                        <input type="number" id="masa_{code}" value="{default_life}" style="width:50px;" onkeyup="hitungPenyusutan('{code}')"> <small>Thn (Masa)</small>
                        <input type="number" id="pakai_{code}" value="0" style="width:50px;" onkeyup="hitungPenyusutan('{code}')"> <small>Thn (Pakai)</small>
                    </div>
                    <div id="info_{code}" style="font-size:0.8em; color:#28a745; margin-top:2px;"></div>
                </td>
            </tr>
            """
    account_inputs += "</tbody></table>"
    
    # MENAMPILKAN DATA YANG SUDAH TERSIMPAN (PREVIEW TABLE)
    try:
        saved_data = db.execute("""
            SELECT jd.account_code, coa.account_name, coa.account_type, SUM(jd.debit) as deb, SUM(jd.credit) as cred
            FROM journal_details jd
            JOIN chart_of_accounts coa ON jd.account_code = coa.account_code
            JOIN journal_entries je ON jd.entry_id = je.id
            WHERE je.description LIKE '%Saldo Awal%'
            GROUP BY jd.account_code
            ORDER BY CASE coa.account_type
                    WHEN 'Aset Lancar' THEN 1 WHEN 'Aset Tetap' THEN 2 WHEN 'Liabilitas' THEN 3
                    WHEN 'Ekuitas' THEN 4 WHEN 'Pendapatan' THEN 5 WHEN 'Beban' THEN 6 ELSE 7
                END, jd.account_code
        """).fetchall()
    except:
        saved_data = []

    saved_table_rows = ""
    total_saved_debit = Decimal('0')
    total_saved_credit = Decimal('0')
    curr_saved_type = ""

    if not saved_data:
        saved_table_rows = "<tr><td colspan='4' style='text-align:center; padding:20px; color:#999;'>Belum ada data tersimpan. Silakan input di atas.</td></tr>"
    else:
        for row in saved_data:
            # Header Kategori di Tabel Preview
            if row['account_type'] != curr_saved_type:
                saved_table_rows += f"<tr style='background:#e9ecef;'><td colspan='4' style='padding:8px; font-weight:bold;'>{row['account_type']}</td></tr>"
                curr_saved_type = row['account_type']
            
            d_val = Decimal(str(row['deb']))
            c_val = Decimal(str(row['cred']))
            total_saved_debit += d_val
            total_saved_credit += c_val
            
            saved_table_rows += f"""
            <tr>
                <td style="padding:8px; border:1px solid #ddd;">{row['account_code']}</td>
                <td style="padding:8px; border:1px solid #ddd;">{row['account_name']}</td>
                <td style="padding:8px; border:1px solid #ddd; text-align:right;">{format_currency(d_val) if d_val > 0 else '-'}</td>
                <td style="padding:8px; border:1px solid #ddd; text-align:right;">{format_currency(c_val) if c_val > 0 else '-'}</td>
            </tr>
            """
    
    # Footer Total Tabel Preview
    balance_status = "‚úÖ BALANCE" if total_saved_debit == total_saved_credit and total_saved_debit > 0 else "‚ùå TIDAK BALANCE"
    balance_color = "#28a745" if total_saved_debit == total_saved_credit and total_saved_debit > 0 else "#dc3545"

    preview_html = f"""
    <div style="margin-top: 40px; background: white; padding: 20px; border: 1px solid #ddd; border-radius: 5px; box-shadow: 0 2px 5px rgba(0,0,0,0.05);">
        <h3 style="margin-top: 0; color: #333; border-bottom: 2px solid #28a745; padding-bottom: 10px;">üìä Data Neraca Saldo Tersimpan</h3>
        <table style="width: 100%; border-collapse: collapse;">
            <thead>
                <tr style="background: #343a40; color: white;">
                    <th style="padding:10px;">Kode</th>
                    <th style="padding:10px;">Nama Akun</th>
                    <th style="padding:10px; text-align:right;">Debit</th>
                    <th style="padding:10px; text-align:right;">Kredit</th>
                </tr>
            </thead>
            <tbody>
                {saved_table_rows}
                <tr style="background: #f8f9fa; font-weight:bold;">
                    <td colspan="2" style="padding:10px; text-align:right;">TOTAL</td>
                    <td style="padding:10px; text-align:right; color:#28a745;">{format_currency(total_saved_debit)}</td>
                    <td style="padding:10px; text-align:right; color:#dc3545;">{format_currency(total_saved_credit)}</td>
                </tr>
                <tr style="background: {balance_color}; color: white; font-weight:bold;">
                    <td colspan="4" style="padding:10px; text-align:center;">STATUS: {balance_status}</td>
                </tr>
            </tbody>
        </table>
    </div>
    """

    body = f"""
    <style>
        .btn-submit {{ background: #007bff; color: white; border: none; padding: 12px 30px; font-size: 16px; border-radius: 4px; cursor: pointer; }}
        .btn-submit:hover {{ background: #0056b3; }}
    </style>
    
    {f'<div style="background:#f8d7da; color:#721c24; padding:15px; border-radius:4px; margin-bottom:15px;">{error_message}</div>' if error_message else ''}
    {f'<div style="background:#d4edda; color:#155724; padding:15px; border-radius:4px; margin-bottom:15px;">{success_message}</div>' if success_message else ''}

    <form method="POST" action="/admin/opening-balance" onsubmit="return confirm('Simpan Neraca Saldo? Data lama akan digantikan.');">
        {company_form_html}
        
        <div style="background: white; padding: 20px; border: 1px solid #ddd; border-radius: 5px;">
            <h3 style="margin-top:0;">üìù Input Saldo Awal</h3>
            <p style="font-size:0.9em; color:#666;">Masukkan saldo positif. Gunakan fitur kalkulator di bawah akun aset tetap untuk menghitung akumulasi.</p>
            {account_inputs}
            
            <div style="text-align:center; margin-top:30px;">
                <button type="submit" class="btn-submit">üíæ Simpan Perubahan</button>
            </div>
        </div>
    </form>

    {preview_html}

    <script>
    const mapAkum = {{
        '1201': '1202', 
        '1203': '1204', 
        '1205': '1206'
    }};

    function formatRupiah(angka) {{
        return new Intl.NumberFormat('id-ID', {{ style: 'currency', currency: 'IDR', maximumFractionDigits: 0 }}).format(angka);
    }}

    function hitungPenyusutan(kodeAset) {{
        const kodeAkum = mapAkum[kodeAset];
        if(!kodeAkum) return;

        const asetVal = parseFloat(document.getElementById('balance_'+kodeAset).value) || 0;
        const masa = parseFloat(document.getElementById('masa_'+kodeAset).value) || 0;
        const pakai = parseFloat(document.getElementById('pakai_'+kodeAset).value) || 0;
        
        let akumulasi = 0;
        let perTahun = 0;

        if(masa > 0) {{
            perTahun = asetVal / masa;
            akumulasi = perTahun * pakai;
        }}

        // Isi Input Akumulasi otomatis
        const inputAkum = document.getElementById('balance_'+kodeAkum);
        if(inputAkum) inputAkum.value = Math.round(akumulasi);
        
        // Tampilkan Info Text
        const infoDiv = document.getElementById('info_'+kodeAset);
        if(infoDiv) {{
            if(akumulasi > 0) infoDiv.innerHTML = `Penyusutan/Thn: ${{formatRupiah(perTahun)}} | Total Akumulasi: ${{formatRupiah(akumulasi)}} (Auto-filled)`;
            else infoDiv.innerHTML = "";
        }}
    }}
    </script>
    """

    return render_page("Neraca Saldo Awal", body, sidebar_content=get_admin_sidebar_html())

def render_add_account_form(error=None):
    """Helper untuk merender form tambah akun."""
    body = f"""
    <form action="/admin/add-account" method="POST">
        <label for="account_code">Kode Akun (Berupa Angka XXXX):</label>
        <input type="text" id="account_code" name="account_code" required>
        
        <label for="account_name">Nama Akun:</label>
        <input type="text" id="account_name" name="account_name" required>
        
        <label for="account_type">Tipe Akun:</label>
        <select id="account_type" name="account_type">
            <option value="Aset Tetap">Aset Tetap</option>
            <option value="Aset Lancar">Aset Lancar</option> 
            <option value="Liabilitas">Liabilitas</option>
            <option value="Ekuitas">Ekuitas</option>
            <option value="Pendapatan">Pendapatan</option>
            <option value="Beban">Beban</option>
        </select>
        
        <input type="submit" value="Simpan Akun Baru">
    </form>
    <a href="{url_for('chart_of_accounts')}">Batal</a>
    """
    return render_page("Input Akun", body, sidebar_content=get_admin_sidebar_html(), error_message=error)

    # Redirect kembali ke daftar akun, kirim pesan error jika ada
    if error_msg:
        return redirect(url_for('chart_of_accounts', error=error_msg))
    else:
        return redirect(url_for('chart_of_accounts'))
    
# =========================
#RIWAYAT TRANSAKSI
# =========================
@app.route("/admin/delete-transactions", methods=['POST'])
def delete_transactions():
    """
    Memproses penghapusan transaksi jurnal dan inventory log terkait.
    Menggunakan deskripsi (Nomor Faktur) sebagai kunci penghapusan log inventory.
    """
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
        
    ids_to_delete = request.form.getlist('selected_ids')
    db = get_db()
    
    with db:
        for entry_id in ids_to_delete:
            # 1. AMBIL DATA JURNAL DULU
            # Kita hanya perlu deskripsi (yang berisi No. Faktur)
            journal = db.execute("SELECT description FROM journal_entries WHERE id = ?", (entry_id,)).fetchone()
            
            if journal:
                unique_identifier = journal['description']
                
                # CATATAN PENTING:
                # Logika ini mengasumsikan bahwa:
                # a) Semua Pembelian/Masuk (purchase) di inventory_log dan journal_entries
                #    menggunakan Nomor Faktur sebagai unique_identifier/description.
                # b) Nomor Faktur ini UNIK di seluruh tabel.

                # 2. HAPUS DATA DI INVENTORY LOG
                # Hapus semua baris di inventory_log yang memiliki deskripsi yang cocok
                db.execute("DELETE FROM inventory_log WHERE description = ?", (unique_identifier,))
                
                # 3. HAPUS DETAIL JURNAL
                db.execute("DELETE FROM journal_details WHERE entry_id = ?", (entry_id,))
                
                # 4. HAPUS HEADER JURNAL
                db.execute("DELETE FROM journal_entries WHERE id = ?", (entry_id,))
            
            # Jika jurnal tidak ditemukan (mungkin sudah terhapus), lanjutkan loop.

    return redirect(url_for('transactions_list'))

# ==================================
# JURNAL UMUM (GENERAL JOURNAL)
# ==================================
@app.route("/admin/general-journal", methods=["GET", "POST"])
def general_journal():
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    error_message = None
    success_message = None
    
    # Handle POST request (simpan jurnal baru)
    if request.method == 'POST':
        try:
            entry_date_str = request.form.get('entry_date')
            description = request.form['description']
            entries_data = request.form.get('entries_data')
            
            current_time_str = datetime.now().strftime('%H:%M:%S')
            entry_timestamp = f"{entry_date_str} {current_time_str}"
            
            if not entries_data:
                error_message = "Data entri jurnal tidak valid."
            else:
                entries = json.loads(entries_data)
                
                if len(entries) < 2:
                    error_message = "Jurnal harus memiliki minimal 2 entri (1 debit dan 1 kredit)."
                else:
                    total_debit = sum(Decimal(str(entry['debit'])) for entry in entries)
                    total_credit = sum(Decimal(str(entry['credit'])) for entry in entries)
                    
                    if total_debit != total_credit:
                        error_message = f"Jurnal tidak balance! Debit: {format_currency(total_debit)}, Kredit: {format_currency(total_credit)}"
                    else:
                        entry_date_obj = datetime.strptime(entry_date_str, '%Y-%m-%d')
                        journal_code = generate_journal_code(db, entry_date_obj)
                        
                        with db:
                            cursor = db.execute(
                                "INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                (journal_code, entry_timestamp, description)
                            )
                            entry_id = cursor.lastrowid
                            
                            for entry in entries:
                                db.execute(
                                    "INSERT INTO journal_details (entry_id, account_code, debit, credit) VALUES (?, ?, ?, ?)",
                                    (entry_id, entry['account_code'], entry['debit'], entry['credit'])
                                )
                            
                            success_message = f"Jurnal berhasil disimpan dengan kode: {journal_code}"
                            
        except Exception as e:
            error_message = f"Terjadi kesalahan: {str(e)}"
    
    accounts = db.execute(
        "SELECT account_code, account_name, account_type FROM chart_of_accounts ORDER BY account_code"
    ).fetchall()
    
    account_options = ""
    for account in accounts:
        account_options += f'<option value="{account["account_code"]}">{account["account_code"]} - {account["account_name"]} ({account["account_type"]})</option>'
    
    today_date = datetime.now().strftime('%Y-%m-%d')

    # --- PERUBAHAN PADA JAVASCRIPT UNTUK TOMBOL HAPUS ---
    js_code = f"""
    <script>
    let entryCount = 0;
    
    function addJournalEntry() {{
        entryCount++;
        const entriesContainer = document.getElementById('journal-entries');
        
        const newEntry = document.createElement('div');
        newEntry.className = 'journal-entry';
        // Style card putih untuk setiap baris
        newEntry.style.border = '1px solid #ddd';
        newEntry.style.padding = '15px';
        newEntry.style.marginBottom = '15px';
        newEntry.style.borderRadius = '8px';
        newEntry.style.backgroundColor = '#fff';
        newEntry.style.boxShadow = '0 2px 4px rgba(0,0,0,0.05)';
        
        newEntry.innerHTML = `
            <div style="display: grid; grid-template-columns: 2fr 1fr 1fr 0.5fr; gap: 15px; align-items: end;">
                <div>
                    <label style="font-weight: bold; display: block; margin-bottom: 5px;">Akun:</label>
                    <select name="account_code_${{entryCount}}" required style="width: 100%; padding: 10px; border: 1px solid #ccc; border-radius: 4px;">
                        <option value="">Pilih Akun</option>
                        {account_options}
                    </select>
                </div>
                <div>
                    <label style="font-weight: bold; display: block; margin-bottom: 5px;">Debit:</label>
                    <input type="number" name="debit_${{entryCount}}" step="0.01" min="0" value="0" 
                           onchange="updateTotals()" style="width: 100%; padding: 10px; border: 1px solid #ccc; border-radius: 4px;">
                </div>
                <div>
                    <label style="font-weight: bold; display: block; margin-bottom: 5px;">Kredit:</label>
                    <input type="number" name="credit_${{entryCount}}" step="0.01" min="0" value="0" 
                           onchange="updateTotals()" style="width: 100%; padding: 10px; border: 1px solid #ccc; border-radius: 4px;">
                </div>
                <div>
                    <button type="button" onclick="removeEntry(this)" class="btn-styled btn-red" style="width: 100%;">
                        Hapus
                    </button>
                </div>
            </div>
        `;
        
        entriesContainer.appendChild(newEntry);
        updateTotals();
    }}
    
    function removeEntry(button) {{
        const entry = button.closest('.journal-entry');
        entry.remove();
        updateTotals();
    }}
    
    function updateTotals() {{
        let totalDebit = 0;
        let totalCredit = 0;
        
        const entries = document.getElementsByClassName('journal-entry');
        for (let entry of entries) {{
            const debitInput = entry.querySelector('input[name^="debit_"]');
            const creditInput = entry.querySelector('input[name^="credit_"]');
            
            if (debitInput && creditInput) {{
                totalDebit += parseFloat(debitInput.value) || 0;
                totalCredit += parseFloat(creditInput.value) || 0;
            }}
        }}
        
        document.getElementById('total-debit').textContent = totalDebit.toLocaleString('id-ID', {{
            minimumFractionDigits: 2,
            maximumFractionDigits: 2
        }});
        document.getElementById('total-credit').textContent = totalCredit.toLocaleString('id-ID', {{
            minimumFractionDigits: 2,
            maximumFractionDigits: 2
        }});
        
        const balanceStatus = document.getElementById('balance-status');
        if (Math.abs(totalDebit - totalCredit) < 0.01) {{
            balanceStatus.innerHTML = '<span style="color: #28a745; display: flex; align-items: center; gap:5px;"><b>‚úì Balance</b></span>';
        }} else {{
            balanceStatus.innerHTML = '<span style="color: #dc3545; display: flex; align-items: center; gap:5px;"><b>‚úó Tidak Balance</b></span>';
        }}
    }}
    
    function prepareSubmit() {{
        const entries = [];
        const entryElements = document.getElementsByClassName('journal-entry');
        
        for (let i = 0; i < entryElements.length; i++) {{
            const entry = entryElements[i];
            const accountSelect = entry.querySelector('select[name^="account_code_"]');
            const debitInput = entry.querySelector('input[name^="debit_"]');
            const creditInput = entry.querySelector('input[name^="credit_"]');
            
            if (accountSelect && debitInput && creditInput) {{
                const accountCode = accountSelect.value;
                const debit = parseFloat(debitInput.value) || 0;
                const credit = parseFloat(creditInput.value) || 0;
                
                if (accountCode && (debit > 0 || credit > 0)) {{
                    entries.push({{
                        account_code: accountCode,
                        debit: debit,
                        credit: credit
                    }});
                }}
            }}
        }}
        
        if (entries.length < 2) {{
            alert('Jurnal harus memiliki minimal 2 entri!');
            return false;
        }}
        
        const totalDebit = entries.reduce((sum, entry) => sum + entry.debit, 0);
        const totalCredit = entries.reduce((sum, entry) => sum + entry.credit, 0);
        
        if (Math.abs(totalDebit - totalCredit) > 0.01) {{
            alert(`Jurnal tidak balance! Debit: ${{totalDebit}}, Kredit: ${{totalCredit}}`);
            return false;
        }}
        
        document.getElementById('entries-data').value = JSON.stringify(entries);
        return true;
    }}
    
    document.addEventListener('DOMContentLoaded', function() {{
        addJournalEntry();
        addJournalEntry();
    }});
    </script>
    """
    
    body = f"""
    <style>
        /* --- BUTTON STYLES MODERN (SERAGAM DENGAN LOGOUT) --- */
        .btn-styled {{
            padding: 10px 20px;
            border: none;
            border-radius: 30px; /* Bentuk Kapsul */
            font-weight: bold;
            cursor: pointer;
            font-size: 0.95em;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            gap: 8px;
            color: white !important;
            box-shadow: 0 4px 6px rgba(0,0,0,0.2); /* Shadow */
            transition: all 0.2s ease;
            text-decoration: none;
        }}
        
        /* Efek Hover Umum */
        .btn-styled:hover {{
            transform: translateY(-2px);
            box-shadow: 0 6px 10px rgba(0,0,0,0.25);
        }}

        /* Warna-warna Tombol */
        .btn-galaxy {{ background-color: #2A4B7C; }} /* Biru Galaxy (Pengganti Biru Biasa) */
        .btn-galaxy:hover {{ background-color: #1e3a63; }}

        .btn-red {{ background-color: #DC3545; }} /* Merah Tetap */
        .btn-red:hover {{ background-color: #b02a37; }}

        .btn-green {{ background-color: #28a745; }} /* Hijau Tetap */
        .btn-green:hover {{ background-color: #1e7e34; }}

        .btn-gray {{ background-color: #6c757d; }} /* Abu-abu Tetap */
        .btn-gray:hover {{ background-color: #545b62; }}

        /* --------------------------------------------------- */

        .journal-form {{
            background-color: #f8f9fa;
            padding: 25px;
            border-radius: 8px;
            border: 1px solid #dee2e6;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        }}
        .totals-display {{
            background-color: #e9ecef;
            padding: 15px;
            border-radius: 5px;
            margin: 20px 0;
            font-weight: bold;
            border: 1px solid #ced4da;
        }}
        
        /* Header Table Style */
        .table-journal th {{
            background-color: #e9ecef !important;
            color: black !important;
            text-align: center;
            padding: 12px;
            border: 1px solid #ddd;
        }}
        .table-journal td {{ padding: 10px; border: 1px solid #ddd; }}
        .journal-link {{ color: black; font-weight: bold; text-decoration: none; }}
        .journal-link:hover {{ text-decoration: underline; }}
    </style>
    
    <div class="journal-form">
        <h3 style="margin-top: 0; color: #333;">Input Jurnal Umum</h3>
        <hr style="margin-bottom: 20px;">
        
        <form id="journal-form" action="/admin/general-journal" method="POST" onsubmit="return prepareSubmit()" style="border: none; padding: 0; background: none;">
            <input type="hidden" id="entries-data" name="entries_data">
            
            <div style="display: grid; grid-template-columns: 1fr 2fr; gap: 20px; margin-bottom: 20px;">
                <div>
                    <label for="entry_date" style="font-weight: bold; display: block; margin-bottom: 5px;">Tanggal Transaksi:</label>
                    <input type="date" id="entry_date" name="entry_date" required value="{today_date}" 
                           style="width: 100%; padding: 10px; border: 1px solid #ccc; border-radius: 4px;">
                </div>
                
                <div>
                    <label for="description" style="font-weight: bold; display: block; margin-bottom: 5px;">Keterangan/Deskripsi Jurnal:</label>
                    <input type="text" id="description" name="description" required 
                           style="width: 100%; padding: 10px; border: 1px solid #ccc; border-radius: 4px;" 
                           placeholder="Contoh: Pembelian perlengkapan secara tunai">
                </div>
            </div>
            
            <h4 style="margin-bottom: 10px;">Entri Jurnal:</h4>
            <div id="journal-entries"></div>
            
            <button type="button" onclick="addJournalEntry()" class="btn-styled btn-galaxy">
                + Tambah Baris Akun
            </button>
            
            <div class="totals-display">
                <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px; text-align: center;">
                    <div>Total Debit: <br><span style="font-size: 1.2em;">Rp <span id="total-debit">0</span></span></div>
                    <div>Total Kredit: <br><span style="font-size: 1.2em;">Rp <span id="total-credit">0</span></span></div>
                    <div style="display: flex; align-items: center; justify-content: center;">
                        <div id="balance-status" style="font-size: 1.1em;"></div>
                    </div>
                </div>
            </div>
            
            <div style="margin-top: 20px; text-align: right; display: flex; justify-content: flex-end; gap: 10px;">
                <button type="button" onclick="window.location.href='/admin'" class="btn-styled btn-gray">
                    Batal
                </button>
                
                <button type="submit" class="btn-styled btn-green">
                    üíæ Simpan Jurnal
                </button>
            </div>
        </form>
    </div>
    
    <div style="display: flex; justify-content: space-between; align-items: center; margin-top: 40px; margin-bottom: 20px;">
        <h3 style="margin: 0;">Daftar Jurnal Umum</h3>
        <a href="{url_for('all_journals_detail')}" class="btn-styled btn-galaxy">
            Lihat Detail Semua Jurnal
        </a>
    </div>
    """
    
    recent_journals = db.execute("""
        SELECT je.id, je.journal_code, je.entry_timestamp, je.description,
               COUNT(jd.id) as entry_count
        FROM journal_entries je
        LEFT JOIN journal_details jd ON je.id = jd.entry_id
        GROUP BY je.id
        ORDER BY je.entry_timestamp DESC
        LIMIT 10
    """).fetchall()
    
    if recent_journals:
        journal_table = """
        <table class="table-journal" style="width: 100%; border-collapse: collapse; margin-top: 10px; background: white;">
            <thead>
                <tr>
                    <th>Kode Jurnal</th>
                    <th>Tanggal/Waktu</th>
                    <th>Deskripsi</th>
                    <th style="text-align: center;">Jml Entri</th>
                    <th style="text-align: center;">Aksi</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for journal in recent_journals:
            entry_timestamp_str = journal['entry_timestamp'] 
            try:
                dt_obj = datetime.strptime(entry_timestamp_str, '%Y-%m-%d %H:%M:%S')
                date_display = dt_obj.strftime('%d-%m-%Y')
                time_display = dt_obj.strftime('%H:%M:%S')
            except:
                date_display = entry_timestamp_str
                time_display = ""

            journal_table += f"""
                <tr>
                    <td style="font-weight: bold; padding: 10px; border: 1px solid #ddd;">
                        <a href="/admin/journal-detail/{journal['id']}" class="journal-link">{journal['journal_code']}</a>
                    </td>
                    <td style="padding: 10px; border: 1px solid #ddd;">
                        {date_display}
                        <br><small style="color: #555;">{time_display}</small>
                    </td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{journal['description']}</td>
                    <td style="text-align: center; padding: 10px; border: 1px solid #ddd;">{journal['entry_count']}</td>
                    <td style="text-align: center; padding: 10px; border: 1px solid #ddd;">
                        <a href="/admin/journal-detail/{journal['id']}" 
                           style="background-color: #17a2b8; color: white; text-decoration: none; padding: 6px 12px; border-radius: 4px; font-size: 0.9em;">
                            Detail
                        </a>
                    </td>
                </tr>
            """
        
        journal_table += """
            </tbody>
        </table>
        """
        body += journal_table
    else:
        body += "<div style='padding: 20px; background-color: #fff; border: 1px solid #ddd; border-radius: 5px; text-align: center; color: #666;'>Belum ada jurnal yang dicatat.</div>"
    
    body += js_code
    
    return render_page("Jurnal Umum", body, sidebar_content=get_admin_sidebar_html(), error_message=error_message)

@app.route("/admin/all-journals-detail")
def all_journals_detail():
    """Menampilkan SEMUA detail jurnal dalam SATU TABEL seperti format jurnal umum"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    
    all_journal_details = db.execute("""
        SELECT 
            je.entry_timestamp,
            jd.account_code,
            coa.account_name,
            jd.debit,
            jd.credit,
            je.journal_code,
            je.description
        FROM journal_details jd
        JOIN journal_entries je ON jd.entry_id = je.id
        JOIN chart_of_accounts coa ON jd.account_code = coa.account_code
        ORDER BY je.entry_timestamp, je.id, jd.debit DESC, jd.credit DESC
    """).fetchall()
    
    # --- [PERUBAHAN DI SINI] ---
    # Hapus header lama, buat layout flexbox baru untuk sejajarkan Total & Tombol
    
    body = f"""
    <style>
        .journal-table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        /* Header Tabel Abu-abu seperti request sebelumnya */
        .journal-table th {{
            background-color: #e9ecef !important;
            color: black !important;
            padding: 12px;
            border: 1px solid #ddd;
            font-weight: bold;
            text-align: center;
        }}
        .journal-table td {{
            padding: 10px;
            border: 1px solid #ddd;
            vertical-align: top;
        }}
        .currency {{
            text-align: right;
            font-family: 'Courier New', Courier;
        }}
        .total-row {{
            font-weight: bold;
            background-color: #f2f2f2;
            border-top: 2px solid #333;
        }}
        .date-column {{ width: 12%; text-align: center; }}
        .account-code-column {{ width: 10%; }}
        .account-name-column {{ width: 35%; }}
        .debit-column {{ width: 18%; }}
        .credit-column {{ width: 18%; }}
        
        /* Container Header Baru */
        .header-actions {{
            display: flex;
            justify-content: space-between; /* Kiri kanan mentok */
            align-items: center; /* Rata tengah vertikal */
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 1px solid #eee;
        }}
        
        /* Style Tombol SAMA PERSIS dengan Logout */
        .btn-back {{
            background-color: #2A4B7C !important; /* Galaxy Blue */
            color: white !important;
            box-shadow: 0 4px 8px rgba(0,0,0,0.25) !important;
            padding: 8px 20px;
            text-decoration: none;
            font-weight: bold;
            font-size: 0.9em;
            display: inline-flex;
            align-items: center;
            gap: 8px;
            border: none;
            border-radius: 30px;
            transition: all 0.3s ease;
            cursor: pointer;
        }}
        .btn-back:hover {{
            background-color: #607d8b !important; /* Hover Blue Gray */
            transform: translateY(-2px);
            box-shadow: 0 6px 12px rgba(0,0,0,0.3) !important;
        }}
    </style>
    
    <div class="header-actions">
        <h3 style="margin: 0; color: #333;">Total Transaksi: {len(all_journal_details)} entri</h3>
        
        <a href="{url_for('general_journal')}" class="btn-back">
            <span>‚¨Ö Kembali ke Jurnal Umum</span>
        </a>
    </div>
    """
    
    if all_journal_details:
        total_debit = sum(Decimal(str(row['debit'])) for row in all_journal_details)
        total_credit = sum(Decimal(str(row['credit'])) for row in all_journal_details)
        
        table_rows = ""
        current_date_str = None
        
        for detail in all_journal_details:
            entry_timestamp_str = detail['entry_timestamp']
            try:
                entry_datetime_obj = datetime.strptime(entry_timestamp_str, '%Y-%m-%d %H:%M:%S')
                formatted_date = entry_datetime_obj.strftime('%d-%m-%Y')
            except:
                formatted_date = entry_timestamp_str
            
            # Logic pemisah tanggal
            if formatted_date != current_date_str:
                if current_date_str is not None:
                    # Baris kosong pemisah antar tanggal
                    table_rows += f"""
                    <tr>
                        <td colspan="5" style="padding: 5px; background-color: #f9f9f9;"></td>
                    </tr>
                    """
                current_date_str = formatted_date
            
            debit_display = format_currency(detail['debit']) if detail['debit'] > 0 else ""
            credit_display = format_currency(detail['credit']) if detail['credit'] > 0 else ""
            
            table_rows += f"""
            <tr>
                <td class="date-column">{formatted_date}</td>
                <td class="account-code-column">{detail['account_code']}</td>
                <td class="account-name-column">{detail['account_name']}</td>
                <td class="currency debit-column">{debit_display}</td>
                <td class="currency credit-column">{credit_display}</td>
            </tr>
            """
        body += f"""
        <table class="journal-table">
            <thead>
                <tr>
                    <th class="date-column">Tanggal</th>
                    <th class="account-code-column">Kode Akun</th>
                    <th class="account-name-column">Nama Akun</th>
                    <th class="debit-column">Debit</th>
                    <th class="credit-column">Kredit</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
            <tfoot>
                <tr class="total-row">
                    <td colspan="3" style="text-align: right; padding-right: 20px;"><strong>TOTAL</strong></td>
                    <td class="currency"><strong>{format_currency(total_debit)}</strong></td>
                    <td class="currency"><strong>{format_currency(total_credit)}</strong></td>
                </tr>
            </tfoot>
        </table>
        
        <div style="margin-top: 20px; padding: 15px; background-color: #f8f9fa; border-radius: 5px;">
            <h4>Status Balance:</h4>
            <p>
                <strong>Total Debit:</strong> {format_currency(total_debit)} | 
                <strong>Total Kredit:</strong> {format_currency(total_credit)} | 
                <strong>Status:</strong> 
                <span style="color: {'green' if total_debit == total_credit else 'red'}; font-weight: bold;">
                    {'‚úì BALANCE' if total_debit == total_credit else '‚úó TIDAK BALANCE'}
                </span>
            </p>
        </div>
        """
    else:
        body += "<p>Belum ada detail transaksi untuk ditampilkan.</p>"
    
    # Judul halaman tetap ada di tab browser, tapi di body sudah dihapus/disesuaikan
    return render_page("Semua Detail Jurnal", body, sidebar_content=get_admin_sidebar_html())

# Tambahkan rute baru untuk menampilkan semua jurnal
@app.route("/admin/all-journals")
def all_journals():
    """Menampilkan SEMUA jurnal dalam satu halaman"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    
    # Ambil SEMUA jurnal
    all_journals = db.execute("""
        SELECT je.id, je.journal_code, je.entry_timestamp, je.description,
               COUNT(jd.id) as entry_count
        FROM journal_entries je
        LEFT JOIN journal_details jd ON je.id = jd.entry_id
        GROUP BY je.id
        ORDER BY je.entry_timestamp DESC
    """).fetchall()
    
    body = f"""
    <style>
        .btn-small {{
            padding: 5px 10px;
            font-size: 0.8em;
            margin: 2px;
        }}
    </style>
    
    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
        <h2 style="margin: 0;">Semua Jurnal Umum</h2>
        <a href="{url_for('general_journal')}" style="text-decoration: none;">
            <input type="button" value="Kembali ke Jurnal Umum" class="btn-blue">
        </a>
    </div>
    
    <p>Total: {len(all_journals)} jurnal</p>
    """
    
    if all_journals:
        journal_table = """
        <table style="width: 100%; border-collapse: collapse; margin-top: 20px;">
            <thead>
                <tr style="background-color: #f2f2f2;">
                    <th style="padding: 10px; border: 1px solid #ddd;">Kode Jurnal</th>
                    <th style="padding: 10px; border: 1px solid #ddd;">Tanggal/Waktu</th>
                    <th style="padding: 10px; border: 1px solid #ddd;">Deskripsi</th>
                    <th style="padding: 10px; border: 1px solid #ddd;">Jumlah Entri</th>
                    <th style="padding: 10px; border: 1px solid #ddd; text-align: center;">Aksi</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for journal in all_journals:
            journal_date, journal_time = journal['entry_timestamp'].split(' ', 1)
            journal_table += f"""
                <tr>
                    <td style="padding: 10px; border: 1px solid #ddd;">{journal['journal_code']}</td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{journal_date}<br><small>{journal_time}</small></td>
                    <td style="padding: 10px; border: 1px solid #ddd;">{journal['description']}</td>
                    <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">{journal['entry_count']}</td>
                    <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">
                        <a href="/admin/journal-detail/{journal['id']}" 
                           style="color: #007bff; text-decoration: none; padding: 5px 10px; border: 1px solid #007bff; border-radius: 3px; font-size: 0.8em;">
                            Lihat Detail
                        </a>
                    </td>
                </tr>
            """
        
        journal_table += """
            </tbody>
        </table>
        """
        body += journal_table
    else:
        body += "<p>Belum ada jurnal yang dicatat.</p>"
    
    return render_page("Semua Jurnal Umum", body, sidebar_content=get_admin_sidebar_html())

@app.route("/admin/journal-detail/<int:journal_id>")
def journal_detail(journal_id):
    """Menampilkan detail lengkap dari sebuah jurnal"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    
    # Ambil data header jurnal
    journal_header = db.execute("""
        SELECT journal_code, entry_timestamp, description 
        FROM journal_entries 
        WHERE id = ?
    """, (journal_id,)).fetchone()
    
    if not journal_header:
        return redirect(url_for('general_journal', error="Jurnal tidak ditemukan"))
    
    # Ambil detail jurnal
    journal_details = db.execute("""
        SELECT jd.account_code, jd.debit, jd.credit, coa.account_name, coa.account_type
        FROM journal_details jd
        JOIN chart_of_accounts coa ON jd.account_code = coa.account_code
        WHERE jd.entry_id = ?
        ORDER BY jd.debit DESC, jd.credit DESC
    """, (journal_id,)).fetchall()
    
    # Hitung total
    total_debit = sum(Decimal(str(row['debit'])) for row in journal_details)
    total_credit = sum(Decimal(str(row['credit'])) for row in journal_details)
    
    # Format detail jurnal
    details_html = ""
    for detail in journal_details:
        debit_display = format_currency(detail['debit']) if detail['debit'] > 0 else ""
        credit_display = format_currency(detail['credit']) if detail['credit'] > 0 else ""
        
        details_html += f"""
        <tr>
            <td style="padding: 8px; border: 1px solid #ddd;">{detail['account_code']}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{detail['account_name']}</td>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{debit_display}</td>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{credit_display}</td>
        </tr>
        """
    
    entry_datetime_obj = datetime.strptime(journal_header['entry_timestamp'], '%Y-%m-%d %H:%M:%S')
    formatted_timestamp = entry_datetime_obj.strftime('%d-%m-%Y %H:%M:%S')

    body = f"""
    <h3>Detail Jurnal: {journal_header['journal_code']}</h3>
    
    <div style="background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin-bottom: 20px;">
        <p><strong>Tanggal/Waktu:</strong> {journal_header['entry_timestamp']}</p>
        <p><strong>Deskripsi:</strong> {journal_header['description']}</p>
    </div>
    
    <table style="width: 100%; border-collapse: collapse;">
        <thead>
            <tr style="background-color: #e9ecef;">
                <th style="padding: 10px; border: 1px solid #ddd;">Kode Akun</th>
                <th style="padding: 10px; border: 1px solid #ddd;">Nama Akun</th>
                <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Debit</th>
                <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Kredit</th>
            </tr>
        </thead>
        <tbody>
            {details_html}
        </tbody>
        <tfoot>
            <tr style="background-color: #f2f2f2; font-weight: bold;">
                <td colspan="2" style="padding: 10px; border: 1px solid #ddd; text-align: right;">Total:</td>
                <td style="padding: 10px; border: 1px solid #ddd; text-align: right;">{format_currency(total_debit)}</td>
                <td style="padding: 10px; border: 1px solid #ddd; text-align: right;">{format_currency(total_credit)}</td>
            </tr>
        </tfoot>
    </table>
    
    <div style="margin-top: 20px;">
        <a href="/admin/general-journal" style="text-decoration: none; margin-right: 10px;">
            <input type="button" value="Kembali ke Jurnal Umum" class="btn-blue">
        </a>
        <a href="/admin/all-journals-detail" style="text-decoration: none;">
            <input type="button" value="Lihat Semua Detail Jurnal" class="btn-blue">
        </a>
    </div>
    """
    
    return render_page(f"Detail Jurnal {journal_header['journal_code']}", body, sidebar_content=get_admin_sidebar_html())

def format_currency(amount):
    """Helper function untuk format mata uang Rupiah"""
    if amount is None:
        amount = 0
    return f"Rp {float(amount):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def format_currency(amount):
    """Helper function untuk format mata uang Rupiah"""
    if amount is None:
        amount = 0
    try:
        # Handle Decimal objects
        amount_float = float(amount)
        return f"Rp {amount_float:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError):
        return "Rp 0,00"

# ==================================
# BUKU BESAR
# ==================================
def generate_ledger_html(account_code, db):
    """
    Helper function untuk mengambil data dan merender HTML untuk SATU buku besar.
    (Versi ini menggunakan header deskripsi, bukan tabel)
    """
    
    # --- (Bagian 1: Pengambilan Data - Tidak Berubah) ---
    acc_name_row = db.execute("SELECT account_name, account_type FROM chart_of_accounts WHERE account_code = ?", (account_code,)).fetchone()
    if not acc_name_row:
        return f"<p>Error: Akun {account_code} tidak ditemukan.</p>"
    
    acc_name = acc_name_row['account_name']
    acc_type = acc_name_row['account_type']

    is_normal_debit = acc_type in ('Aset Lancar', 'Aset Tetap', 'Beban')
    pos_saldo = "Debit" if is_normal_debit else "Kredit"
    
    saldo_awal_data = db.execute(
        "SELECT (SUM(debit) - SUM(credit)) as balance FROM journal_details WHERE account_code = ? AND entry_id = 1", 
        (account_code,)
    ).fetchone()
    
    saldo_awal = saldo_awal_data['balance'] if saldo_awal_data['balance'] else 0.0
    
    query = """
    SELECT j.entry_timestamp, j.description, d.debit, d.credit, j.id as entry_id
    FROM journal_details d
    JOIN journal_entries j ON d.entry_id = j.id
    WHERE d.account_code = ?
    ORDER BY j.entry_timestamp, j.id;
    """
    transactions = db.execute(query, (account_code,)).fetchall()
    
    table_rows = ""
    saldo = 0.0
    
    for trx in transactions:
        entry_datetime_obj = datetime.strptime(trx['entry_timestamp'], '%Y-%m-%d %H:%M:%S')
        entry_date = entry_datetime_obj.strftime('%d-%m-%Y')
        description = trx['description']
        if trx['entry_id'] == 1:
            description = "<b>Saldo Awal</b>"
            
        if is_normal_debit:
            saldo += trx['debit']
            saldo -= trx['credit']
        else:
            saldo -= trx['debit']
            saldo += trx['credit']
            
        table_rows += f"""
        <tr>
            <td>{entry_date}</td> 
            <td>{description}</td>
            <td class="currency">{format_currency(trx['debit'])}</td>
            <td class="currency">{format_currency(trx['credit'])}</td>
            <td class="currency">{format_currency(saldo)}</td>
        </tr>
        """
    # --- (Akhir Bagian 1) ---
        
    # --- [BAGIAN BARU] BUAT HEADER DESKRIPSI (Ganti Tabel) ---
    header_html = f"""
    <style>
        .ledger-header-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;  /* 2 kolom */
            grid-template-rows: auto auto;   /* 2 baris */
            gap: 8px 15px; /* Jarak baris, Jarak kolom */
            border: 1px solid #ccc;
            padding: 15px;
            margin-bottom: 20px;
            background-color: #f9f9f9;
            border-radius: 5px;
        }}
        .header-item {{
            display: flex;
        }}
        .header-item .label {{
            font-weight: bold;
            width: 110px; /* Lebar label tetap (misal: "Nama Akun") */
        }}
        .header-item .value {{
            flex: 1;
        }}
        
        /* --- PERUBAHAN DI SINI --- */
        .header-item .saldo-awal {{
            font-weight: bold;
            font-family: 'Courier New', Courier;
            flex: 1;
            text-align: left; /* Diubah dari 'right' */
            /* padding-right: 10px; Dihapus */
        }}
        /* --- AKHIR PERUBAHAN --- */
        
        .report-table {{ width: 100%; border-collapse: collapse; }}
        .report-table th, .report-table td {{ padding: 8px; border: 1px solid #ddd; }}
        .report-table th {{ background-color: #f2f2f2; }}
        .total-row {{ font-weight: bold; background-color: #f2f2f2; border-top: 2px solid #333; }}
        .currency {{ text-align: right; font-family: 'Courier New', Courier; }}
    </style>
    
    <div class="ledger-header-grid">
        <div class="header-item">
            <span class="label">Kode Akun</span>
            <span class="value">: {account_code}</span>
        </div>
        <div class="header-item">
            <span class="label">Pos Saldo</span>
            <span class="value">: {pos_saldo}</span>
        </div>
        
        <div class="header-item">
            <span class="label">Nama Akun</span>
            <span class="value">: {acc_name}</span>
        </div>
        <div class="header-item">
            <span class="label">Saldo Awal</span>
            <span class="saldo-awal">: {format_currency(saldo_awal)}</span>
        </div>
    </div>
    """
    # --- [AKHIR BAGIAN BARU] ---

    # --- (Bagian 3: Pembuatan Tabel Utama - PERUBAHAN DI SINI) ---
    ledger_html = f"""
    {header_html}
    <table class="report-table">
        <thead>
            <tr>
                <th style="width: 15%; text-align: center;">Tanggal</th>
                <th style="text-align: center;">Deskripsi</th>
                <th style="width: 20%; text-align: center;">Debit</th>
                <th style="width: 20%; text-align: center;">Kredit</th>
                <th style="width: 20%; text-align: center;">Saldo</th>
            </tr>
        </thead>
        <tbody>
            {table_rows}
        </tbody>
        <tfoot>
            <tr class="total-row">
                <td colspan="4" style="text-align: right;"><b>Saldo Akhir</b></td>
                <td class="currency"><b>{format_currency(saldo)}</b></td>
            </tr>
        </tfoot>
    </table>
    """
    return ledger_html
    
def format_currency(value):
    """
    Helper untuk memformat angka menjadi format mata uang Rupiah
    Contoh: 15000.0 -> Rp 15.000,00
    Contoh: -5000.0 -> (Rp 5.000,00)
    """
    # Gunakan f-string formatting untuk pemisah ribuan (,) dan 2 desimal (.)
    # Ganti koma dan titik untuk standar Indonesia
    formatted = "{:,.2f}".format(value).replace(",", "X").replace(".", ",").replace("X", ".")
    
    if value < 0:
        return f'(Rp {formatted.replace("-", "")})'
    else:
        return f'Rp {formatted}'

def get_net_income(db):
    """
    Helper untuk menghitung Laba/Rugi Bersih saat ini.
    Mengembalikan angka (float).
    """
    # 1. Ambil semua total Pendapatan
    pendapatan_query = """
    SELECT SUM(d.credit) - SUM(d.debit) as total
    FROM journal_details d
    JOIN chart_of_accounts c ON d.account_code = c.account_code
    WHERE c.account_type = 'Pendapatan'
    """
    pendapatan_data = db.execute(pendapatan_query).fetchone()
    total_pendapatan = pendapatan_data['total'] if pendapatan_data['total'] else 0.0

    # 2. Ambil semua total Beban
    beban_query = """
    SELECT SUM(d.debit) - SUM(d.credit) as total
    FROM journal_details d
    JOIN chart_of_accounts c ON d.account_code = c.account_code
    WHERE c.account_type = 'Beban'
    """
    beban_data = db.execute(beban_query).fetchone()
    total_beban = beban_data['total'] if beban_data['total'] else 0.0
    
    # 3. Hitung Laba/Rugi
    laba_rugi_bersih = total_pendapatan - total_beban
    return laba_rugi_bersih

@app.route("/admin/ledger")
def ledger():
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    # Dapatkan kode akun dari URL (?code=...)
    account_code = request.args.get('code')
    
    # KONDISI 1: TIDAK ADA KODE AKUN (Tampilkan halaman pilihan)
    if not account_code:
        accounts = db.execute("SELECT account_code, account_name, account_type FROM chart_of_accounts ORDER BY account_type, account_code").fetchall()
        
        # --- PERUBAHAN DIMULAI DI SINI ---
        # Tambahkan Tombol "Tampilkan Semua" yang mengarah ke rute baru
        body = f"""
        <a href="{url_for('ledger_all')}" style="text-decoration: none;">
            <input type="button" value="Tampilkan Semua Akun" class="btn-blue">
        </a>
        <hr>
        <h2>Atau, Pilih Akun Satu per Satu:</h2>
        """
        # --- PERUBAHAN SELESAI ---
        
        current_type = ""
        for acc in accounts:
            # Buat grup berdasarkan Tipe Akun
            if acc['account_type'] != current_type:
                if current_type != "":
                    body += "</ul>" # Tutup list sebelumnya
                body += f"<br><b>{acc['account_type']}</b><ul>" # Mulai list baru
                current_type = acc['account_type']
            
            # Buat tautan ke halaman ini lagi, TAPI dengan parameter code=
            body += f'<li><a href="/admin/ledger?code={acc["account_code"]}">{acc["account_code"]} - {acc["account_name"]}</a></li>'
        
        body += "</ul>" # Tutup list terakhir
        
        return render_page("Pilih Buku Besar", body, sidebar_content=get_admin_sidebar_html())

    # KONDISI 2: ADA KODE AKUN (Tampilkan satu ledger)
    else:
        # Panggil helper function yang baru kita buat
        single_ledger_html = generate_ledger_html(account_code, db)
        
        body = f"""
        <p><a href="/admin/ledger">&larr; Kembali ke Pilihan Akun</a></p>
        {single_ledger_html}
        """
        
        # Ambil nama akun lagi hanya untuk judul halaman
        acc_name_row = db.execute("SELECT account_name FROM chart_of_accounts WHERE account_code = ?", (account_code,)).fetchone()
        acc_name = acc_name_row['account_name'] if acc_name_row else "Error"
        
        return render_page(f"Buku Besar - {acc_name}", body, sidebar_content=get_admin_sidebar_html())
    
@app.route("/admin/ledger-all")
def ledger_all():
    """Rute BARU untuk menampilkan SEMUA buku besar sekaligus."""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))

    db = get_db()
    # Ambil SEMUA akun
    accounts = db.execute("SELECT account_code FROM chart_of_accounts ORDER BY account_code").fetchall()
    
    # Mulai body HTML
    all_ledgers_html = f"""
    <p><a href="{url_for('ledger')}">&larr; Kembali ke Pilihan Akun</a></p>
    <hr>
    """
    
    # Loop melalui setiap akun
    for acc in accounts:
        account_code = acc['account_code']
        # Panggil helper function untuk setiap akun
        all_ledgers_html += generate_ledger_html(account_code, db)
        all_ledgers_html += "<hr>" # Beri pemisah antar buku besar

    # Render halaman penuh
    return render_page("Semua Buku Besar", all_ledgers_html, sidebar_content=get_admin_sidebar_html())

# ==================================
#  BUKU BESAR PEMBANTU
# ==================================
def ensure_ledger_tables():
    """Pastikan tabel ledger sudah dibuat sebelum operasi"""
    db = get_db()
    
    try:
        # Cek apakah tabel ledger_ar sudah ada
        db.execute("SELECT 1 FROM ledger_ar LIMIT 1")
        return True
    except:
        # Jika tabel belum ada, buat sekarang
        print("üì¶ Membuat tabel ledger...")
        return init_ledger_separate()

@app.route("/admin/ledger-ar", methods=['GET', 'POST'])
def ledger_ar():
    """Buku Besar Pembantu Piutang (Account Receivable)"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    # Pastikan tabel sudah dibuat
    if not ensure_ledger_tables():
        return redirect(url_for('init_ledger_separate_route'))
    
    db = get_db()
    
    # Handle POST request (input transaksi piutang)
    if request.method == 'POST':
        try:
            transaction_date = request.form['transaction_date']
            debtor_name = request.form['debtor_name']
            address = request.form.get('address', '')
            phone = request.form.get('phone', '')
            description = request.form['description']
            reference = request.form['reference']
            debit = float(request.form['debit'] or 0)
            credit = float(request.form['credit'] or 0)
            
            # Validasi
            if debit == 0 and credit == 0:
                return redirect(url_for('ledger_ar', error="Harus mengisi Debit atau Kredit"))
            
            # Simpan ke database AR
            with db:
                db.execute(
                    """INSERT INTO ledger_ar 
                    (transaction_date, debtor_name, address, phone, description, reference, debit, credit) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (transaction_date, debtor_name, address, phone, description, reference, debit, credit)
                )
            
            return redirect(url_for('ledger_ar', success="Transaksi piutang berhasil disimpan!"))
            
        except Exception as e:
            return redirect(url_for('ledger_ar', error=f"Error: {str(e)}"))
    
    # GET request - tampilkan form dan data
    try:
        # Ambil semua transaksi piutang dan group by debtor_name
        ar_transactions = db.execute(
            "SELECT * FROM ledger_ar ORDER BY debtor_name, transaction_date, id"
        ).fetchall()
    except Exception as e:
        # Jika masih error, redirect ke initialization
        return redirect(url_for('init_ledger_separate_route'))
    
    # Handle messages
    success_message = request.args.get('success')
    error_message = request.args.get('error')
    
    # Group transactions by debtor_name
    debtors_data = {}
    for trx in ar_transactions:
        debtor_name = trx['debtor_name']
        if debtor_name not in debtors_data:
            debtors_data[debtor_name] = {
                'transactions': [],
                'address': trx['address'],
                'phone': trx['phone'],
                'total_debit': 0,
                'total_credit': 0,
                'saldo_akhir': 0
            }
        debtors_data[debtor_name]['transactions'].append(trx)
    
    # Hitung total untuk setiap debitur
    for debtor_name, data in debtors_data.items():
        saldo = 0
        for trx in data['transactions']:
            data['total_debit'] += trx['debit']
            data['total_credit'] += trx['credit']
            saldo += trx['debit'] - trx['credit']  # Untuk AR: debit tambah piutang, credit kurang piutang
        data['saldo_akhir'] = saldo
    
    # Format tabel per debitur
    debtors_html = ""
    for debtor_name, data in debtors_data.items():
        transactions_html = ""
        saldo_berjalan = 0
        
        for trx in data['transactions']:
            saldo_berjalan += trx['debit'] - trx['credit']
            date_obj = datetime.strptime(trx['transaction_date'], '%Y-%m-%d')
            formatted_date = date_obj.strftime('%d-%m-%Y')
            
            transactions_html += f"""
            <tr>
                <td>{formatted_date}</td>
                <td>{trx['description']}</td>
                <td>{trx['reference']}</td>
                <td class="currency">{format_currency(trx['debit'])}</td>
                <td class="currency">{format_currency(trx['credit'])}</td>
                <td class="currency">{format_currency(saldo_berjalan)}</td>
            </tr>
            """
        
        # Tentukan pos saldo
        pos_saldo = "Debit" if data['saldo_akhir'] >= 0 else "Kredit"
        
        debtors_html += f"""
        <div class="creditor-section">
            <!-- HEADER seperti gambar 2 -->
            <div class="creditor-header">
                <table class="header-table">
                    <tr>
                        <td class="header-label">Nama Debitur</td>
                        <td class="header-value">{debtor_name}</td>
                        <td class="header-label">Pos Saldo</td>
                        <td class="header-value">{pos_saldo}</td>
                    </tr>
                    <tr>
                        <td class="header-label">Alamat</td>
                        <td class="header-value">{data['address'] or '-'}</td>
                        <td class="header-label">Saldo Awal</td>
                        <td class="header-value currency">Rp 0,00</td>
                    </tr>
                    <tr>
                        <td class="header-label">Telepon</td>
                        <td class="header-value">{data['phone'] or '-'}</td>
                        <td class="header-label"></td>
                        <td class="header-value"></td>
                    </tr>
                </table>
            </div>
            
            <!-- Tabel Transaksi -->
            <table class="ledger-table">
                <thead>
                    <tr>
                        <th>Tanggal</th>
                        <th>Keterangan</th>
                        <th>Referensi</th>
                        <th>Debit</th>
                        <th>Kredit</th>
                        <th>Saldo</th>
                    </tr>
                </thead>
                <tbody>
                    {transactions_html}
                </tbody>
                <tfoot>
                    <tr class="total-row">
                        <td colspan="3" style="text-align: right; font-weight: bold;">Total</td>
                        <td class="currency" style="font-weight: bold;">{format_currency(data['total_debit'])}</td>
                        <td class="currency" style="font-weight: bold;">{format_currency(data['total_credit'])}</td>
                        <td class="currency" style="font-weight: bold;">{format_currency(data['saldo_akhir'])}</td>
                    </tr>
                    <tr class="saldo-akhir-row">
                        <td colspan="5" style="text-align: right; font-weight: bold;">Saldo Akhir</td>
                        <td class="currency" style="font-weight: bold; background-color: #e9ecef;">{format_currency(data['saldo_akhir'])}</td>
                    </tr>
                </tfoot>
            </table>
        </div>
        <div style="margin-bottom: 40px;"></div>
        """

    body = f"""
    <style>
        /* Same styles as AP but with different colors */
        .ledger-container {{
            display: grid;
            grid-template-columns: 400px 1fr;
            gap: 30px;
        }}
        .form-section {{
            background-color: #f8f9fa;
            padding: 25px;
            border-radius: 8px;
            border: 1px solid #dee2e6;
            height: fit-content;
        }}
        .data-section {{
            background: white;
            padding: 0;
        }}
        .form-group {{
            margin-bottom: 20px;
        }}
        .form-label {{
            font-weight: bold;
            display: block;
            margin-bottom: 5px;
            color: #333;
        }}
        .form-input {{
            width: 100%;
            padding: 10px;
            border: 1px solid #ccc;
            border-radius: 4px;
            box-sizing: border-box;
        }}
        .amount-inputs {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 15px;
        }}
        .alert-success {{
            background-color: #d4edda;
            color: #155724;
            padding: 12px;
            border-radius: 4px;
            margin-bottom: 20px;
            border: 1px solid #c3e6cb;
        }}
        .alert-error {{
            background-color: #f8d7da;
            color: #721c24;
            padding: 12px;
            border-radius: 4px;
            margin-bottom: 20px;
            border: 1px solid #f5c6cb;
        }}
        
        /* Style untuk header debitur */
        .creditor-section {{
            margin-bottom: 30px;
            border: 1px solid #ddd;
            border-radius: 8px;
            overflow: hidden;
        }}
        .creditor-header {{
            background-color: #f8f9fa;
            padding: 0;
            border-bottom: 1px solid #ddd;
        }}
        .header-table {{
            width: 100%;
            border-collapse: collapse;
        }}
        .header-table td {{
            padding: 12px 15px;
            border: 1px solid #ddd;
        }}
        .header-label {{
            font-weight: bold;
            background-color: #e9ecef;
            width: 25%;
        }}
        .header-value {{
            width: 25%;
        }}
        
        /* Style untuk tabel transaksi */
        .ledger-table {{
            width: 100%;
            border-collapse: collapse;
        }}
        .ledger-table th {{
            background-color: #f2f2f2;
            padding: 12px;
            border: 1px solid #ddd;
            text-align: left;
            font-weight: bold;
        }}
        .ledger-table td {{
            padding: 10px;
            border: 1px solid #ddd;
        }}
        .currency {{
            text-align: right;
            font-family: 'Courier New', monospace;
        }}
        .total-row {{
            background-color: #f8f9fa;
        }}
        .saldo-akhir-row {{
            background-color: #e9ecef;
            border-top: 2px solid #333;
        }}
        
        .summary-card {{
            background: linear-gradient(135deg, #007bff, #0056b3);
            color: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
    </style>

    <h2>üìó Buku Besar Pembantu Piutang (AR)</h2>
    
    {f'<div class="alert-success">{success_message}</div>' if success_message else ''}
    {f'<div class="alert-error">{error_message}</div>' if error_message else ''}

    <!-- Summary Card -->
    <div class="summary-card">
        <h3 style="margin: 0 0 10px 0; color: white;">Summary Piutang</h3>
        <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 15px;">
            <div>
                <small>Total Debitur</small><br>
                <strong style="font-size: 1.2em;">{len(debtors_data)} debitur</strong>
            </div>
            <div>
                <small>Total Transaksi</small><br>
                <strong style="font-size: 1.2em;">{len(ar_transactions)} transaksi</strong>
            </div>
            <div>
                <small>Total Piutang</small><br>
                <strong style="font-size: 1.2em;">{format_currency(sum(data['saldo_akhir'] for data in debtors_data.values()))}</strong>
            </div>
        </div>
    </div>

    <div class="ledger-container">
        <!-- Form Input -->
        <div class="form-section">
            <h3 style="margin-top: 0; color: #333;">Input Transaksi Piutang</h3>
            
            <form method="POST">
                <div class="form-group">
                    <label class="form-label">Tanggal Transaksi:</label>
                    <input type="date" name="transaction_date" value="{datetime.now().strftime('%Y-%m-%d')}" class="form-input" required>
                </div>

                <div class="form-group">
                    <label class="form-label">Nama Debitur:</label>
                    <input type="text" name="debtor_name" class="form-input" placeholder="Nama lengkap debitur" required>
                </div>

                <div class="form-group">
                    <label class="form-label">Alamat:</label>
                    <input type="text" name="address" class="form-input" placeholder="Alamat debitur (opsional)">
                </div>

                <div class="form-group">
                    <label class="form-label">No. Telepon:</label>
                    <input type="text" name="phone" class="form-input" placeholder="No. telepon (opsional)">
                </div>

                <div class="form-group">
                    <label class="form-label">Keterangan:</label>
                    <input type="text" name="description" class="form-input" placeholder="Contoh: Penjualan kredit, Pembayaran piutang, dll" required>
                </div>

                <div class="form-group">
                    <label class="form-label">No. Referensi:</label>
                    <input type="text" name="reference" class="form-input" placeholder="INV-001, PAY-001, dll" required>
                </div>

                <div class="form-group">
                    <label class="form-label">Jumlah:</label>
                    <div class="amount-inputs">
                        <div>
                            <label class="form-label">Debit (Tambah Piutang):</label>
                            <input type="number" name="debit" step="0.01" min="0" value="0" class="form-input" placeholder="0">
                        </div>
                        <div>
                            <label class="form-label">Kredit (Kurangi Piutang):</label>
                            <input type="number" name="credit" step="0.01" min="0" value="0" class="form-input" placeholder="0">
                        </div>
                    </div>
                    <small style="color: #666;">
                        <strong>Keterangan:</strong><br>
                        ‚Ä¢ Debit = Penambahan piutang (jual kredit)<br>
                        ‚Ä¢ Kredit = Pengurangan piutang (pembayaran)
                    </small>
                </div>

                <div style="text-align: center; margin-top: 25px;">
                    <input type="submit" value="üíæ Simpan Transaksi Piutang" 
                           style="background-color: #007bff; color: white; padding: 12px 24px; border: none; border-radius: 5px; cursor: pointer; font-weight: bold; width: 100%;">
                </div>
            </form>
        </div>

        <!-- Data Transaksi -->
        <div class="data-section">
            <h3>Riwayat Transaksi Piutang</h3>
            
            {debtors_html if debtors_data else '''
            <div style="text-align: center; padding: 40px; color: #666; background-color: #f8f9fa; border-radius: 8px;">
                <h4>üìù Belum ada transaksi piutang</h4>
                <p>Silakan input transaksi pertama menggunakan form di samping.</p>
            </div>
            '''}
        </div>
    </div>
    """
    
    return render_page("Buku Besar Piutang (AR)", body, sidebar_content=get_admin_sidebar_html())

@app.route("/admin/ledger-ap", methods=['GET', 'POST'])
def ledger_ap():
    """Buku Besar Pembantu Utang (Account Payable)"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    # Pastikan tabel sudah dibuat
    if not ensure_ledger_tables():
        return redirect(url_for('init_ledger_separate_route'))
    
    db = get_db()
    
    # Handle POST request (input transaksi utang)
    if request.method == 'POST':
        try:
            transaction_date = request.form['transaction_date']
            creditor_name = request.form['creditor_name']
            address = request.form.get('address', '')
            phone = request.form.get('phone', '')
            description = request.form['description']
            reference = request.form['reference']
            debit = float(request.form['debit'] or 0)
            credit = float(request.form['credit'] or 0)
            
            # Validasi
            if debit == 0 and credit == 0:
                return redirect(url_for('ledger_ap', error="Harus mengisi Debit atau Kredit"))
            
            # Simpan ke database AP
            with db:
                db.execute(
                    """INSERT INTO ledger_ap 
                    (transaction_date, creditor_name, address, phone, description, reference, debit, credit) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (transaction_date, creditor_name, address, phone, description, reference, debit, credit)
                )
            
            return redirect(url_for('ledger_ap', success="Transaksi utang berhasil disimpan!"))
            
        except Exception as e:
            return redirect(url_for('ledger_ap', error=f"Error: {str(e)}"))
    
    # GET request - tampilkan form dan data
    try:
        # Ambil semua transaksi utang dan group by creditor_name
        ap_transactions = db.execute(
            "SELECT * FROM ledger_ap ORDER BY creditor_name, transaction_date, id"
        ).fetchall()
    except Exception as e:
        # Jika masih error, redirect ke initialization
        return redirect(url_for('init_ledger_separate_route'))
    
    # Handle messages
    success_message = request.args.get('success')
    error_message = request.args.get('error')
    
    # Group transactions by creditor_name
    creditors_data = {}
    for trx in ap_transactions:
        creditor_name = trx['creditor_name']
        if creditor_name not in creditors_data:
            creditors_data[creditor_name] = {
                'transactions': [],
                'address': trx['address'],
                'phone': trx['phone'],
                'total_debit': 0,
                'total_credit': 0,
                'saldo_akhir': 0
            }
        creditors_data[creditor_name]['transactions'].append(trx)
    
    # Hitung total untuk setiap kreditur
    for creditor_name, data in creditors_data.items():
        saldo = 0
        for trx in data['transactions']:
            data['total_debit'] += trx['debit']
            data['total_credit'] += trx['credit']
            saldo += trx['credit'] - trx['debit']  # Untuk AP: kredit tambah utang, debit kurang utang
        data['saldo_akhir'] = saldo
    
    # Format tabel per kreditur
    creditors_html = ""
    for creditor_name, data in creditors_data.items():
        transactions_html = ""
        saldo_berjalan = 0
        
        for trx in data['transactions']:
            saldo_berjalan += trx['credit'] - trx['debit']
            date_obj = datetime.strptime(trx['transaction_date'], '%Y-%m-%d')
            formatted_date = date_obj.strftime('%d-%m-%Y')
            
            transactions_html += f"""
            <tr>
                <td>{formatted_date}</td>
                <td>{trx['description']}</td>
                <td>{trx['reference']}</td>
                <td class="currency">{format_currency(trx['debit'])}</td>
                <td class="currency">{format_currency(trx['credit'])}</td>
                <td class="currency">{format_currency(saldo_berjalan)}</td>
            </tr>
            """
        
        # Tentukan pos saldo
        pos_saldo = "Kredit" if data['saldo_akhir'] >= 0 else "Debit"
        
        creditors_html += f"""
        <div class="creditor-section">
            <!-- HEADER seperti gambar 2 -->
            <div class="creditor-header">
                <table class="header-table">
                    <tr>
                        <td class="header-label">Nama Kreditur</td>
                        <td class="header-value">{creditor_name}</td>
                        <td class="header-label">Pos Saldo</td>
                        <td class="header-value">{pos_saldo}</td>
                    </tr>
                    <tr>
                        <td class="header-label">Alamat</td>
                        <td class="header-value">{data['address'] or '-'}</td>
                        <td class="header-label">Saldo Awal</td>
                        <td class="header-value currency">Rp 0,00</td>
                    </tr>
                    <tr>
                        <td class="header-label">Telepon</td>
                        <td class="header-value">{data['phone'] or '-'}</td>
                        <td class="header-label"></td>
                        <td class="header-value"></td>
                    </tr>
                </table>
            </div>
            
            <!-- Tabel Transaksi -->
            <table class="ledger-table">
                <thead>
                    <tr>
                        <th>Tanggal</th>
                        <th>Keterangan</th>
                        <th>Referensi</th>
                        <th>Debit</th>
                        <th>Kredit</th>
                        <th>Saldo</th>
                    </tr>
                </thead>
                <tbody>
                    {transactions_html}
                </tbody>
                <tfoot>
                    <tr class="total-row">
                        <td colspan="3" style="text-align: right; font-weight: bold;">Total</td>
                        <td class="currency" style="font-weight: bold;">{format_currency(data['total_debit'])}</td>
                        <td class="currency" style="font-weight: bold;">{format_currency(data['total_credit'])}</td>
                        <td class="currency" style="font-weight: bold;">{format_currency(data['saldo_akhir'])}</td>
                    </tr>
                    <tr class="saldo-akhir-row">
                        <td colspan="5" style="text-align: right; font-weight: bold;">Saldo Akhir</td>
                        <td class="currency" style="font-weight: bold; background-color: #e9ecef;">{format_currency(data['saldo_akhir'])}</td>
                    </tr>
                </tfoot>
            </table>
        </div>
        <div style="margin-bottom: 40px;"></div>
        """

    body = f"""
    <style>
        .ledger-container {{
            display: grid;
            grid-template-columns: 400px 1fr;
            gap: 30px;
        }}
        .form-section {{
            background-color: #f8f9fa;
            padding: 25px;
            border-radius: 8px;
            border: 1px solid #dee2e6;
            height: fit-content;
        }}
        .data-section {{
            background: white;
            padding: 0;
        }}
        .form-group {{
            margin-bottom: 20px;
        }}
        .form-label {{
            font-weight: bold;
            display: block;
            margin-bottom: 5px;
            color: #333;
        }}
        .form-input {{
            width: 100%;
            padding: 10px;
            border: 1px solid #ccc;
            border-radius: 4px;
            box-sizing: border-box;
        }}
        .amount-inputs {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 15px;
        }}
        .alert-success {{
            background-color: #d4edda;
            color: #155724;
            padding: 12px;
            border-radius: 4px;
            margin-bottom: 20px;
            border: 1px solid #c3e6cb;
        }}
        .alert-error {{
            background-color: #f8d7da;
            color: #721c24;
            padding: 12px;
            border-radius: 4px;
            margin-bottom: 20px;
            border: 1px solid #f5c6cb;
        }}
        
        /* Style untuk header kreditur */
        .creditor-section {{
            margin-bottom: 30px;
            border: 1px solid #ddd;
            border-radius: 8px;
            overflow: hidden;
        }}
        .creditor-header {{
            background-color: #f8f9fa;
            padding: 0;
            border-bottom: 1px solid #ddd;
        }}
        .header-table {{
            width: 100%;
            border-collapse: collapse;
        }}
        .header-table td {{
            padding: 12px 15px;
            border: 1px solid #ddd;
        }}
        .header-label {{
            font-weight: bold;
            background-color: #e9ecef;
            width: 25%;
        }}
        .header-value {{
            width: 25%;
        }}
        
        /* Style untuk tabel transaksi */
        .ledger-table {{
            width: 100%;
            border-collapse: collapse;
        }}
        .ledger-table th {{
            background-color: #f2f2f2;
            padding: 12px;
            border: 1px solid #ddd;
            text-align: left;
            font-weight: bold;
        }}
        .ledger-table td {{
            padding: 10px;
            border: 1px solid #ddd;
        }}
        .currency {{
            text-align: right;
            font-family: 'Courier New', monospace;
        }}
        .total-row {{
            background-color: #f8f9fa;
        }}
        .saldo-akhir-row {{
            background-color: #e9ecef;
            border-top: 2px solid #333;
        }}
        
        .summary-card {{
            background: linear-gradient(135deg, #28a745, #1e7e34);
            color: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
    </style>

    <h2>üìò Buku Besar Pembantu Utang (AP)</h2>
    
    {f'<div class="alert-success">{success_message}</div>' if success_message else ''}
    {f'<div class="alert-error">{error_message}</div>' if error_message else ''}

    <!-- Summary Card -->
    <div class="summary-card">
        <h3 style="margin: 0 0 10px 0; color: white;">Summary Utang</h3>
        <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 15px;">
            <div>
                <small>Total Kreditur</small><br>
                <strong style="font-size: 1.2em;">{len(creditors_data)} kreditur</strong>
            </div>
            <div>
                <small>Total Transaksi</small><br>
                <strong style="font-size: 1.2em;">{len(ap_transactions)} transaksi</strong>
            </div>
            <div>
                <small>Total Utang</small><br>
                <strong style="font-size: 1.2em;">{format_currency(sum(data['saldo_akhir'] for data in creditors_data.values()))}</strong>
            </div>
        </div>
    </div>

    <div class="ledger-container">
        <!-- Form Input -->
        <div class="form-section">
            <h3 style="margin-top: 0; color: #333;">Input Transaksi Utang</h3>
            
            <form method="POST">
                <div class="form-group">
                    <label class="form-label">Tanggal Transaksi:</label>
                    <input type="date" name="transaction_date" value="{datetime.now().strftime('%Y-%m-%d')}" class="form-input" required>
                </div>

                <div class="form-group">
                    <label class="form-label">Nama Kreditur:</label>
                    <input type="text" name="creditor_name" class="form-input" placeholder="Nama lengkap kreditur" required>
                </div>

                <div class="form-group">
                    <label class="form-label">Alamat:</label>
                    <input type="text" name="address" class="form-input" placeholder="Alamat kreditur (opsional)">
                </div>

                <div class="form-group">
                    <label class="form-label">No. Telepon:</label>
                    <input type="text" name="phone" class="form-input" placeholder="No. telepon (opsional)">
                </div>

                <div class="form-group">
                    <label class="form-label">Keterangan:</label>
                    <input type="text" name="description" class="form-input" placeholder="Contoh: Beli kredit, Pembayaran utang, dll" required>
                </div>

                <div class="form-group">
                    <label class="form-label">No. Referensi:</label>
                    <input type="text" name="reference" class="form-input" placeholder="PO-001, PAY-001, dll" required>
                </div>

                <div class="form-group">
                    <label class="form-label">Jumlah:</label>
                    <div class="amount-inputs">
                        <div>
                            <label class="form-label">Debit (Kurangi Utang):</label>
                            <input type="number" name="debit" step="0.01" min="0" value="0" class="form-input" placeholder="0">
                        </div>
                        <div>
                            <label class="form-label">Kredit (Tambah Utang):</label>
                            <input type="number" name="credit" step="0.01" min="0" value="0" class="form-input" placeholder="0">
                        </div>
                    </div>
                    <small style="color: #666;">
                        <strong>Keterangan:</strong><br>
                        ‚Ä¢ Debit = Pengurangan utang (pembayaran)<br>
                        ‚Ä¢ Kredit = Penambahan utang (beli kredit)
                    </small>
                </div>

                <div style="text-align: center; margin-top: 25px;">
                    <input type="submit" value="üíæ Simpan Transaksi Utang" 
                           style="background-color: #28a745; color: white; padding: 12px 24px; border: none; border-radius: 5px; cursor: pointer; font-weight: bold; width: 100%;">
                </div>
            </form>
        </div>

        <!-- Data Transaksi -->
        <div class="data-section">
            <h3>Riwayat Transaksi Utang</h3>
            
            {creditors_html if creditors_data else '''
            <div style="text-align: center; padding: 40px; color: #666; background-color: #f8f9fa; border-radius: 8px;">
                <h4>üìù Belum ada transaksi utang</h4>
                <p>Silakan input transaksi pertama menggunakan form di samping.</p>
            </div>
            '''}
        </div>
    </div>
    """
    
    return render_page("Buku Besar Utang (AP)", body, sidebar_content=get_admin_sidebar_html())

# ==================================
# ‚ø¶ NERACA SALDO SETELAH BUKU BESAR
# ==================================

@app.route("/admin/trial-balance")
def trial_balance():
    """Menampilkan Neraca Saldo berdasarkan saldo akhir dari Buku Besar (SEBELUM penyesuaian)"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    
    # Parameter periode (opsional)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Query untuk mendapatkan saldo akhir semua akun (SEBELUM penyesuaian)
    query = """
    SELECT 
        coa.account_code,
        coa.account_name,
        coa.account_type,
        -- Hitung saldo akhir menggunakan logika normal balance (SEBELUM penyesuaian)
        -- Exclude jurnal penyesuaian (Penyusutan, Penyesuaian, Adjustment)
        CASE 
            WHEN coa.account_type IN ('Aset Lancar', 'Aset Tetap', 'Beban') THEN
                (COALESCE((
                    SELECT (SUM(debit) - SUM(credit)) 
                    FROM journal_details 
                    WHERE account_code = coa.account_code AND entry_id = 1
                ), 0) +
                COALESCE(SUM(CASE WHEN je.id != 1 AND je.description NOT LIKE '%Penyusutan%' AND je.description NOT LIKE '%Penyesuaian%' AND je.description NOT LIKE '%Adjustment%' THEN jd.debit ELSE 0 END), 0) -
                COALESCE(SUM(CASE WHEN je.id != 1 AND je.description NOT LIKE '%Penyusutan%' AND je.description NOT LIKE '%Penyesuaian%' AND je.description NOT LIKE '%Adjustment%' THEN jd.credit ELSE 0 END), 0))
            ELSE
                (COALESCE((
                    SELECT (SUM(credit) - SUM(debit)) 
                    FROM journal_details 
                    WHERE account_code = coa.account_code AND entry_id = 1
                ), 0) +
                COALESCE(SUM(CASE WHEN je.id != 1 AND je.description NOT LIKE '%Penyusutan%' AND je.description NOT LIKE '%Penyesuaian%' AND je.description NOT LIKE '%Adjustment%' THEN jd.credit ELSE 0 END), 0) -
                COALESCE(SUM(CASE WHEN je.id != 1 AND je.description NOT LIKE '%Penyusutan%' AND je.description NOT LIKE '%Penyesuaian%' AND je.description NOT LIKE '%Adjustment%' THEN jd.debit ELSE 0 END), 0))
        END as ending_balance_before_adjustment
    FROM chart_of_accounts coa
    LEFT JOIN journal_details jd ON coa.account_code = jd.account_code
    LEFT JOIN journal_entries je ON jd.entry_id = je.id
    """
    
    params = []
    
    # Filter berdasarkan tanggal jika ada (exclude penyesuaian)
    if start_date and end_date:
        query += " WHERE (je.entry_timestamp BETWEEN ? AND ? OR je.id = 1) AND je.description NOT LIKE '%Penyusutan%' AND je.description NOT LIKE '%Penyesuaian%' AND je.description NOT LIKE '%Adjustment%'"
        params.extend([f"{start_date} 00:00:00", f"{end_date} 23:59:59"])
    else:
        query += " WHERE je.description NOT LIKE '%Penyusutan%' AND je.description NOT LIKE '%Penyesuaian%' AND je.description NOT LIKE '%Adjustment%'"
    
    query += """
    GROUP BY coa.account_code, coa.account_name, coa.account_type
    HAVING ending_balance_before_adjustment != 0
    ORDER BY 
        CASE coa.account_type
            WHEN 'Aset Lancar' THEN 1
            WHEN 'Aset Tetap' THEN 2
            WHEN 'Liabilitas' THEN 3
            WHEN 'Ekuitas' THEN 4
            WHEN 'Pendapatan' THEN 5
            WHEN 'Beban' THEN 6
            ELSE 7
        END,
        coa.account_code
    """
    
    accounts = db.execute(query, params).fetchall()
    
    # Hitung grand total sesuai normal balance
    grand_total_debit = Decimal('0')
    grand_total_credit = Decimal('0')
    
    for account in accounts:
        balance = Decimal(str(account['ending_balance_before_adjustment'] or '0'))
        account_type = account['account_type']
        
        # Untuk neraca saldo, saldo debit untuk Aset & Beban, saldo kredit untuk Liabilitas, Ekuitas & Pendapatan
        if account_type in ('Aset Lancar', 'Aset Tetap', 'Beban'):
            if balance > 0:
                grand_total_debit += balance
            else:
                grand_total_credit += abs(balance)
        else:  # Liabilitas, Ekuitas, Pendapatan
            if balance > 0:
                grand_total_credit += balance
            else:
                grand_total_debit += abs(balance)
    
    # Format tabel neraca saldo yang disederhanakan
    table_rows = ""
    for account in accounts:
        balance = Decimal(str(account['ending_balance_before_adjustment'] or '0'))
        account_type = account['account_type']
        
        # Tentukan kolom debit dan kredit untuk saldo akhir
        if account_type in ('Aset Lancar', 'Aset Tetap', 'Beban'):
            # Normal balance di debit
            debit_display = format_currency(balance) if balance >= 0 else ""
            credit_display = format_currency(abs(balance)) if balance < 0 else ""
        else:
            # Normal balance di kredit
            debit_display = format_currency(abs(balance)) if balance < 0 else ""
            credit_display = format_currency(balance) if balance >= 0 else ""
        
        table_rows += f"""
        <tr>
            <td style="padding: 8px; border: 1px solid #ddd;">{account['account_code']}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{account['account_name']}</td>
            <td style="padding: 8px; border: 1px solid #ddd;">{account['account_type']}</td>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{debit_display}</td>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{credit_display}</td>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">
                <a href="/admin/ledger?code={account['account_code']}" 
                   style="color: #007bff; text-decoration: none;" 
                   title="Lihat Buku Besar">üìñ</a>
            </td>
        </tr>
        """
    
    # Status balance check
    balance_status = "SEIMBANG" if grand_total_debit == grand_total_credit else "TIDAK SEIMBANG"
    status_color = "#155724" if balance_status == "SEIMBANG" else "#721c24"
    status_bg_color = "#d4edda" if balance_status == "SEIMBANG" else "#f8d7da"
    
    body = f"""
    <style>
        .trial-balance-table th {{
            background-color: #e9ecef;
            font-weight: bold;
        }}
        .total-row {{
            background-color: #f8f9fa;
            font-weight: bold;
            border-top: 2px solid #333;
        }}
        .balance-check {{
            background-color: {status_bg_color};
            padding: 10px;
            border-radius: 5px;
            margin: 10px 0;
            text-align: center;
            font-weight: bold;
        }}
        .account-link {{
            color: #007bff;
            text-decoration: none;
        }}
        .account-link:hover {{
            text-decoration: underline;
        }}
    </style>
    
    <p>Laporan ini menunjukkan saldo akhir semua akun berdasarkan data dari Buku Besar <strong>sebelum jurnal penyesuaian dan jurnal penutup</strong>.</p>
    
    
    <div class="balance-check">
        Status Balance: <span style="color: {status_color};"><b>{balance_status}</b></span>
        {f' - Selisih: {format_currency(abs(grand_total_debit - grand_total_credit))}' if grand_total_debit != grand_total_credit else ''}
    </div>
    
    <table class="trial-balance-table" style="width: 100%; border-collapse: collapse; margin-top: 20px;">
        <thead>
            <tr>
                <th style="padding: 10px; border: 1px solid #ddd;">Kode Akun</th>
                <th style="padding: 10px; border: 1px solid #ddd;">Nama Akun</th>
                <th style="padding: 10px; border: 1px solid #ddd;">Tipe Akun</th>
                <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Saldo Debit</th>
                <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Saldo Kredit</th>
                <th style="padding: 10px; border: 1px solid #ddd; text-align: center;">Buku Besar</th>
            </tr>
        </thead>
        <tbody>
            {table_rows if table_rows else '<tr><td colspan="6" style="text-align: center; padding: 20px;">Tidak ada saldo pada periode yang dipilih.</td></tr>'}
        </tbody>
        <tfoot>
            <tr class="total-row">
                <td colspan="3" style="padding: 10px; border: 1px solid #ddd; text-align: right;"><strong>TOTAL:</strong></td>
                <td style="padding: 10px; border: 1px solid #ddd; text-align: right;"><strong>{format_currency(grand_total_debit)}</strong></td>
                <td style="padding: 10px; border: 1px solid #ddd; text-align: right;"><strong>{format_currency(grand_total_credit)}</strong></td>
                <td style="padding: 10px; border: 1px solid #ddd;"></td>
            </tr>
        </tfoot>
    </table>
    
    <div style="margin-top: 30px; background-color: #f8f9fa; padding: 15px; border-radius: 5px;">
        <h4>Keterangan:</h4>
        <ul>
            <li><strong>Normal Balance:</strong> 
                <ul>
                    <li><strong>Debit:</strong> Aset Lancar, Aset Tetap, Beban</li>
                    <li><strong>Kredit:</strong> Liabilitas, Ekuitas, Pendapatan</li>
                </ul>
            </li>
            <li><strong>Ikon Buku Besar (üìñ):</strong> Klik untuk melihat detail transaksi di Buku Besar</li>
            <li><strong>Status Balance:</strong> Total Saldo Debit harus sama dengan Total Saldo Kredit</li>
            <li><strong>Catatan Penting:</strong> Data ini belum termasuk jurnal penyesuaian seperti penyusutan aset tetap</li>
        </ul>
    </div>
    """
    
    return render_page("Neraca Saldo", body, sidebar_content=get_admin_sidebar_html())

# ==================================
# JURNAL PENYESUAIAN
# ==================================
@app.route("/admin/adjusting-entries", methods=['GET', 'POST'])
def adjusting_entri():
    """Halaman untuk menginput jurnal penyesuaian dan penyusutan"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    error_message = None
    success_message = None
    
    # LOGIKA POST (PENGOLAHAN DATA DATABASE)
    if request.method == 'POST':
        try:
            with db:
                # --- A. Handle Perhitungan Penyusutan Otomatis ---
                if 'calculate_depreciation' in request.form:
                    depreciation_date = request.form.get('depreciation_date')
                    
                    if not depreciation_date:
                        error_message = "Tanggal penyusutan harus diisi"
                    else:
                        # Ambil aset tetap
                        aset_tetap_data = db.execute("""
                            SELECT 
                                jd.account_code,
                                coa.account_name,
                                SUM(jd.debit) as nilai_aset
                            FROM journal_details jd
                            JOIN chart_of_accounts coa ON jd.account_code = coa.account_code
                            JOIN journal_entries je ON jd.entry_id = je.id
                            WHERE je.description LIKE '%Saldo Awal%'
                            AND coa.account_type = 'Aset Tetap'
                            AND coa.account_name NOT LIKE '%akumulasi%'
                            GROUP BY jd.account_code, coa.account_name
                        """).fetchall()
                        
                        total_depreciation_entries = 0
                        depreciation_details = []
                        
                        for aset in aset_tetap_data:
                            account_code = aset['account_code']
                            account_name = aset['account_name']
                            nilai_aset = Decimal(str(aset['nilai_aset'] or '0'))
                            
                            if nilai_aset == 0 or 'tanah' in account_name.lower():
                                continue
                            
                            masa_manfaat_key = f"depreciation_life_{account_code}"
                            masa_manfaat = int(request.form.get(masa_manfaat_key, 0))
                            
                            if masa_manfaat > 0 and nilai_aset > 0:
                                penyusutan_per_tahun = nilai_aset / masa_manfaat
                                journal_code = generate_journal_code(db, datetime.strptime(depreciation_date, '%Y-%m-%d'))
                                
                                cursor = db.execute(
                                    "INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                    (journal_code, f"{depreciation_date} 00:00:00", f"Penyusutan {account_name} - {depreciation_date}")
                                )
                                entry_id = cursor.lastrowid
                                
                                # Mapping Akun Manual
                                beban_penyusutan_code = ""
                                akumulasi_penyusutan_code = ""
                                
                                # Logika mapping sederhana (bisa disesuaikan)
                                if account_code == '1201':   # Peralatan
                                    beban_penyusutan_code = '6201'; akumulasi_penyusutan_code = '1202'
                                elif account_code == '1203': # Kendaraan
                                    beban_penyusutan_code = '6202'; akumulasi_penyusutan_code = '1204'
                                elif account_code == '1205': # Bangunan
                                    beban_penyusutan_code = '6203'; akumulasi_penyusutan_code = '1206'
                                elif account_code == '1207': # tambahan
                                    beban_penyusutan_code = '6204'; akumulasi_penyusutan_code = '1208'
                                elif account_code == '1209': # tambahan
                                    beban_penyusutan_code = '6206'; akumulasi_penyusutan_code = '1210'
                                elif account_code == '1211': # tambahan
                                    beban_penyusutan_code = '6206'; akumulasi_penyusutan_code = '1212'
                                else:
                                    # Fallback untuk akun tambahan, asumsikan pola +1
                                    continue 

                                # Cek akun beban ada atau tidak
                                beban_akun = db.execute("SELECT account_name FROM chart_of_accounts WHERE account_code = ?", (beban_penyusutan_code,)).fetchone()
                                if not beban_akun: continue # Skip jika akun beban tidak ada
                                
                                # Debit Beban
                                db.execute("INSERT INTO journal_details (entry_id, account_code, debit, credit) VALUES (?, ?, ?, ?)", 
                                           (entry_id, beban_penyusutan_code, float(penyusutan_per_tahun), 0.0))
                                # Kredit Akumulasi
                                db.execute("INSERT INTO journal_details (entry_id, account_code, debit, credit) VALUES (?, ?, ?, ?)", 
                                           (entry_id, akumulasi_penyusutan_code, 0.0, float(penyusutan_per_tahun)))
                                
                                total_depreciation_entries += 1
                                depreciation_details.append({'nama_aset': account_name, 'penyusutan': penyusutan_per_tahun})
                        
                        if total_depreciation_entries > 0:
                            success_message = f"‚úÖ Berhasil membuat {total_depreciation_entries} jurnal penyusutan."
                        elif not error_message:
                            error_message = "‚ùå Tidak ada jurnal penyusutan yang dibuat."

                # --- B. Handle Jurnal Penyesuaian Manual ---
                elif 'add_adjusting_entry' in request.form:
                    entry_date = request.form.get('entry_date')
                    description = request.form.get('description')
                    
                    if not entry_date or not description:
                        error_message = "Tanggal dan keterangan jurnal harus diisi"
                    else:
                        journal_code = generate_journal_code(db, datetime.strptime(entry_date, '%Y-%m-%d'))
                        cursor = db.execute(
                            "INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                            (journal_code, f"{entry_date} 00:00:00", f"Penyesuaian: {description}")
                        )
                        entry_id = cursor.lastrowid
                        
                        debit_total = Decimal('0')
                        credit_total = Decimal('0')
                        
                        # Loop input Debit (menggunakan nama dari JS: debit_amount_1, dst)
                        # Kita loop range aman, misal 1 sampai 50
                        for i in range(1, 51):
                            debit_account = request.form.get(f'debit_account_{i}')
                            debit_amount = request.form.get(f'debit_amount_{i}')
                            
                            if debit_account and debit_amount:
                                val = Decimal(str(debit_amount))
                                if val > 0:
                                    db.execute("INSERT INTO journal_details (entry_id, account_code, debit, credit) VALUES (?, ?, ?, ?)", 
                                               (entry_id, debit_account, float(val), 0.0))
                                    debit_total += val
                        
                        # Loop input Kredit
                        for j in range(1, 51):
                            credit_account = request.form.get(f'credit_account_{j}')
                            credit_amount = request.form.get(f'credit_amount_{j}')
                            
                            if credit_account and credit_amount:
                                val = Decimal(str(credit_amount))
                                if val > 0:
                                    db.execute("INSERT INTO journal_details (entry_id, account_code, debit, credit) VALUES (?, ?, ?, ?)", 
                                               (entry_id, credit_account, 0.0, float(val)))
                                    credit_total += val
                            
                        if debit_total != credit_total:
                            # Raise error to trigger rollback
                            raise Exception(f"Jurnal tidak balance! Debit: {format_currency(debit_total)}, Kredit: {format_currency(credit_total)}")
                        else:
                            success_message = f"‚úÖ Jurnal penyesuaian berhasil disimpan! Total: {format_currency(debit_total)}"

        except Exception as e:
            error_message = f"‚ùå Terjadi kesalahan: {str(e)}"
            try: db.rollback() 
            except: pass

    # LOGIKA GET (PERSIAPAN TAMPILAN HTML)
    # 1. Data Aset
    aset_tetap_data = []
    try:
        aset_tetap_data = db.execute("""
            SELECT jd.account_code, coa.account_name, SUM(jd.debit) as nilai_aset
            FROM journal_details jd
            JOIN chart_of_accounts coa ON jd.account_code = coa.account_code
            JOIN journal_entries je ON jd.entry_id = je.id
            WHERE je.description LIKE '%Saldo Awal%' AND coa.account_type = 'Aset Tetap' AND coa.account_name NOT LIKE '%akumulasi%'
            GROUP BY jd.account_code, coa.account_name HAVING SUM(jd.debit) > 0
        """).fetchall()
    except: pass

    # 2. Data Akun (untuk dropdown)
    accounts = []
    try:
        accounts = db.execute("SELECT account_code, account_name FROM chart_of_accounts ORDER BY account_code").fetchall()
    except: pass
    
    # Buat string option HTML sekali saja untuk dipakai di Python & JS
    acc_opts = "".join([f'<option value="{a["account_code"]}">{a["account_code"]} - {a["account_name"]}</option>' for a in accounts])

    # 3. Form Penyusutan Otomatis
    aset_penyusutan_map = {
        '1201': {'nama': 'Peralatan', 'default_masa_manfaat': 5, 'beban_nama': 'Beban Penyusutan Peralatan'},
        '1203': {'nama': 'Kendaraan', 'default_masa_manfaat': 8, 'beban_nama': 'Beban Penyusutan Kendaraan'},
        '1205': {'nama': 'Bangunan', 'default_masa_manfaat': 20, 'beban_nama': 'Beban Penyusutan Bangunan'},
        '1301': {'nama': 'Tanah', 'default_masa_manfaat': 0, 'beban_nama': None}
    }

    depreciation_form = ""
    if aset_tetap_data:
        depreciation_items = ""
        for aset in aset_tetap_data:
            acc_code = aset['account_code']
            acc_name = aset['account_name']
            val_aset = Decimal(str(aset['nilai_aset'] or '0'))
            
            if 'tanah' in acc_name.lower(): continue
            
            conf = aset_penyusutan_map.get(acc_code, {})
            def_life = conf.get('default_masa_manfaat', 5)
            beban_nm = conf.get('beban_nama', 'Beban Penyusutan')
            penyusutan = val_aset / def_life if def_life > 0 else 0
            
            depreciation_items += f"""
            <div style="background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin-bottom: 15px; border-left: 4px solid #17a2b8;">
                <div style="display: grid; grid-template-columns: 2fr 1fr 1fr 1fr; gap: 15px; align-items: center;">
                    <div><strong style="color: #495057;">{acc_name}</strong><br><small style="color: #6c757d;">Nilai: {format_currency(val_aset)}</small></div>
                    <div>
                        <label style="font-size: 0.9em; font-weight: bold; color: #495057;">Masa Manfaat (thn):</label>
                        <input type="number" name="depreciation_life_{acc_code}" value="{def_life}" min="1" max="100" 
                               style="width: 100%; padding: 8px; border: 1px solid #ced4da; border-radius: 4px;"
                               onchange="hitungPenyusutanOtomatis('{acc_code}', {float(val_aset)})">
                    </div>
                    <div>
                        <label style="font-size: 0.9em; font-weight: bold; color: #495057;">Penyusutan/Thn:</label>
                        <div id="depreciation_amount_{acc_code}" style="font-weight: bold; color: #dc3545; font-size: 1.1em;">{format_currency(penyusutan)}</div>
                    </div>
                    <div>
                         <label style="font-size: 0.9em; font-weight: bold; color: #495057;">Jurnal:</label>
                         <div style="font-size: 0.8em; color: #6c757d;"><span style="color: #28a745;">‚ñ† {beban_nm} (D)</span><br><span style="color: #dc3545;">‚ñ† Akm. Penyu. (K)</span></div>
                    </div>
                </div>
            </div>
            """
        
        depreciation_form = f"""
        <div class="depreciation-form" style="background-color: white; padding: 25px; border-radius: 8px; border: 2px solid #007bff; margin-bottom: 30px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
            <h3 style="color: #007bff; border-bottom: 2px solid #007bff; padding-bottom: 10px; margin-bottom: 20px;">üîÑ Penyusutan Aset Tetap Otomatis</h3>
            <form method="POST" action="/admin/adjusting-entries" onsubmit="return validateDepreciationForm()">
                <div style="margin-bottom: 25px; background-color: #e7f3ff; padding: 15px; border-radius: 5px;">
                    <label style="font-weight: bold; color: #0056b3;">üìÖ Tanggal Penyusutan:</label>
                    <input type="date" name="depreciation_date" required value="{datetime.now().strftime('%Y-%m-%d')}"
                           style="padding: 10px; border: 1px solid #007bff; border-radius: 4px; margin-left: 10px;">
                </div>
                {depreciation_items}
                <div style="margin-top: 25px; text-align: center;">
                    <button type="submit" name="calculate_depreciation" style="background: linear-gradient(135deg, #17a2b8, #138496); color: white; padding: 15px 40px; border: none; border-radius: 5px; cursor: pointer; font-weight: bold;">üìù Buat Jurnal Penyusutan</button>
                </div>
            </form>
        </div>
        """
    else:
        depreciation_form = "<div style='background:#fff3cd; padding:20px; text-align:center;'>‚ö†Ô∏è Belum ada data aset tetap.</div>"

    # 4. Form Manual Adjustment (Menggunakan JS yang Anda berikan, disesuaikan sintaks Pythonnya)
    manual_form = f"""
    <div class="manual-adjustment-form" style="background-color: white; padding: 25px; border-radius: 8px; border: 2px solid #28a745; margin-bottom: 30px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
        <h3 style="color: #28a745; border-bottom: 2px solid #28a745; padding-bottom: 10px; margin-bottom: 20px;">üìù Jurnal Penyesuaian Manual</h3>
        <form method="POST" action="/admin/adjusting-entries" onsubmit="return validateManualForm()">
            <div style="display: grid; grid-template-columns: 1fr 2fr; gap: 20px; margin-bottom: 25px;">
                <div><label style="font-weight:bold; display:block; margin-bottom:5px;">üìÖ Tanggal:</label><input type="date" name="entry_date" required value="{datetime.now().strftime('%Y-%m-%d')}" style="width:100%; padding:10px; border:1px solid #ced4da;"></div>
                <div><label style="font-weight:bold; display:block; margin-bottom:5px;">üìã Keterangan:</label><input type="text" name="description" required placeholder="Contoh: Penyesuaian perlengkapan" style="width:100%; padding:10px; border:1px solid #ced4da;"></div>
            </div>

            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 30px;">
                <div>
                    <h4 style="color: #28a745; border-bottom: 2px solid #28a745;">Debit</h4>
                    <div id="debit-entries">
                        <div class="debit-entry" style="margin-bottom: 15px; padding: 15px; background-color: #f8f9fa; border-radius: 5px; border-left: 4px solid #28a745;">
                            <div style="margin-bottom: 10px;">
                                <label style="display:block; margin-bottom:5px; font-weight:bold; color:#495057;">Akun Debit:</label>
                                <select name="debit_account_1" required style="width:100%; padding:8px; border:1px solid #ced4da;">
                                    <option value="">Pilih Akun Debit</option>{acc_opts}
                                </select>
                            </div>
                            <div>
                                <label style="display:block; margin-bottom:5px; font-weight:bold; color:#495057;">Jumlah:</label>
                                <input type="number" name="debit_amount_1" required step="0.01" min="0" placeholder="0.00" style="width:100%; padding:8px; border:1px solid #ced4da;">
                            </div>
                        </div>
                    </div>
                    <button type="button" onclick="addDebitEntry()" style="background:#28a745; color:white; border:none; padding:8px 15px; border-radius:4px; cursor:pointer;">‚ûï Tambah Debit</button>
                </div>
                
                <div>
                    <h4 style="color: #dc3545; border-bottom: 2px solid #dc3545;">Kredit</h4>
                    <div id="credit-entries">
                        <div class="credit-entry" style="margin-bottom: 15px; padding: 15px; background-color: #f8f9fa; border-radius: 5px; border-left: 4px solid #dc3545;">
                            <div style="margin-bottom: 10px;">
                                <label style="display:block; margin-bottom:5px; font-weight:bold; color:#495057;">Akun Kredit:</label>
                                <select name="credit_account_1" required style="width:100%; padding:8px; border:1px solid #ced4da;">
                                    <option value="">Pilih Akun Kredit</option>{acc_opts}
                                </select>
                            </div>
                            <div>
                                <label style="display:block; margin-bottom:5px; font-weight:bold; color:#495057;">Jumlah:</label>
                                <input type="number" name="credit_amount_1" required step="0.01" min="0" placeholder="0.00" style="width:100%; padding:8px; border:1px solid #ced4da;">
                            </div>
                        </div>
                    </div>
                    <button type="button" onclick="addCreditEntry()" style="background:#dc3545; color:white; border:none; padding:8px 15px; border-radius:4px; cursor:pointer;">‚ûï Tambah Kredit</button>
                </div>
            </div>

            <div style="margin-top: 25px; text-align: center;">
                <button type="submit" name="add_adjusting_entry" style="background: linear-gradient(135deg, #28a745, #20c997); color: white; padding: 15px 40px; border: none; border-radius: 5px; cursor: pointer; font-weight: bold;">üíæ Simpan Jurnal Penyesuaian</button>
            </div>
        </form>
    </div>
    """

    # 5. Recent Adjustments Table
    recent_table_rows = ""
    try:
        recent_entries = db.execute("""
            SELECT je.id as entry_id, je.journal_code, je.entry_timestamp, je.description, jd.account_code, coa.account_name, jd.debit, jd.credit
            FROM journal_entries je JOIN journal_details jd ON je.id = jd.entry_id JOIN chart_of_accounts coa ON jd.account_code = coa.account_code
            WHERE je.description LIKE '%Penyesuaian%' OR je.description LIKE '%Penyusutan%'
            ORDER BY je.entry_timestamp DESC, je.id DESC LIMIT 50
        """).fetchall()
        
        current_id = None
        for row in recent_entries:
            if current_id != row['entry_id']:
                recent_table_rows += f"""<tr style="background:#e7f3ff;"><td colspan="4" style="font-weight:bold; padding:10px; color:#0056b3;">üìÖ {row['entry_timestamp'][:10]} - {row['journal_code']} - {row['description']}</td></tr>"""
                current_id = row['entry_id']
            
            d = format_currency(row['debit']) if row['debit'] > 0 else ''
            c = format_currency(row['credit']) if row['credit'] > 0 else ''
            recent_table_rows += f"""<tr><td style="padding:10px; border:1px solid #eee;">{row['account_code']}</td><td>{row['account_name']}</td><td style="text-align:right; color:#28a745;">{d}</td><td style="text-align:right; color:#dc3545;">{c}</td></tr>"""
    except: pass

    recent_adjustments = f"""
    <div style="background-color: white; padding: 25px; border-radius: 8px; border: 2px solid #6f42c1; margin-bottom: 30px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
        <h3 style="color: #6f42c1;">üìã Jurnal Penyesuaian Terbaru</h3>
        <div style="overflow-x: auto;"><table style="width: 100%; border-collapse: collapse; margin-top: 15px;"><thead><tr style="background-color: #6f42c1; color: white;"><th style="padding:12px;">Kode</th><th>Akun</th><th>Debit</th><th>Kredit</th></tr></thead><tbody>{recent_table_rows}</tbody></table></div>
    </div>
    """
    javascript_code = f"""
    <script>
    let debitCount = 1;
    let creditCount = 1;

    function addDebitEntry() {{
        debitCount++;
        const container = document.getElementById('debit-entries');
        const newEntry = document.createElement('div');
        newEntry.className = 'debit-entry';
        newEntry.style = 'margin-bottom: 15px; padding: 15px; background-color: #f8f9fa; border-radius: 5px; border-left: 4px solid #28a745;';
        
        // Perhatikan penggunaan {{acc_opts}} dari Python di bawah ini
        newEntry.innerHTML = `
            <div style="margin-bottom: 10px;">
                <label style="display: block; margin-bottom: 5px; font-weight: bold; color: #495057;">Akun Debit:</label>
                <select name="debit_account_${{debitCount}}" required style="width: 100%; padding: 8px; border: 1px solid #ced4da; border-radius: 4px;">
                    <option value="">Pilih Akun Debit</option>
                    {acc_opts} 
                </select>
            </div>
            <div>
                <label style="display: block; margin-bottom: 5px; font-weight: bold; color: #495057;">Jumlah:</label>
                <input type="number" name="debit_amount_${{debitCount}}" required step="0.01" min="0" placeholder="0.00" style="width: 100%; padding: 8px; border: 1px solid #ced4da; border-radius: 4px;">
            </div>
            <button type="button" onclick="this.parentElement.remove()" style="margin-top: 10px; background-color: #dc3545; color: white; padding: 5px 10px; border: none; border-radius: 3px; cursor: pointer; font-size: 12px;">‚ùå Hapus</button>
        `;
        container.appendChild(newEntry);
    }}

    function addCreditEntry() {{
        creditCount++;
        const container = document.getElementById('credit-entries');
        const newEntry = document.createElement('div');
        newEntry.className = 'credit-entry';
        newEntry.style = 'margin-bottom: 15px; padding: 15px; background-color: #f8f9fa; border-radius: 5px; border-left: 4px solid #dc3545;';
        
        newEntry.innerHTML = `
            <div style="margin-bottom: 10px;">
                <label style="display: block; margin-bottom: 5px; font-weight: bold; color: #495057;">Akun Kredit:</label>
                <select name="credit_account_${{creditCount}}" required style="width: 100%; padding: 8px; border: 1px solid #ced4da; border-radius: 4px;">
                    <option value="">Pilih Akun Kredit</option>
                    {acc_opts}
                </select>
            </div>
            <div>
                <label style="display: block; margin-bottom: 5px; font-weight: bold; color: #495057;">Jumlah:</label>
                <input type="number" name="credit_amount_${{creditCount}}" required step="0.01" min="0" placeholder="0.00" style="width: 100%; padding: 8px; border: 1px solid #ced4da; border-radius: 4px;">
            </div>
            <button type="button" onclick="this.parentElement.remove()" style="margin-top: 10px; background-color: #dc3545; color: white; padding: 5px 10px; border: none; border-radius: 3px; cursor: pointer; font-size: 12px;">‚ùå Hapus</button>
        `;
        container.appendChild(newEntry);
    }}

    function formatCurrency(amount) {{
        return 'Rp ' + parseFloat(amount).toLocaleString('id-ID', {{minimumFractionDigits: 2, maximumFractionDigits: 2}});
    }}

    function hitungPenyusutanOtomatis(accountCode, nilaiAset) {{
        // Mencari input masa manfaat berdasarkan nama input
        const inputName = 'depreciation_life_' + accountCode;
        const inputElement = document.querySelector('input[name="' + inputName + '"]');
        
        if (inputElement) {{
            const masaManfaat = parseFloat(inputElement.value) || 0;
            if (masaManfaat > 0 && nilaiAset > 0) {{
                const penyusutan = nilaiAset / masaManfaat;
                document.getElementById('depreciation_amount_' + accountCode).textContent = formatCurrency(penyusutan);
            }} else {{
                document.getElementById('depreciation_amount_' + accountCode).textContent = formatCurrency(0);
            }}
        }}
    }}

    function validateDepreciationForm() {{
        const dateInput = document.querySelector('input[name="depreciation_date"]');
        if (!dateInput.value) {{
            alert('‚ùå Harap pilih tanggal penyusutan.');
            return false;
        }}
        return confirm('üìù Apakah Anda yakin ingin membuat jurnal penyusutan?');
    }}

    function validateManualForm() {{
        const dateInput = document.querySelector('input[name="entry_date"]');
        const descriptionInput = document.querySelector('input[name="description"]');
        
        if (!dateInput.value || !descriptionInput.value) {{
            alert('‚ùå Tanggal dan keterangan jurnal harus diisi.');
            return false;
        }}

        // Hitung total
        let totalDebit = 0;
        let totalCredit = 0;
        
        document.querySelectorAll('input[name^="debit_amount_"]').forEach(inp => {{
            totalDebit += parseFloat(inp.value) || 0;
        }});
        
        document.querySelectorAll('input[name^="credit_amount_"]').forEach(inp => {{
            totalCredit += parseFloat(inp.value) || 0;
        }});
        
        if (totalDebit === 0 || totalCredit === 0) {{
            alert('‚ùå Nominal tidak boleh nol.');
            return false;
        }}
        
        // Toleransi floating point
        if (Math.abs(totalDebit - totalCredit) > 1) {{
            alert('‚ùå Jurnal tidak balance! \\nDebit: ' + formatCurrency(totalDebit) + '\\nKredit: ' + formatCurrency(totalCredit));
            return false;
        }}
        
        return confirm('üíæ Simpan jurnal penyesuaian ini?');
    }}
    </script>
    """

    # 6. Assemble Final HTML
    message_html = ""
    if error_message:
        message_html = f'<div style="color:#721c24; background-color:#f8d7da; padding:15px; margin-bottom:20px; border-left:4px solid #dc3545; border-radius:5px;">{error_message}</div>'
    if success_message:
        message_html = f'<div style="color:#155724; background-color:#d4edda; padding:15px; margin-bottom:20px; border-left:4px solid #28a745; border-radius:5px;">{success_message}</div>'

    body_content = f"""
    <style>
        .adjusting-container {{ max-width: 100%; margin: 0 auto; }}
        .info-box {{ background-color: #e7f3ff; padding: 20px; border-radius: 8px; margin-bottom: 25px; border-left: 4px solid #007bff; }}
    </style>
    
    <div class="adjusting-container">
        <div class="info-box">
            <h4 style="color: #0056b3; margin-bottom: 10px;">üí° Informasi Jurnal Penyesuaian</h4>
            <p style="margin-bottom: 0; color: #495057;">Halaman ini digunakan untuk mencatat penyusutan aset tetap secara otomatis dan penyesuaian manual lainnya (beban dibayar dimuka, perlengkapan, dll).</p>
        </div>
        
        {message_html}
        {depreciation_form}
        {manual_form}
        {recent_adjustments}
    </div>
    {javascript_code}
    """
    
    return render_page("Jurnal Penyesuaian", body_content, sidebar_content=get_admin_sidebar_html())

# ==================================
# NERACA LAJUR
# ==================================
@app.route("/admin/work-sheet")
def work_sheet():
    """Halaman Neraca Lajur berdasarkan data dari Trial Balance (Sebelum Penyesuaian)"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    
    # Parameter periode (opsional) - sama dengan trial balance
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    try:
        # AMBIL DATA NERACA SALDO DARI TRIAL BALANCE (SEBELUM PENYESUAIAN)
        trial_balance_query = """
        SELECT 
            coa.account_code,
            coa.account_name,
            coa.account_type,
            CASE 
                WHEN coa.account_type IN ('Aset Lancar', 'Aset Tetap', 'Beban') THEN
                    (COALESCE((
                        SELECT (SUM(debit) - SUM(credit)) 
                        FROM journal_details 
                        WHERE account_code = coa.account_code AND entry_id = 1
                    ), 0) +
                    COALESCE(SUM(CASE WHEN je.id != 1 AND je.description NOT LIKE '%Penyusutan%' AND je.description NOT LIKE '%Penyesuaian%' AND je.description NOT LIKE '%Adjustment%' THEN jd.debit ELSE 0 END), 0) -
                    COALESCE(SUM(CASE WHEN je.id != 1 AND je.description NOT LIKE '%Penyusutan%' AND je.description NOT LIKE '%Penyesuaian%' AND je.description NOT LIKE '%Adjustment%' THEN jd.credit ELSE 0 END), 0))
                ELSE
                    (COALESCE((
                        SELECT (SUM(credit) - SUM(debit)) 
                        FROM journal_details 
                        WHERE account_code = coa.account_code AND entry_id = 1
                    ), 0) +
                    COALESCE(SUM(CASE WHEN je.id != 1 AND je.description NOT LIKE '%Penyusutan%' AND je.description NOT LIKE '%Penyesuaian%' AND je.description NOT LIKE '%Adjustment%' THEN jd.credit ELSE 0 END), 0) -
                    COALESCE(SUM(CASE WHEN je.id != 1 AND je.description NOT LIKE '%Penyusutan%' AND je.description NOT LIKE '%Penyesuaian%' AND je.description NOT LIKE '%Adjustment%' THEN jd.debit ELSE 0 END), 0))
            END as ending_balance_before_adjustment
        FROM chart_of_accounts coa
        LEFT JOIN journal_details jd ON coa.account_code = jd.account_code
        LEFT JOIN journal_entries je ON jd.entry_id = je.id
        """
        
        trial_balance_params = []
        
        if start_date and end_date:
            trial_balance_query += " WHERE (je.entry_timestamp BETWEEN ? AND ? OR je.id = 1) AND je.description NOT LIKE '%Penyusutan%' AND je.description NOT LIKE '%Penyesuaian%' AND je.description NOT LIKE '%Adjustment%'"
            trial_balance_params.extend([f"{start_date} 00:00:00", f"{end_date} 23:59:59"])
        else:
            trial_balance_query += " WHERE je.description NOT LIKE '%Penyusutan%' AND je.description NOT LIKE '%Penyesuaian%' AND je.description NOT LIKE '%Adjustment%'"
        
        trial_balance_query += """
        GROUP BY coa.account_code, coa.account_name, coa.account_type
        HAVING ending_balance_before_adjustment != 0
        ORDER BY coa.account_code
        """
        
        neraca_saldo = db.execute(trial_balance_query, trial_balance_params).fetchall()

        # AMBIL DATA JURNAL PENYESUAIAN SECARA TERPISAH
        adjustment_query = """
            SELECT 
                jd.account_code,
                coa.account_name,
                coa.account_type,
                SUM(jd.debit) as adjustment_debit,
                SUM(jd.credit) as adjustment_credit
            FROM journal_details jd
            JOIN chart_of_accounts coa ON jd.account_code = coa.account_code
            JOIN journal_entries je ON jd.entry_id = je.id
            WHERE (je.description LIKE '%Penyusutan%' 
                OR je.description LIKE '%Penyesuaian%'
                OR je.description LIKE '%Adjustment%'
                OR je.description LIKE '%penyusutan%'
                OR je.description LIKE '%penyesuaian%'
                OR je.description LIKE '%adjustment%')
        """
        
        adjustment_params = []
        
        if start_date and end_date:
            adjustment_query += " AND je.entry_timestamp BETWEEN ? AND ?"
            adjustment_params.extend([f"{start_date} 00:00:00", f"{end_date} 23:59:59"])
        
        adjustment_query += """
            GROUP BY jd.account_code, coa.account_name, coa.account_type
            HAVING adjustment_debit != 0 OR adjustment_credit != 0
        """
        
        jurnal_penyesuaian = db.execute(adjustment_query, adjustment_params).fetchall()

        # Buat dictionary untuk data neraca saldo dan penyesuaian
        neraca_saldo_dict = {}
        for item in neraca_saldo:
            balance = Decimal(str(item['ending_balance_before_adjustment'] or '0'))
            account_type = item['account_type']
            
            if account_type in ('Aset Lancar', 'Aset Tetap', 'Beban'):
                debit_balance = balance if balance >= 0 else Decimal('0')
                credit_balance = abs(balance) if balance < 0 else Decimal('0')
            else:
                debit_balance = abs(balance) if balance < 0 else Decimal('0')
                credit_balance = balance if balance >= 0 else Decimal('0')
            
            neraca_saldo_dict[item['account_code']] = {
                'account_name': item['account_name'],
                'account_type': account_type,
                'debit': debit_balance,
                'credit': credit_balance,
                'original_balance': balance
            }

        adjustment_dict = {}
        for item in jurnal_penyesuaian:
            adjustment_dict[item['account_code']] = {
                'adjustment_debit': Decimal(str(item['adjustment_debit'] or '0')),
                'adjustment_credit': Decimal(str(item['adjustment_credit'] or '0')),
                'account_name': item['account_name'],
                'account_type': item['account_type']
            }

        # PROSES NERACA LAJUR
        work_sheet_data = []
        
        # Inisialisasi Total Kolom
        total_ns_d = Decimal('0')
        total_ns_k = Decimal('0')
        total_adj_d = Decimal('0')
        total_adj_k = Decimal('0')
        total_nsd_d = Decimal('0')
        total_nsd_k = Decimal('0')
        total_lr_d = Decimal('0')
        total_lr_k = Decimal('0')
        total_nr_d = Decimal('0')
        total_nr_k = Decimal('0')

        # Process each account from trial balance
        for account_code, data in neraca_saldo_dict.items():
            debit_neraca = data['debit']
            credit_neraca = data['credit']
            
            adjustment_data = adjustment_dict.get(account_code, {})
            adjustment_debit = adjustment_data.get('adjustment_debit', Decimal('0'))
            adjustment_credit = adjustment_data.get('adjustment_credit', Decimal('0'))
            
            # Hitung Neraca Saldo Disesuaikan
            account_type = data['account_type']
            adjusted_debit = Decimal('0')
            adjusted_credit = Decimal('0')
            
            if account_type in ('Aset Lancar', 'Aset Tetap', 'Beban'):
                # Normal Debit
                saldo_akhir = (debit_neraca - credit_neraca) + (adjustment_debit - adjustment_credit)
                if saldo_akhir >= 0:
                    adjusted_debit = saldo_akhir
                else:
                    adjusted_credit = abs(saldo_akhir)
            else:
                # Normal Kredit
                saldo_akhir = (credit_neraca - debit_neraca) + (adjustment_credit - adjustment_debit)
                if saldo_akhir >= 0:
                    adjusted_credit = saldo_akhir
                else:
                    adjusted_debit = abs(saldo_akhir)
            
            is_laba_rugi = account_type in ['Pendapatan', 'Beban']
            is_neraca = account_type in ['Aset Lancar', 'Aset Tetap', 'Liabilitas', 'Ekuitas']
            
            laba_rugi_debit = Decimal('0')
            laba_rugi_credit = Decimal('0')
            neraca_debit = Decimal('0')
            neraca_credit = Decimal('0')
            
            if is_laba_rugi:
                laba_rugi_debit = adjusted_debit
                laba_rugi_credit = adjusted_credit
            elif is_neraca:
                neraca_debit = adjusted_debit
                neraca_credit = adjusted_credit
            
            work_sheet_data.append({
                'account_code': account_code,
                'account_name': data['account_name'],
                'account_type': account_type,
                'debit_neraca': debit_neraca,
                'credit_neraca': credit_neraca,
                'adjustment_debit': adjustment_debit,
                'adjustment_credit': adjustment_credit,
                'adjusted_debit': adjusted_debit,
                'adjusted_credit': adjusted_credit,
                'laba_rugi_debit': laba_rugi_debit,
                'laba_rugi_credit': laba_rugi_credit,
                'neraca_debit': neraca_debit,
                'neraca_credit': neraca_credit
            })
            
            total_ns_d += debit_neraca
            total_ns_k += credit_neraca
            total_adj_d += adjustment_debit
            total_adj_k += adjustment_credit
            total_nsd_d += adjusted_debit
            total_nsd_k += adjusted_credit
            total_lr_d += laba_rugi_debit
            total_lr_k += laba_rugi_credit
            total_nr_d += neraca_debit
            total_nr_k += neraca_credit

        # Tambahkan juga akun yang HANYA ada di penyesuaian
        for account_code, adj_data in adjustment_dict.items():
            if account_code not in neraca_saldo_dict:
                adjustment_debit = adj_data.get('adjustment_debit', Decimal('0'))
                adjustment_credit = adj_data.get('adjustment_credit', Decimal('0'))
                account_type = adj_data.get('account_type', '')
                
                adjusted_debit = Decimal('0')
                adjusted_credit = Decimal('0')
                
                if account_type in ('Aset Lancar', 'Aset Tetap', 'Beban'):
                    saldo_akhir = adjustment_debit - adjustment_credit
                    if saldo_akhir >= 0: adjusted_debit = saldo_akhir
                    else: adjusted_credit = abs(saldo_akhir)
                else:
                    saldo_akhir = adjustment_credit - adjustment_debit
                    if saldo_akhir >= 0: adjusted_credit = saldo_akhir
                    else: adjusted_debit = abs(saldo_akhir)

                is_laba_rugi = account_type in ['Pendapatan', 'Beban']
                is_neraca = account_type in ['Aset Lancar', 'Aset Tetap', 'Liabilitas', 'Ekuitas']
                
                laba_rugi_debit = Decimal('0')
                laba_rugi_credit = Decimal('0')
                neraca_debit = Decimal('0')
                neraca_credit = Decimal('0')
                
                if is_laba_rugi:
                    laba_rugi_debit = adjusted_debit
                    laba_rugi_credit = adjusted_credit
                elif is_neraca:
                    neraca_debit = adjusted_debit
                    neraca_credit = adjusted_credit
                
                work_sheet_data.append({
                    'account_code': account_code,
                    'account_name': adj_data.get('account_name', ''),
                    'account_type': account_type,
                    'debit_neraca': Decimal('0'),
                    'credit_neraca': Decimal('0'),
                    'adjustment_debit': adjustment_debit,
                    'adjustment_credit': adjustment_credit,
                    'adjusted_debit': adjusted_debit,
                    'adjusted_credit': adjusted_credit,
                    'laba_rugi_debit': laba_rugi_debit,
                    'laba_rugi_credit': laba_rugi_credit,
                    'neraca_debit': neraca_debit,
                    'neraca_credit': neraca_credit
                })
                
                total_adj_d += adjustment_debit
                total_adj_k += adjustment_credit
                total_nsd_d += adjusted_debit
                total_nsd_k += adjusted_credit
                total_lr_d += laba_rugi_debit
                total_lr_k += laba_rugi_credit
                total_nr_d += neraca_debit
                total_nr_k += neraca_credit

        # --- LOGIKA PERHITUNGAN LABA/RUGI & SELISIH ---
        laba_rugi_bersih = total_lr_k - total_lr_d
        
        # Inisialisasi nilai selisih untuk tabel
        selisih_lr_d = Decimal('0')
        selisih_lr_k = Decimal('0')
        selisih_nr_d = Decimal('0')
        selisih_nr_k = Decimal('0')
        
        label_selisih = "Laba Bersih" if laba_rugi_bersih >= 0 else "Rugi Bersih"
        
        # Jika Laba (Kredit > Debit), selisih ditaruh di Debit Laba Rugi dan Kredit Neraca
        if laba_rugi_bersih >= 0:
            selisih_lr_d = laba_rugi_bersih  # Di LR, tambah ke sisi Debit (yang lebih kecil)
            selisih_nr_k = laba_rugi_bersih  # Di Neraca, tambah ke sisi Kredit (Ekuitas bertambah)
        else:
            # Jika Rugi (Debit > Kredit), selisih ditaruh di Kredit Laba Rugi dan Debit Neraca
            selisih_lr_k = abs(laba_rugi_bersih)
            selisih_nr_d = abs(laba_rugi_bersih)

        # Hitung Total Akhir (Balance)
        balance_lr_d = total_lr_d + selisih_lr_d
        balance_lr_k = total_lr_k + selisih_lr_k
        balance_nr_d = total_nr_d + selisih_nr_d
        balance_nr_k = total_nr_k + selisih_nr_k

    except Exception as e:
        error_message = f"‚ùå Terjadi kesalahan: {str(e)}"
        import traceback
        print("Error details:", traceback.format_exc())
        table_rows = f"""<tr><td colspan="12" style="padding: 20px; text-align: center; color: #dc3545;">{error_message}</td></tr>"""
        
        # Reset all totals to 0
        total_ns_d = total_ns_k = total_adj_d = total_adj_k = total_nsd_d = total_nsd_k = Decimal('0')
        total_lr_d = total_lr_k = total_nr_d = total_nr_k = Decimal('0')
        selisih_lr_d = selisih_lr_k = selisih_nr_d = selisih_nr_k = Decimal('0')
        balance_lr_d = balance_lr_k = balance_nr_d = balance_nr_k = Decimal('0')
        label_selisih = ""

    # Format data untuk tabel
    table_rows = ""
    for data in work_sheet_data:
        row_style = "background-color: #fff3cd;" if data['adjustment_debit'] > 0 or data['adjustment_credit'] > 0 else ""
        
        table_rows += f"""
        <tr style="{row_style}">
            <td style="padding: 8px; border: 1px solid #dee2e6; font-weight: bold;">{data['account_code']}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6;">{data['account_name']}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6; text-align: right;">{format_currency(data['debit_neraca'])}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6; text-align: right;">{format_currency(data['credit_neraca'])}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6; text-align: right;">{format_currency(data['adjustment_debit'])}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6; text-align: right;">{format_currency(data['adjustment_credit'])}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6; text-align: right;">{format_currency(data['adjusted_debit'])}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6; text-align: right;">{format_currency(data['adjusted_credit'])}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6; text-align: right;">{format_currency(data['laba_rugi_debit'])}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6; text-align: right;">{format_currency(data['laba_rugi_credit'])}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6; text-align: right;">{format_currency(data['neraca_debit'])}</td>
            <td style="padding: 8px; border: 1px solid #dee2e6; text-align: right;">{format_currency(data['neraca_credit'])}</td>
        </tr>
        """

    # --- BAGIAN FOOTER TABEL (MODIFIKASI UTAMA) ---
    
    # 1. Baris Jumlah Awal
    table_rows += f"""
    <tr style="background-color: #f8f9fa; font-weight: bold;">
        <td colspan="2" style="padding: 10px; border: 1px solid #dee2e6; text-align: left;">Jumlah</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{format_currency(total_ns_d)}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{format_currency(total_ns_k)}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{format_currency(total_adj_d)}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{format_currency(total_adj_k)}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{format_currency(total_nsd_d)}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{format_currency(total_nsd_k)}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{format_currency(total_lr_d)}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{format_currency(total_lr_k)}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{format_currency(total_nr_d)}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{format_currency(total_nr_k)}</td>
    </tr>
    """

    # 2. Baris Selisih (Laba/Rugi Bersih)
    # Kosongkan kolom yang tidak perlu, isi hanya di LR dan Neraca
    selisih_lr_d_str = format_currency(selisih_lr_d) if selisih_lr_d > 0 else ""
    selisih_lr_k_str = format_currency(selisih_lr_k) if selisih_lr_k > 0 else ""
    selisih_nr_d_str = format_currency(selisih_nr_d) if selisih_nr_d > 0 else ""
    selisih_nr_k_str = format_currency(selisih_nr_k) if selisih_nr_k > 0 else ""

    table_rows += f"""
    <tr style="background-color: #fff; font-weight: bold;">
        <td colspan="2" style="padding: 10px; border: 1px solid #dee2e6; text-align: left;">{label_selisih}</td>
        <td style="border: 1px solid #dee2e6;"></td> <td style="border: 1px solid #dee2e6;"></td>
        <td style="border: 1px solid #dee2e6;"></td> <td style="border: 1px solid #dee2e6;"></td>
        <td style="border: 1px solid #dee2e6;"></td> <td style="border: 1px solid #dee2e6;"></td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{selisih_lr_d_str}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{selisih_lr_k_str}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{selisih_nr_d_str}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right;">{selisih_nr_k_str}</td>
    </tr>
    """
    
    # 3. Baris Balance Akhir
    # Hanya dihitung untuk LR dan Neraca, sisanya kosong
    table_rows += f"""
    <tr style="background-color: #f8f9fa; font-weight: bold; border-top: 2px solid #333;">
        <td colspan="2" style="padding: 10px; border: 1px solid #dee2e6; text-align: left;"></td>
        <td style="border: 1px solid #dee2e6;"></td> <td style="border: 1px solid #dee2e6;"></td>
        <td style="border: 1px solid #dee2e6;"></td> <td style="border: 1px solid #dee2e6;"></td>
        <td style="border: 1px solid #dee2e6;"></td> <td style="border: 1px solid #dee2e6;"></td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right; border-bottom: 3px double #333;">{format_currency(balance_lr_d)}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right; border-bottom: 3px double #333;">{format_currency(balance_lr_k)}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right; border-bottom: 3px double #333;">{format_currency(balance_nr_d)}</td>
        <td style="padding: 10px; border: 1px solid #dee2e6; text-align: right; border-bottom: 3px double #333;">{format_currency(balance_nr_k)}</td>
    </tr>
    """

    body = f"""
    <style>
        .work-sheet-container {{
            max-width: 100%;
            margin: 0 auto;
            padding: 20px;
            overflow-x: auto;
        }}
        .work-sheet-table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
            font-size: 0.85em;
            min-width: 1200px;
        }}
        .section-title {{
            color: #2c3e50;
            border-bottom: 2px solid #007bff;
            padding-bottom: 10px;
            margin-bottom: 20px;
        }}
        .info-box {{
            background-color: #e7f3ff;
            padding: 15px;
            border-radius: 8px;
            margin-bottom: 20px;
            border-left: 4px solid #007bff;
        }}
        .summary-box {{
            background-color: #f8f9fa;
            padding: 15px;
            border-radius: 8px;
            margin-bottom: 20px;
            border: 1px solid #dee2e6;
        }}
    </style>
    
    <div class="work-sheet-container">
        <h2 class="section-title">üìä Neraca Lajur (Work Sheet)</h2>
        
        <div class="info-box">
            <h4 style="color: #0056b3; margin-bottom: 10px;">üí° Informasi Neraca Lajur</h4>
            <p style="margin-bottom: 0; color: #495057;">
                Neraca Lajur adalah kertas kerja akuntansi yang digunakan untuk menyusun laporan keuangan.
            </p>
        </div>

        <div style="overflow-x: auto;">
            <table class="work-sheet-table">
                <thead>
                    <tr>
                        <th rowspan="2" style="padding: 12px; border: 1px solid #dee2e6; background-color: #343a40; color: white; text-align: center;">Kode Akun</th>
                        <th rowspan="2" style="padding: 12px; border: 1px solid #dee2e6; background-color: #343a40; color: white; text-align: center;">Nama Akun</th>
                        <th colspan="2" style="padding: 12px; border: 1px solid #dee2e6; background-color: #28a745; color: white; text-align: center;">Neraca Saldo</th>
                        <th colspan="2" style="padding: 12px; border: 1px solid #dee2e6; background-color: #ffc107; color: black; text-align: center;">Penyesuaian</th>
                        <th colspan="2" style="padding: 12px; border: 1px solid #dee2e6; background-color: #17a2b8; color: white; text-align: center;">Disesuaikan</th>
                        <th colspan="2" style="padding: 12px; border: 1px solid #dee2e6; background-color: #6f42c1; color: white; text-align: center;">Laba/Rugi</th>
                        <th colspan="2" style="padding: 12px; border: 1px solid #dee2e6; background-color: #e83e8c; color: white; text-align: center;">Neraca</th>
                    </tr>
                    <tr>
                        <th style="padding: 10px; border: 1px solid #dee2e6; background-color: #28a745; color: white; text-align: center;">Debit</th>
                        <th style="padding: 10px; border: 1px solid #dee2e6; background-color: #28a745; color: white; text-align: center;">Kredit</th>
                        <th style="padding: 10px; border: 1px solid #dee2e6; background-color: #ffc107; color: black; text-align: center;">Debit</th>
                        <th style="padding: 10px; border: 1px solid #dee2e6; background-color: #ffc107; color: black; text-align: center;">Kredit</th>
                        <th style="padding: 10px; border: 1px solid #dee2e6; background-color: #17a2b8; color: white; text-align: center;">Debit</th>
                        <th style="padding: 10px; border: 1px solid #dee2e6; background-color: #17a2b8; color: white; text-align: center;">Kredit</th>
                        <th style="padding: 10px; border: 1px solid #dee2e6; background-color: #6f42c1; color: white; text-align: center;">Debit</th>
                        <th style="padding: 10px; border: 1px solid #dee2e6; background-color: #6f42c1; color: white; text-align: center;">Kredit</th>
                        <th style="padding: 10px; border: 1px solid #dee2e6; background-color: #e83e8c; color: white; text-align: center;">Debit</th>
                        <th style="padding: 10px; border: 1px solid #dee2e6; background-color: #e83e8c; color: white; text-align: center;">Kredit</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
        </div>
    """
    return render_page("Neraca Lajur", body, sidebar_content=get_admin_sidebar_html())

# =========================
# INVENTORY CARD
# =========================
@app.route("/admin/inventory-journal")
def inventory_journal():
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    try:
        items = db.execute("SELECT id, item_name FROM inventory_items ORDER BY item_name ASC").fetchall()
    except Exception as e:
        return render_page("Error", f"<p>Gagal mengambil data: {e}</p>", sidebar_content=get_admin_sidebar_html())
    
    body = f"""
    <style>
        .inventory-list {{
            list-style: none;
            padding: 0;
        }}
        .inventory-list li {{
            margin-bottom: 10px;
        }}
        .inventory-link {{
            display: block;
            padding: 15px;
            background-color: white;
            border: 1px solid #ddd;
            border-radius: 8px;
            text-decoration: none;
            color: #333;
            font-weight: bold;
            transition: background 0.2s;
        }}
        .inventory-link:hover {{
            background-color: #f0f8ff;
            border-color: #007bff;
            color: #007bff;
        }}
    </style>

    <a href="{url_for('inventory_card_all')}" style="text-decoration: none;">
        <input type="button" value="Tampilkan Semua Kartu Stok" class="btn-blue">
    </a>
    <hr>
    <h2>Atau, Pilih Kartu Stok Satu per Satu:</h2>
    """
    
    if items:
        body += "<ul class='inventory-list'>"
        for item in items:
            body += f'<li><a href="/admin/inventory-card?item_id={item["id"]}" class="inventory-link">üì¶ {item["item_name"]}</a></li>'
        body += "</ul>"
    else:
        # Tampilkan pesan jika tidak ada data
        body += """
        <div style="padding: 20px; background-color: #fff3cd; border: 1px solid #ffeeba; color: #856404; border-radius: 5px;">
            ‚ö†Ô∏è Belum ada data item inventory. Silakan tambahkan data awal di database atau jalankan ulang inisialisasi.
        </div>
        """
    
    return render_page("Jurnal Inventory", body, sidebar_content=get_admin_sidebar_html())

def generate_inventory_card_html(item_id, db):
    """
    Helper untuk membuat HTML untuk SATU kartu stok.
    Perbaikan: Menambahkan 'white-space: nowrap' agar teks tidak turun ke bawah.
    """
    item = db.execute("SELECT * FROM inventory_items WHERE id = ?", (item_id,)).fetchone()
    if not item:
        return f"<p>Error: Item {item_id} tidak ditemukan.</p>"
        
    # Panggil fungsi logic perhitungan (pastikan process_average_card sudah ada)
    table_rows, final_qty, final_cost = process_average_card(item_id, db)
    
    # Style untuk sub-header
    header_sub_style = 'border: 1px solid #999; padding: 8px; font-size: 0.9em;'
    
    # Buat HTML
    card_html = f"""
    <style>
        .stok-table {{ width: 100%; border-collapse: collapse; margin-top: 10px; border: 1px solid #999; }}
        
        /* HEADER */
        .stok-table th {{ 
            border: 1px solid #999; 
            padding: 10px; 
            vertical-align: middle; 
            text-align: center; 
            color: black; 
            white-space: nowrap; /* Header juga satu baris */
        }}
        
        .bg-gray {{ background-color: #e0e0e0; }}
        .bg-green {{ background-color: #d4edda; }}
        
        /* ISI DATA (TD) */
        .stok-table td {{ 
            border: 1px solid #ddd; 
            padding: 8px 10px; /* Sedikit diperluas paddingnya */
            vertical-align: middle; /* Teks rata tengah secara vertikal */
            white-space: nowrap;    /* PENTING: Ini yang mencegah teks turun ke bawah */
        }}
        
        .num-col {{ 
            text-align: right; 
            font-family: 'Courier New', Courier, monospace; 
        }}
        
        .center-col {{ text-align: center; }}
    </style>

    <h2>Kartu Stok: {item['item_name']} (Metode AVERAGE)</h2>
    <p>Unit Saat Ini: <b>{final_qty:,.0f} unit</b>, Total Nilai: <b>{format_currency(final_cost)}</b></p>
    
    <div style="overflow-x: auto;">
    <table class="stok-table">
        <thead>
            <tr class="bg-gray">
                <th rowspan="2">Tanggal</th>
                <th rowspan="2">No. Faktur</th>
                <th rowspan="2">Keterangan</th>
                <th colspan="3">Masuk (Pembelian)</th>
                <th colspan="3">Keluar (HPP/Beban)</th>
                <th colspan="3" class="bg-green">Saldo</th>
            </tr>
            <tr style="background-color: #f9f9f9;">
                <th style="{{header_sub_style}}">Unit</th>
                <th style="{{header_sub_style}}">Harga</th>
                <th style="{{header_sub_style}}">Total</th>
                
                <th style="{{header_sub_style}}">Unit</th>
                <th style="{{header_sub_style}}">Harga</th>
                <th style="{{header_sub_style}}">Total</th>
                
                <th style="{{header_sub_style}} background-color: #e9f7ef;">Unit</th>
                <th style="{{header_sub_style}} background-color: #e9f7ef;">Rata-rata</th>
                <th style="{{header_sub_style}} background-color: #e9f7ef;">Total</th>
            </tr>
        </thead>
        <tbody>
            {table_rows}
        </tbody>
    </table>
    </div>
    """
    return card_html

@app.route("/admin/inventory-card")
def inventory_card():
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    item_id = request.args.get('item_id')
    if not item_id:
        return redirect(url_for('inventory_journal'))
        
    db = get_db()
    item = db.execute("SELECT * FROM inventory_items WHERE id = ?", (item_id,)).fetchone()
    if not item:
        return redirect(url_for('inventory_journal'))
        
    # --- [PERUBAHAN] Panggil helper untuk tabel riwayat ---
    history_table_html = generate_inventory_card_html(item_id, db)
    # --- [AKHIR PERUBAHAN] ---
    
    # --- Logika form ---
    form_bawah_html = "" 
    if item['sales_account']:
        form_bawah_html = f"""
        <div style="padding: 10px; border: 1px dashed #ccc; background-color: #f9f9f9; border-radius: 5px;">
            <h4>Keterangan Penjualan</h4>
            <p>Penjualan untuk <b>{item['item_name']}</b> dicatat secara otomatis saat konsumen melakukan pembelian di Laman Utama.</p>
            <p>Gunakan formulir di atas hanya untuk mencatat <b>Pembelian Stok (Stok Masuk)</b>.</p>
        </div>
        """
    else:
        form_bawah_html = f"""
        <form action="/admin/inventory-tx" method="POST" style="width: 100%;">
            <input type="hidden" name="item_id" value="{item_id}">
            <input type="hidden" name="trx_type" value="sale">
            <h4>Form Pemakaian Stok (Use)</h4>
            
            <label for="s_date">Tanggal Pakai:</label>
            <input type="date" id="s_date" name="trx_date" required>
            <label for="s_desc">Deskripsi (Misal: Dipakai untuk...):</label>
            <input type="text" id="s_desc" name="description" value="Pemakaian Stok" required>
            <label for="s_qty">Kuantitas Dipakai:</label>
            <input type="number" step="any" id="s_qty" name="qty" required>
            
            <input type="hidden" name="sale_price_per_unit" value="0"> 
            
            <input type="submit" value="Simpan Pemakaian" class="btn-red">
        </form>
        """

    # --- [PERUBAHAN] Gabungkan form dan tabel riwayat ---
    body = f"""
    <form action="/admin/inventory-tx" method="POST" style="width: 100%;">
        <input type="hidden" name="item_id" value="{item_id}">
        <input type="hidden" name="trx_type" value="purchase">
        <h4>Form Beli Stok (Purchase)</h4>
        
        <label for="p_date">Tanggal:</label>
        <input type="date" id="p_date" name="trx_date" required>
        <label for="p_invoice">No. Faktur:</label>
        <input type="text" id="p_invoice" name="invoice_no" placeholder="Contoh: F-2025-001" required>
        <label for="p_desc">Deskripsi (Opsional):</label>
        <input type="text" id="p_desc" name="description" placeholder="(misal: Pembelian Stok dari Supplier A)">
        <label for="p_qty">Kuantitas:</label>
        <input type="number" step="any" id="p_qty" name="qty" required>
        <label for="p_cost">Harga Beli Per Unit (Cost):</label>
        <input type="number" step="any" id="p_cost" name="cost_per_unit" required>
        
        <input type="submit" value="Simpan Pembelian" class="btn-blue">
    </form>
    <hr style="margin: 25px 0;">

    {form_bawah_html}
    <hr style="margin: 25px 0;">

    {history_table_html}
    """
    
    return render_page(f"Kartu Stok - {item['item_name']}", body, sidebar_content=get_admin_sidebar_html())

def process_average_card(item_id, db):
    """
    Process inventory card using average method.
    Perbaikan: Warna data faktur jadi hitam.
    """
    logs = db.execute("""
        SELECT * FROM inventory_log 
        WHERE item_id = ? 
        ORDER BY trx_date, id
    """, (item_id,)).fetchall()
    
    table_rows = ""
    running_qty = 0
    running_cost = 0.0
    average_cost = 0.0
    
    for log in logs:
        date_str = log['trx_date']
        
        # Logika pisah faktur & deskripsi
        full_desc = log['description'] or ""
        faktur = "-"
        keterangan = full_desc
        
        if " - " in full_desc:
            parts = full_desc.split(" - ", 1)
            faktur = parts[0]
            keterangan = parts[1]
        elif full_desc.startswith("F"):
            faktur = full_desc
            keterangan = "-"
            
        # Format Tanggal (DD-MM-YYYY)
        try:
            d_obj = datetime.strptime(date_str, '%Y-%m-%d')
            date_display = d_obj.strftime('%d-%m-%Y')
        except:
            date_display = date_str

        if log['trx_type'] == 'purchase':
            # Pembelian
            qty_in = log['qty']
            cost_in = log['cost_per_unit']
            total_in = qty_in * cost_in
            
            if running_qty + qty_in > 0:
                average_cost = ((running_qty * average_cost) + total_in) / (running_qty + qty_in)
            
            running_qty += qty_in
            running_cost = running_qty * average_cost
            
            table_rows += f"""
            <tr>
                <td class="center-col">{date_display}</td>
                <td class="center-col" style="font-weight:bold; color:black;">{faktur}</td>
                <td>{keterangan}</td>
                
                <td class="num-col">{qty_in:,.0f}</td>
                <td class="num-col">{format_currency(cost_in)}</td>
                <td class="num-col">{format_currency(total_in)}</td>
                
                <td class="center-col" style="color:#ccc;">-</td>
                <td class="center-col" style="color:#ccc;">-</td>
                <td class="center-col" style="color:#ccc;">-</td>
                
                <td class="num-col" style="background-color: #f1f8f4; font-weight:bold;">{running_qty:,.0f}</td>
                <td class="num-col" style="background-color: #f1f8f4;">{format_currency(average_cost)}</td>
                <td class="num-col" style="background-color: #f1f8f4;">{format_currency(running_cost)}</td>
            </tr>
            """
            
        elif log['trx_type'] == 'sale':
            # Penjualan
            qty_out = log['qty']
            total_out = qty_out * average_cost
            
            running_qty -= qty_out
            running_cost = running_qty * average_cost
            
            table_rows += f"""
            <tr>
                <td class="center-col">{date_display}</td>
                <td class="center-col" style="font-weight:bold; color:black;">{faktur}</td>
                <td>{keterangan}</td>
                
                <td class="center-col" style="color:#ccc;">-</td>
                <td class="center-col" style="color:#ccc;">-</td>
                <td class="center-col" style="color:#ccc;">-</td>
                
                <td class="num-col" style="color: #dc3545;">{qty_out:,.0f}</td>
                <td class="num-col">{format_currency(average_cost)}</td>
                <td class="num-col">{format_currency(total_out)}</td>
                
                <td class="num-col" style="background-color: #f1f8f4; font-weight:bold;">{running_qty:,.0f}</td>
                <td class="num-col" style="background-color: #f1f8f4;">{format_currency(average_cost)}</td>
                <td class="num-col" style="background-color: #f1f8f4;">{format_currency(running_cost)}</td>
            </tr>
            """
    
    return table_rows, running_qty, running_cost

@app.route("/admin/inventory-card-all")
def inventory_card_all():
    """Rute BARU untuk menampilkan SEMUA kartu stok."""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))

    db = get_db()
    items = db.execute("SELECT id, item_name FROM inventory_items ORDER BY item_name").fetchall()
    
    all_cards_html = f"""
    <p><a href="{url_for('inventory_journal')}">‚Üê Kembali ke Pilihan Inventory</a></p>
    <hr>
    """
    
    # Loop dan panggil helper
    for item in items:
        all_cards_html += generate_inventory_card_html(item['id'], db)
        all_cards_html += "<hr style='margin-top: 30px;'>"
        
    return render_page("Semua Kartu Stok", all_cards_html, sidebar_content=get_admin_sidebar_html())

@app.route("/admin/inventory-tx", methods=['POST'])
def inventory_tx():
    """Menangani input Beli (Purchase) dan Pakai (Use) dari Admin."""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))

    # 1. Ambil data Form
    item_id = request.form['item_id']
    trx_type = request.form['trx_type']
    trx_date = request.form['trx_date']
    # 'description' TIDAK diambil di sini lagi (spesifik per tipe)
    qty = float(request.form['qty'])
    
    db = get_db()
    item = db.execute("SELECT * FROM inventory_items WHERE id = ?", (item_id,)).fetchone()
    
    # --- [PERBAIKAN WAKTU REAL-TIME] ---
    # Ambil jam:menit:detik saat ini
    current_time = datetime.now().strftime('%H:%M:%S')
    # Gabungkan tanggal yang dipilih dengan jam saat ini
    trx_timestamp_str = f"{trx_date} {current_time}"
    
    # Untuk generate kode jurnal, kita butuh objek datetime
    trx_datetime = datetime.strptime(trx_date, '%Y-%m-%d')

    with db:
        if trx_type == 'purchase':
            # Ambil data invoice & deskripsi
            invoice_no = request.form['invoice_no']
            desc_note = request.form['description'] 
            
            # Gabungkan
            description = invoice_no
            if desc_note:
                description = f"{invoice_no} - {desc_note}"
            
            cost_per_unit = float(request.form['cost_per_unit'])
            total_cost = qty * cost_per_unit
            
            # Simpan ke Log Inventory
            db.execute(
                "INSERT INTO inventory_log (item_id, trx_date, trx_type, description, qty, cost_per_unit) VALUES (?, ?, ?, ?, ?, ?)",
                (item_id, trx_date, trx_type, description, qty, cost_per_unit)
            )
            
            # Jurnal
            jurnal_desc = description 
            j_code = generate_journal_code(db, trx_datetime)
            
            # Gunakan trx_timestamp_str yang sudah ada jam-nya
            cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                (j_code, trx_timestamp_str, jurnal_desc))
            entry_id = cursor.lastrowid
            
            # Debit: Persediaan
            db.execute("INSERT INTO journal_details (entry_id, account_code, debit) VALUES (?, ?, ?)",
                       (entry_id, item['inventory_account'], total_cost))
            # Kredit: Kas (Asumsi)
            db.execute("INSERT INTO journal_details (entry_id, account_code, credit) VALUES (?, ?, ?)",
                       (entry_id, '1101', total_cost))

        elif trx_type == 'sale': 
            # Ambil deskripsi manual
            description = request.form['description']

            # (Logika Hitung HPP Average tetap sama)
            logs = db.execute("SELECT * FROM inventory_log WHERE item_id = ? AND trx_date <= ? ORDER BY trx_date, id", (item_id, trx_date)).fetchall()
            average_stack = []
            for log in logs:
                if log['trx_type'] == 'purchase':
                    average_stack.append([log['qty'], log['cost_per_unit']])
                elif log['trx_type'] == 'sale':
                    qty_to_sell_hist = log['qty']
                    while qty_to_sell_hist > 0 and average_stack:
                        if average_stack[0][0] <= qty_to_sell_hist:
                            qty_to_sell_hist -= average_stack.pop(0)[0]
                        else:
                            average_stack[0][0] -= qty_to_sell_hist
                            qty_to_sell_hist = 0
            
            qty_to_use = qty
            total_cogs = 0.0
            temp_qty_to_use = qty_to_use
            while temp_qty_to_use > 0 and average_stack:
                if average_stack[0][0] <= temp_qty_to_use:
                    qty_from_stack = average_stack[0][0]
                    cost_from_stack = average_stack[0][1]
                    total_cogs += (qty_from_stack * cost_from_stack)
                    temp_qty_to_use -= qty_from_stack
                    average_stack.pop(0)
                else:
                    total_cogs += (temp_qty_to_use * average_stack[0][1])
                    average_stack[0][0] -= temp_qty_to_use
                    temp_qty_to_use = 0
            
            if item['sales_account'] is None:
                # Catat Log Inventory
                db.execute(
                    "INSERT INTO inventory_log (item_id, trx_date, trx_type, description, qty, sale_price_per_unit) VALUES (?, ?, ?, ?, ?, ?)",
                    (item_id, trx_date, trx_type, description, qty, 0)
                )
                
                jurnal_desc = description 
                j_code = generate_journal_code(db, trx_datetime)
                
                cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                    (j_code, trx_timestamp_str, jurnal_desc))
                entry_id = cursor.lastrowid
                
                db.execute("INSERT INTO journal_details (entry_id, account_code, debit) VALUES (?, ?, ?)",
                           (entry_id, item['expense_cogs_account'], total_cogs))
                db.execute("INSERT INTO journal_details (entry_id, account_code, credit) VALUES (?, ?, ?)",
                           (entry_id, item['inventory_account'], total_cogs))

    return redirect(url_for('inventory_card', item_id=item_id))

# ==================================
# LAPORAN RUGI LABA
# ==================================
@app.route("/admin/income-statement")
def income_statement():
    """Laporan Laba Rugi"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    error_message = None
    
    # Ambil data perusahaan dan periode dari database
    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    
    # Set default jika data perusahaan belum ada
    nama_perusahaan = company_info['company_name'] if company_info else 'Budidaya Gurame'
    periode_akuntansi = company_info['accounting_period'] if company_info else '2025'
    
    try:
        # Query untuk mengambil semua akun pendapatan dan beban
        query = """
        SELECT
            c.account_code,
            c.account_name,
            c.account_type,
            SUM(d.debit) as total_debit,
            SUM(d.credit) as total_credit
        FROM journal_details d
        JOIN chart_of_accounts c ON d.account_code = c.account_code
        WHERE c.account_type IN ('Pendapatan', 'Beban')
        GROUP BY c.account_code
        ORDER BY c.account_type DESC, c.account_code;
        """
        accounts = db.execute(query).fetchall()
        
        pendapatan_html = ""
        beban_html = ""
        total_pendapatan = Decimal('0')
        total_beban = Decimal('0')

        for acc in accounts:
            if acc['account_type'] == 'Pendapatan':
                # Saldo normal Pendapatan adalah Kredit
                balance = Decimal(str(acc['total_credit'] or '0')) - Decimal(str(acc['total_debit'] or '0'))
                total_pendapatan += balance
                pendapatan_html += f"""
                <tr>
                    <td class="item-name indent">{acc['account_name']}</td>
                    <td class="amount">{format_currency(balance)}</td>
                </tr>
                """
            elif acc['account_type'] == 'Beban':
                # Saldo normal Beban adalah Debit
                balance = Decimal(str(acc['total_debit'] or '0')) - Decimal(str(acc['total_credit'] or '0'))
                total_beban += balance
                beban_html += f"""
                <tr>
                    <td class="item-name indent">{acc['account_name']}</td>
                    <td class="amount">{format_currency(balance)}</td>
                </tr>
                """
                
        # Hitung laba/rugi bersih
        laba_rugi_bersih = total_pendapatan - total_beban
        
        # Tentukan label hasil
        hasil_label = "Laba Bersih" if laba_rugi_bersih >= 0 else "Rugi Bersih"
        
    except Exception as e:
        error_message = f"‚ùå Terjadi kesalahan: {str(e)}"
        pendapatan_html = '<tr><td class="item-name indent">Tidak ada data</td><td class="amount">Rp 0,00</td></tr>'
        beban_html = '<tr><td class="item-name indent">Tidak ada data</td><td class="amount">Rp 0,00</td></tr>'
        total_pendapatan = Decimal('0')
        total_beban = Decimal('0')
        laba_rugi_bersih = Decimal('0')
        hasil_label = "Laba Bersih"

    body = f"""
    <style>
        .income-container {{
            max-width: 800px;
            margin: 0 auto;
            padding: 20px;
            font-family: Arial, sans-serif;
        }}
        .header-container {{
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            margin-bottom: 30px;
        }}
        .company-header {{
            text-align: center;
            flex-grow: 1;
        }}
        .company-name {{
            font-size: 24px;
            font-weight: bold;
            margin-bottom: 5px;
        }}
        .report-title {{
            font-size: 18px;
            font-weight: bold;
            margin-bottom: 5px;
        }}
        .period {{
            font-size: 14px;
            margin-bottom: 20px;
        }}
        .income-table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        .income-table td {{
            padding: 8px 12px;
            border: none;
        }}
        .income-table .item-name {{
            width: 70%;
            padding-left: 50px;
        }}
        .income-table .amount {{
            width: 30%;
            text-align: right;
            padding-right: 50px;
            font-weight: normal;
        }}
        .income-table .section-header {{
            font-weight: bold;
            background-color: #f8f9fa;
            border-top: 1px solid #dee2e6;
            border-bottom: 1px solid #dee2e6;
        }}
        .income-table .subtotal {{
            border-top: 1px solid #000;
            font-weight: bold;
        }}
        .income-table .total {{
            border-top: 2px solid #000;
            font-weight: bold;
            background-color: #f5f5f5;
        }}
        .negative-amount {{
            color: #dc3545;
        }}
        .indent {{
            padding-left: 70px !important;
        }}
        .download-btn {{
            background-color: #28a745;
            color: white;
            padding: 10px 20px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            text-decoration: none;
            display: inline-block;
        }}
        .download-btn:hover {{
            background-color: #218838;
            text-decoration: none;
            color: white;
        }}
        .separator {{
            border-top: 2px solid #000;
            margin: 20px 0;
        }}
        .no-data {{
            color: #6c757d;
            font-style: italic;
        }}
    </style>

    <div class="income-container">
        <!-- HEADER DENGAN JUDUL DI TENGAH DAN BUTTON DI KANAN -->
        <div class="header-container">
            <div style="width: 120px;"></div> <!-- Spacer untuk balance -->
            <div class="company-header">
                <div class="company-name">{nama_perusahaan}</div>
                <div class="report-title">Laporan Laba Rugi</div>
                <div class="period">Untuk Periode yang Berakhir 31 Desember {periode_akuntansi}</div>
            </div>
            <div>
                <a href="{url_for('download_income_statement_excel')}" class="download-btn">
                    üì• Download (Excel)
                </a>
            </div>
        </div>

        <div class="separator"></div>

        <!-- TABEL LAPORAN LABA RUGI -->
        <table class="income-table">
            <!-- PENDAPATAN -->
            <tr class="section-header">
                <td class="item-name">Pendapatan</td>
                <td class="amount"></td>
            </tr>
            {pendapatan_html if pendapatan_html else '<tr><td class="item-name indent no-data">Tidak ada data pendapatan</td><td class="amount no-data">Rp 0,00</td></tr>'}
            <tr class="subtotal">
                <td class="item-name">Total Pendapatan</td>
                <td class="amount">{format_currency(total_pendapatan)}</td>
            </tr>
            
            <!-- BEBAN -->
            <tr class="section-header">
                <td class="item-name">Beban</td>
                <td class="amount"></td>
            </tr>
            {beban_html if beban_html else '<tr><td class="item-name indent no-data">Tidak ada data beban</td><td class="amount no-data">Rp 0,00</td></tr>'}
            <tr class="subtotal">
                <td class="item-name">Total Beban</td>
                <td class="amount">{format_currency(total_beban)}</td>
            </tr>
            
            <!-- HASIL -->
            <tr class="total">
                <td class="item-name">{hasil_label}</td>
                <td class="amount">{format_currency(abs(laba_rugi_bersih))}</td>
            </tr>
        </table>
        <div class="separator"></div>
    </div>
    """
    
    return render_page("Laporan Laba Rugi", body, sidebar_content=get_admin_sidebar_html(), error_message=error_message)

@app.route("/admin/download/income-statement-excel")
def download_income_statement_excel():
    """Membuat dan mengirim file .xlsx dari Laporan Laba Rugi."""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
        
    db = get_db()
    
    # Ambil data perusahaan
    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    nama_perusahaan = company_info['company_name'] if company_info else 'Budidaya Gurame'
    periode_akuntansi = company_info['accounting_period'] if company_info else '2025'
    
    try:
        # Query untuk mengambil semua akun pendapatan dan beban
        query = """
        SELECT
            c.account_code,
            c.account_name,
            c.account_type,
            SUM(d.debit) as total_debit,
            SUM(d.credit) as total_credit
        FROM journal_details d
        JOIN chart_of_accounts c ON d.account_code = c.account_code
        WHERE c.account_type IN ('Pendapatan', 'Beban')
        GROUP BY c.account_code
        ORDER BY c.account_type DESC, c.account_code;
        """
        accounts = db.execute(query).fetchall()
        
        # Buat Workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Laba Rugi"
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        
        # Header Perusahaan
        ws.append([nama_perusahaan])
        ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.merge_cells('A1:B1')
        
        ws.append(["Laporan Laba Rugi"])
        ws.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True, size=12)
        ws.cell(row=2, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.merge_cells('A2:B2')
        
        ws.append([f"Untuk Periode yang Berakhir 31 Desember {periode_akuntansi}"])
        ws.cell(row=3, column=1).font = openpyxl.styles.Font(size=11)
        ws.cell(row=3, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.merge_cells('A3:B3')
        
        ws.append([])  # Empty row
        
        total_pendapatan = Decimal('0')
        total_beban = Decimal('0')
        
        # Data Pendapatan
        ws.append(["PENDAPATAN"])
        ws.cell(row=5, column=1).font = openpyxl.styles.Font(bold=True)
        
        pendapatan_accounts = [acc for acc in accounts if acc['account_type'] == 'Pendapatan']
        if pendapatan_accounts:
            for acc in pendapatan_accounts:
                balance = Decimal(str(acc['total_credit'] or '0')) - Decimal(str(acc['total_debit'] or '0'))
                total_pendapatan += balance
                ws.append(["  " + acc['account_name'], float(balance)])
                ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        else:
            ws.append(["  Tidak ada data pendapatan", 0])
            ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        
        # Total Pendapatan
        ws.append(["Total Pendapatan", float(total_pendapatan)])
        ws.cell(row=ws.max_row, column=1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=ws.max_row, column=2).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        ws.cell(row=ws.max_row, column=2).border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='double')
        )
        
        ws.append([])  # Empty row
        
        # Data Beban
        ws.append(["BEBAN"])
        ws.cell(row=ws.max_row, column=1).font = openpyxl.styles.Font(bold=True)
        
        beban_accounts = [acc for acc in accounts if acc['account_type'] == 'Beban']
        if beban_accounts:
            for acc in beban_accounts:
                balance = Decimal(str(acc['total_debit'] or '0')) - Decimal(str(acc['total_credit'] or '0'))
                total_beban += balance
                ws.append(["  " + acc['account_name'], float(balance)])
                ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        else:
            ws.append(["  Tidak ada data beban", 0])
            ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        
        # Total Beban
        ws.append(["Total Beban", float(total_beban)])
        ws.cell(row=ws.max_row, column=1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=ws.max_row, column=2).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        ws.cell(row=ws.max_row, column=2).border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='double')
        )
        
        ws.append([])  # Empty row
        
        # Laba/Rugi Bersih
        laba_rugi_bersih = total_pendapatan - total_beban
        hasil_label = "Laba Bersih" if laba_rugi_bersih >= 0 else "Rugi Bersih"
        
        ws.append([hasil_label, float(abs(laba_rugi_bersih))])
        ws.cell(row=ws.max_row, column=1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=ws.max_row, column=2).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        ws.cell(row=ws.max_row, column=2).border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style='double'),
            bottom=openpyxl.styles.Side(style='double')
        )
        
    except Exception as e:
        # Jika terjadi error, buat file Excel dengan pesan error
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Laba Rugi"
        ws.append(["Error"])
        ws.append([f"Terjadi kesalahan: {str(e)}"])
    
    # Simpan ke buffer memori
    mem_file = BytesIO()
    wb.save(mem_file)
    mem_file.seek(0)
    
    # Kembalikan sebagai file download
    return Response(
        mem_file.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": f"attachment; filename=laporan_laba_rugi_{periode_akuntansi}.xlsx"}
    )

# DOWNLOAD LAPORAN LABA RUGI UNTUK LANDING PAGE
@app.route("/download-income-statement")
def download_income_statement_public():
    """Download laporan laba rugi untuk publik (dari landing page)"""
    db = get_db()
    
    # Ambil data perusahaan
    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    nama_perusahaan = company_info['company_name'] if company_info else 'Budidaya Gurame'
    periode_akuntansi = company_info['accounting_period'] if company_info else '2025'
    
    try:
        # Query untuk mengambil semua akun pendapatan dan beban
        query = """
        SELECT
            c.account_code,
            c.account_name,
            c.account_type,
            SUM(d.debit) as total_debit,
            SUM(d.credit) as total_credit
        FROM journal_details d
        JOIN chart_of_accounts c ON d.account_code = c.account_code
        WHERE c.account_type IN ('Pendapatan', 'Beban')
        GROUP BY c.account_code
        ORDER BY c.account_type DESC, c.account_code;
        """
        accounts = db.execute(query).fetchall()
        
        # Buat Workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Laba Rugi"
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        
        # Header Perusahaan
        ws.append([nama_perusahaan])
        ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.merge_cells('A1:B1')
        
        ws.append(["Laporan Laba Rugi"])
        ws.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True, size=12)
        ws.cell(row=2, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.merge_cells('A2:B2')
        
        ws.append([f"Untuk Periode yang Berakhir 31 Desember {periode_akuntansi}"])
        ws.cell(row=3, column=1).font = openpyxl.styles.Font(size=11)
        ws.cell(row=3, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')
        ws.merge_cells('A3:B3')
        
        ws.append([])  # Empty row
        
        total_pendapatan = Decimal('0')
        total_beban = Decimal('0')
        
        # Data Pendapatan
        ws.append(["PENDAPATAN"])
        ws.cell(row=5, column=1).font = openpyxl.styles.Font(bold=True)
        
        pendapatan_accounts = [acc for acc in accounts if acc['account_type'] == 'Pendapatan']
        if pendapatan_accounts:
            for acc in pendapatan_accounts:
                balance = Decimal(str(acc['total_credit'] or '0')) - Decimal(str(acc['total_debit'] or '0'))
                total_pendapatan += balance
                ws.append(["  " + acc['account_name'], float(balance)])
                ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        else:
            ws.append(["  Tidak ada data pendapatan", 0])
            ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        
        # Total Pendapatan
        ws.append(["Total Pendapatan", float(total_pendapatan)])
        ws.cell(row=ws.max_row, column=1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=ws.max_row, column=2).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        ws.cell(row=ws.max_row, column=2).border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='double')
        )
        
        ws.append([])  # Empty row
        
        # Data Beban
        ws.append(["BEBAN"])
        ws.cell(row=ws.max_row, column=1).font = openpyxl.styles.Font(bold=True)
        
        beban_accounts = [acc for acc in accounts if acc['account_type'] == 'Beban']
        if beban_accounts:
            for acc in beban_accounts:
                balance = Decimal(str(acc['total_debit'] or '0')) - Decimal(str(acc['total_credit'] or '0'))
                total_beban += balance
                ws.append(["  " + acc['account_name'], float(balance)])
                ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        else:
            ws.append(["  Tidak ada data beban", 0])
            ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        
        # Total Beban
        ws.append(["Total Beban", float(total_beban)])
        ws.cell(row=ws.max_row, column=1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=ws.max_row, column=2).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        ws.cell(row=ws.max_row, column=2).border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='double')
        )
        
        ws.append([])  # Empty row
        
        # Laba/Rugi Bersih
        laba_rugi_bersih = total_pendapatan - total_beban
        hasil_label = "Laba Bersih" if laba_rugi_bersih >= 0 else "Rugi Bersih"
        
        ws.append([hasil_label, float(abs(laba_rugi_bersih))])
        ws.cell(row=ws.max_row, column=1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=ws.max_row, column=2).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=ws.max_row, column=2).number_format = '"Rp"#,##0.00'
        ws.cell(row=ws.max_row, column=2).border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style='double'),
            bottom=openpyxl.styles.Side(style='double')
        )
        
    except Exception as e:
        # Jika terjadi error, buat file Excel dengan pesan error
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Laba Rugi"
        ws.append(["Error"])
        ws.append([f"Terjadi kesalahan: {str(e)}"])
    
    # Simpan ke buffer memori
    mem_file = BytesIO()
    wb.save(mem_file)
    mem_file.seek(0)
    
    # Kembalikan sebagai file download
    return Response(
        mem_file.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": f"attachment; filename=laporan_laba_rugi_{periode_akuntansi}.xlsx"}
    )

# ==================================
# PERUBAHAN MODAL
# ==================================
@app.route("/admin/equity-change")
def equity_change():
    """Laporan Perubahan Modal"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    error_message = None
    
    # Ambil data perusahaan dan periode dari database
    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    
    # Set default jika data perusahaan belum ada
    nama_perusahaan = company_info['company_name'] if company_info else 'Budidaya Gurame'
    periode_akuntansi = company_info['accounting_period'] if company_info else '2025'
    
    try:
        modal_awal_data = db.execute(
            "SELECT (SUM(credit) - SUM(debit)) as modal_awal FROM journal_details "
            "WHERE account_code = '3101' AND entry_id = 1"
        ).fetchone()
        modal_awal = Decimal(str(modal_awal_data['modal_awal'] or '0'))
        
        laba_rugi_bersih = Decimal(str(get_net_income(db)))
        
        prive_data = db.execute(
            "SELECT (SUM(debit) - SUM(credit)) as total_prive FROM journal_details "
            "WHERE account_code = '3102'"
        ).fetchone()
        total_prive = Decimal(str(prive_data['total_prive'] or '0'))
        
        modal_akhir = modal_awal + laba_rugi_bersih - total_prive
        
        equity_data = {
            'modal_awal': modal_awal,
            'laba_bersih': laba_rugi_bersih if laba_rugi_bersih > 0 else Decimal('0'),
            'rugi_bersih': abs(laba_rugi_bersih) if laba_rugi_bersih < 0 else Decimal('0'),
            'prive': total_prive,
            'modal_akhir': modal_akhir
        }
        
    except Exception as e:
        error_message = f"‚ùå Terjadi kesalahan: {str(e)}"
        equity_data = { 'modal_awal': 0, 'laba_bersih': 0, 'rugi_bersih': 0, 'prive': 0, 'modal_akhir': 0 }

    body = f"""
    <style>
        .equity-container {{
            max-width: 800px;
            margin: 0 auto;
            padding: 20px;
            font-family: Arial, sans-serif;
        }}
        .header-container {{
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            margin-bottom: 30px;
        }}
        .company-header {{
            text-align: center;
            flex-grow: 1;
        }}
        .company-name {{
            font-size: 24px;
            font-weight: bold;
            margin-bottom: 5px;
        }}
        .report-title {{
            font-size: 18px;
            font-weight: bold;
            margin-bottom: 5px;
        }}
        .period {{
            font-size: 14px;
            margin-bottom: 20px;
        }}
        .equity-table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        .equity-table td {{
            padding: 8px 12px;
            border: none;
        }}
        .equity-table .item-name {{
            width: 70%;
            padding-left: 50px;
        }}
        .equity-table .amount {{
            width: 30%;
            text-align: right;
            padding-right: 50px;
            font-weight: normal;
        }}
        .equity-table .subtotal {{
            border-top: 1px solid #000;
            font-weight: bold;
        }}
        .equity-table .total {{
            border-top: 2px solid #000;
            font-weight: bold;
            background-color: #f5f5f5;
        }}
        .negative-amount {{
            color: #dc3545;
        }}
        .indent {{
            padding-left: 70px !important;
        }}
        .download-btn {{
            background-color: #28a745;
            color: white;
            padding: 10px 20px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            text-decoration: none;
            display: inline-block;
        }}
        .download-btn:hover {{
            background-color: #218838;
        }}
        .company-info {{
            margin-top: 30px;
            padding: 15px;
            background-color: #f8f9fa;
            border-radius: 5px;
            font-size: 14px;
        }}
        .separator {{
            border-top: 2px solid #000;
            margin: 20px 0;
        }}
    </style>

    <div class="equity-container">
        <!-- HEADER DENGAN JUDUL DI TENGAH DAN BUTTON DI KANAN -->
        <div class="header-container">
            <div style="width: 120px;"></div> <!-- Spacer untuk balance -->
            <div class="company-header">
                <div class="company-name">{nama_perusahaan}</div>
                <div class="report-title">Laporan Perubahan Modal</div>
                <div class="period">Per 31 Desember {periode_akuntansi}</div>
            </div>
            <div>
                <a href="{url_for('download_equity_change_excel')}" class="download-btn">
                    üì• Download (Excel)
                </a>
            </div>
        </div>

        <div class="separator"></div>

        <!-- TABEL PERUBAHAN MODAL -->
        <table class="equity-table">
            <tr>
                <td class="item-name">Modal Awal</td>
                <td class="amount">{format_currency(equity_data['modal_awal'])}</td>
            </tr>
            <tr>
                <td class="item-name indent">
                    {'Laba Bersih' if equity_data['laba_bersih'] > 0 else 'Rugi Bersih'}
                </td>
                <td class="amount">
                    {f"({format_currency(equity_data['rugi_bersih'])})" if equity_data['rugi_bersih'] > 0 else format_currency(equity_data['laba_bersih'])}
                </td>
            </tr>
            <tr class="subtotal">
                <td class="item-name"></td>
                <td class="amount">{format_currency(equity_data['modal_awal'] + (equity_data['laba_bersih'] - equity_data['rugi_bersih']))}</td>
            </tr>
            <tr>
                <td class="item-name indent">Prive</td>
                <td class="amount negative-amount">({format_currency(equity_data['prive'])})</td>
            </tr>
            <tr class="total">
                <td class="item-name">Modal Akhir</td>
                <td class="amount">{format_currency(equity_data['modal_akhir'])}</td>
            </tr>
        </table>
        <div class="separator"></div>
    </div>
    """
    
    return render_page("Laporan Perubahan Modal", body, sidebar_content=get_admin_sidebar_html(), error_message=error_message if 'error_message' in locals() else None)

@app.route("/admin/download/equity-change-excel")
def download_equity_change_excel():
    """Membuat dan mengirim file .xlsx dari Laporan Perubahan Modal."""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
        
    db = get_db()
    
    # 1. Ambil data
    modal_awal_data = db.execute(
        "SELECT (SUM(credit) - SUM(debit)) as modal_awal FROM journal_details "
        "WHERE account_code = '3101' AND entry_id = 1"
    ).fetchone()
    modal_awal = Decimal(str(modal_awal_data['modal_awal'] or '0'))
    laba_rugi_bersih = Decimal(str(get_net_income(db)))
    prive_data = db.execute(
        "SELECT (SUM(debit) - SUM(credit)) as total_prive FROM journal_details "
        "WHERE account_code = '3102'"
    ).fetchone()
    total_prive = Decimal(str(prive_data['total_prive'] or '0'))
    
    penambahan_modal = laba_rugi_bersih - total_prive
    modal_akhir = modal_awal + penambahan_modal
    
    # 2. Buat Workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Perubahan Modal"
    
    # 3. Tulis Data
    ws.append(["Keterangan", "Jumlah"])
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    ws["B1"].font = openpyxl.styles.Font(bold=True)

    ws.append(["Modal Awal", float(modal_awal)])
    
    if laba_rugi_bersih >= 0:
        ws.append(["Laba Bersih", float(laba_rugi_bersih)])
    else:
        ws.append(["Rugi Bersih", float(laba_rugi_bersih)]) # Simpan sebagai negatif
        
    ws.append(["Prive", float(-total_prive)]) # Simpan sebagai negatif
    
    ws.append(["Penambahan (Pengurangan) Modal", float(penambahan_modal)])
    ws.cell(row=ws.max_row, column=1).font = openpyxl.styles.Font(bold=True)
    
    ws.append(["Modal Akhir", float(modal_akhir)])
    ws.cell(row=ws.max_row, column=1).font = openpyxl.styles.Font(bold=True)
    
    # Format angka
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '"Rp"#,##0.00;("Rp"#,##0.00)'
            
    # Atur lebar kolom
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    # 4. Simpan ke buffer memori
    mem_file = BytesIO()
    wb.save(mem_file)
    mem_file.seek(0)

    # 5. Kembalikan sebagai file download
    return Response(
        mem_file.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": "attachment; filename=laporan_perubahan_modal.xlsx"}
    )

# ==================================
# LAPORAN POSISI KEUANGAN
# ==================================
@app.route("/admin/financial-position")
def financial_position():
    """Laporan Neraca dengan Layout Sejajar (Aligned Footer)"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    
    db = get_db()
    
    # --- 1. SETUP INFO PERUSAHAAN ---
    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    nama_perusahaan = company_info['company_name'] if company_info else 'Budidaya Gurame'
    periode_akuntansi = company_info['accounting_period'] if company_info else '2025'
    
    # --- 2. PARAMETER TANGGAL ---
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    cutoff_date = f"{end_date} 23:59:59" if end_date else f"{periode_akuntansi}-12-31 23:59:59"
    period_text = f"Per {end_date}" if end_date else f"Per 31 Desember {periode_akuntansi}"
    if start_date and end_date:
        period_text = f"Periode {start_date} s/d {end_date}"

    # --- 3. FUNGSI HELPERS ---
    def get_balance_by_type(account_types):
        query = f"""
        SELECT 
            coa.account_code, coa.account_name, coa.account_type,
            COALESCE(SUM(jd.debit), 0) as total_debit,
            COALESCE(SUM(jd.credit), 0) as total_credit
        FROM chart_of_accounts coa
        LEFT JOIN journal_details jd ON coa.account_code = jd.account_code
        LEFT JOIN journal_entries je ON jd.entry_id = je.id
        WHERE coa.account_type IN ({','.join(['?']*len(account_types))})
        AND (je.entry_timestamp <= ? OR je.entry_timestamp IS NULL)
        GROUP BY coa.account_code, coa.account_name, coa.account_type
        """
        params = account_types + [cutoff_date]
        return db.execute(query, params).fetchall()

    def get_specific_balance(account_code, is_credit_normal=True):
        query = """
        SELECT SUM(jd.debit) as d, SUM(jd.credit) as c
        FROM journal_details jd
        JOIN journal_entries je ON jd.entry_id = je.id
        WHERE jd.account_code = ? AND je.entry_timestamp <= ?
        """
        row = db.execute(query, [account_code, cutoff_date]).fetchone()
        debit = Decimal(str(row['d'] or 0))
        credit = Decimal(str(row['c'] or 0))
        return (credit - debit) if is_credit_normal else (debit - credit)

    def calculate_net_income_upto_date():
        rev_rows = get_balance_by_type(['Pendapatan'])
        total_revenue = sum([Decimal(str(r['total_credit'])) - Decimal(str(r['total_debit'])) for r in rev_rows])
        exp_rows = get_balance_by_type(['Beban'])
        total_expense = sum([Decimal(str(r['total_debit'])) - Decimal(str(r['total_credit'])) for r in exp_rows])
        return total_revenue - total_expense

    try:
        # --- HITUNG ASET ---
        aset_data = get_balance_by_type(['Aset Lancar', 'Aset Tetap'])
        total_aset_lancar = Decimal('0')
        total_aset_tetap = Decimal('0')
        aset_lancar_html = ""
        aset_tetap_html = ""

        for row in aset_data:
            saldo = Decimal(row['total_debit']) - Decimal(row['total_credit'])
            if saldo != 0:
                html_row = f'<tr><td style="padding-left: 30px;">{row["account_name"]}</td><td class="currency">{format_currency(saldo)}</td></tr>'
                if row['account_type'] == 'Aset Lancar':
                    total_aset_lancar += saldo
                    aset_lancar_html += html_row
                elif row['account_type'] == 'Aset Tetap':
                    total_aset_tetap += saldo
                    aset_tetap_html += html_row
        total_aset = total_aset_lancar + total_aset_tetap

        # --- HITUNG LIABILITAS ---
        liabilitas_data = get_balance_by_type(['Liabilitas'])
        total_liabilitas = Decimal('0')
        liabilitas_html = ""
        for row in liabilitas_data:
            saldo = Decimal(row['total_credit']) - Decimal(row['total_debit'])
            if saldo != 0:
                total_liabilitas += saldo
                liabilitas_html += f'<tr><td style="padding-left: 30px;">{row["account_name"]}</td><td class="currency">{format_currency(saldo)}</td></tr>'

        # --- HITUNG MODAL AKHIR ---
        modal_awal_calc = get_specific_balance('3101', is_credit_normal=True) 
        prive_calc = get_specific_balance('3102', is_credit_normal=False)
        laba_bersih_calc = calculate_net_income_upto_date()
        total_modal_akhir = modal_awal_calc + laba_bersih_calc - prive_calc

        modal_html = f"""
        <tr>
            <td style="padding-left: 30px;">Modal Akhir</td>
            <td class="currency">{format_currency(total_modal_akhir)}</td>
        </tr>
        """
        total_liabilitas_modal = total_liabilitas + total_modal_akhir

    except Exception as e:
        import traceback
        print("Error details:", traceback.format_exc())
        return f"Terjadi kesalahan: {str(e)}", 500

    # --- HTML & CSS ---
    empty_row = '<tr><td colspan="2" style="text-align: center; color: #6c757d;">- 0 -</td></tr>'

    body = f"""
    <style>
        .balance-sheet-container {{ max-width: 1100px; margin: 0 auto; padding: 20px; font-family: Arial, sans-serif; }}
        .company-header {{ text-align: center; margin-bottom: 30px; }}
        .company-name {{ font-size: 24px; font-weight: bold; margin-bottom: 5px; }}
        .report-title {{ font-size: 18px; font-weight: bold; margin-bottom: 5px; }}
        .period {{ font-size: 14px; margin-bottom: 20px; color: #555; }}
        
        /* Flex Container untuk 2 Kolom */
        .balance-columns {{ 
            display: flex; 
            gap: 30px; 
            margin-top: 20px; 
            align-items: stretch; /* Memastikan kedua kolom tingginya sama */
        }}
        
        /* Kartu per Kolom (Aset / Pasiva) */
        .balance-section {{ 
            flex: 1; 
            min-width: 300px; 
            display: flex;       /* Nested Flexbox */
            flex-direction: column; /* Susun ke bawah */
            border: 1px solid #ddd;
            padding: 15px;
            background-color: #fff;
        }}

        .balance-sheet-table {{ width: 100%; border-collapse: collapse; margin-bottom: 10px; }}
        .balance-sheet-table th, .balance-sheet-table td {{ padding: 8px 12px; border: 1px solid #ddd; }}
        .section-header {{ background-color: #f1f1f1; font-weight: bold; text-align: left; }}
        .total-row {{ background-color: #f8f9fa; font-weight: bold; border-top: 2px solid #333; }}
        .currency {{ text-align: right; font-family: 'Courier New', monospace; }}

        /* Bagian Total Paling Bawah (Sticky Bottom) */
        .grand-total-container {{
            margin-top: auto; /* Ini kuncinya: dorong ke paling bawah */
            padding-top: 15px;
        }}
        .grand-total-table {{
            width: 100%;
            border-collapse: collapse;
            background-color: #e2e6ea;
            border-top: 3px solid #333;
        }}
        .grand-total-table td {{
            padding: 12px;
            font-weight: bold;
            font-size: 1.1em;
        }}

        .download-btn {{ display: inline-block; background-color: #28a745; color: white; padding: 8px 15px; text-decoration: none; border-radius: 4px; margin-bottom: 10px; font-size: 14px; }}
        .status-box {{ margin-top: 30px; text-align: center; padding: 15px; border-radius: 8px; border: 1px solid #ddd; }}
        .status-balanced {{ background-color: #d4edda; color: #155724; border-color: #c3e6cb; }}
        .status-unbalanced {{ background-color: #f8d7da; color: #721c24; border-color: #f5c6cb; }}
    </style>

    <div class="balance-sheet-container">
        <div style="text-align: right;">
            <a href="{url_for('download_financial_position_excel', end_date=end_date)}" class="download-btn">üì• Download Excel</a>
        </div>

        <div class="company-header">
            <div class="company-name">{nama_perusahaan}</div>
            <div class="report-title">Laporan Posisi Keuangan (Neraca)</div>
            <div class="period">{period_text}</div>
        </div>

        <div class="balance-columns">
            <div class="balance-section">
                <div> <h3 style="color: #2c3e50; border-bottom: 2px solid #28a745; padding-bottom: 10px;">ASET</h3>
                    <table class="balance-sheet-table">
                        <tr class="section-header"><th colspan="2">Aset Lancar</th></tr>
                        {aset_lancar_html if aset_lancar_html else empty_row}
                        <tr class="total-row">
                            <td>Total Aset Lancar</td>
                            <td class="currency">{format_currency(total_aset_lancar)}</td>
                        </tr>
                        
                        <tr class="section-header"><th colspan="2">Aset Tetap</th></tr>
                        {aset_tetap_html if aset_tetap_html else empty_row}
                        <tr class="total-row">
                            <td>Total Aset Tetap</td>
                            <td class="currency">{format_currency(total_aset_tetap)}</td>
                        </tr>
                    </table>
                </div>

                <div class="grand-total-container">
                    <table class="grand-total-table">
                        <tr>
                            <td>TOTAL ASET</td>
                            <td class="currency">{format_currency(total_aset)}</td>
                        </tr>
                    </table>
                </div>
            </div>

            <div class="balance-section">
                <div> <h3 style="color: #2c3e50; border-bottom: 2px solid #dc3545; padding-bottom: 10px;">LIABILITAS & EKUITAS</h3>
                    <table class="balance-sheet-table">
                        <tr class="section-header"><th colspan="2">Liabilitas</th></tr>
                        {liabilitas_html if liabilitas_html else empty_row}
                        <tr class="total-row">
                            <td>Total Liabilitas</td>
                            <td class="currency">{format_currency(total_liabilitas)}</td>
                        </tr>

                        <tr class="section-header"><th colspan="2">Ekuitas</th></tr>
                        {modal_html} 
                        <tr class="total-row">
                            <td>Total Ekuitas</td>
                            <td class="currency">{format_currency(total_modal_akhir)}</td>
                        </tr>
                    </table>
                </div>

                <div class="grand-total-container">
                    <table class="grand-total-table">
                        <tr>
                            <td>TOTAL LIABILITAS & EKUITAS</td>
                            <td class="currency">{format_currency(total_liabilitas_modal)}</td>
                        </tr>
                    </table>
                </div>
            </div>
        </div>

        <div class="status-box {'status-balanced' if total_aset == total_liabilitas_modal else 'status-unbalanced'}">
    </div>
    """
    
    return render_page("Laporan Posisi Keuangan", body, sidebar_content=get_admin_sidebar_html())

@app.route("/admin/download/financial-position-excel")
def download_financial_position_excel():
    """Membuat dan mengirim file .xlsx dari Laporan Posisi Keuangan (Neraca)"""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
        
    db = get_db()
    
    # --- AMBIL PARAMETER TANGGAL ---
    end_date = request.args.get('end_date', '')
    
    # --- SETUP INFO PERUSAHAAN ---
    company_info = db.execute("SELECT * FROM company_info LIMIT 1").fetchone()
    nama_perusahaan = company_info['company_name'] if company_info else 'Budidaya Gurame'
    periode_akuntansi = company_info['accounting_period'] if company_info else '2025'
    
    # --- LOGIC CUTOFF DATE ---
    cutoff_date = f"{end_date} 23:59:59" if end_date else f"{periode_akuntansi}-12-31 23:59:59"
    period_text = f"Per {end_date}" if end_date else f"Per 31 Desember {periode_akuntansi}"

    # --- FUNGSI HELPERS (SAMA DENGAN YANG DI ATAS) ---
    def get_balance_by_type(account_types):
        query = f"""
        SELECT 
            coa.account_code, coa.account_name, coa.account_type,
            COALESCE(SUM(jd.debit), 0) as total_debit,
            COALESCE(SUM(jd.credit), 0) as total_credit
        FROM chart_of_accounts coa
        LEFT JOIN journal_details jd ON coa.account_code = jd.account_code
        LEFT JOIN journal_entries je ON jd.entry_id = je.id
        WHERE coa.account_type IN ({','.join(['?']*len(account_types))})
        AND (je.entry_timestamp <= ? OR je.entry_timestamp IS NULL)
        GROUP BY coa.account_code, coa.account_name, coa.account_type
        """
        params = account_types + [cutoff_date]
        return db.execute(query, params).fetchall()

    def get_specific_balance(account_code, is_credit_normal=True):
        query = """
        SELECT SUM(jd.debit) as d, SUM(jd.credit) as c
        FROM journal_details jd
        JOIN journal_entries je ON jd.entry_id = je.id
        WHERE jd.account_code = ? AND je.entry_timestamp <= ?
        """
        row = db.execute(query, [account_code, cutoff_date]).fetchone()
        debit = Decimal(str(row['d'] or 0))
        credit = Decimal(str(row['c'] or 0))
        return (credit - debit) if is_credit_normal else (debit - credit)

    def calculate_net_income_upto_date():
        rev_rows = get_balance_by_type(['Pendapatan'])
        total_revenue = sum([Decimal(str(r['total_credit'])) - Decimal(str(r['total_debit'])) for r in rev_rows])
        exp_rows = get_balance_by_type(['Beban'])
        total_expense = sum([Decimal(str(r['total_debit'])) - Decimal(str(r['total_credit'])) for r in exp_rows])
        return total_revenue - total_expense

    try:
        # --- HITUNG ASET ---
        aset_data = get_balance_by_type(['Aset Lancar', 'Aset Tetap'])
        total_aset_lancar = Decimal('0')
        total_aset_tetap = Decimal('0')
        
        aset_lancar_items = []
        aset_tetap_items = []
        
        for row in aset_data:
            saldo = Decimal(row['total_debit']) - Decimal(row['total_credit'])
            if saldo != 0:
                if row['account_type'] == 'Aset Lancar':
                    total_aset_lancar += saldo
                    aset_lancar_items.append({
                        'nama': row['account_name'],
                        'saldo': saldo
                    })
                elif row['account_type'] == 'Aset Tetap':
                    total_aset_tetap += saldo
                    aset_tetap_items.append({
                        'nama': row['account_name'],
                        'saldo': saldo
                    })
        
        total_aset = total_aset_lancar + total_aset_tetap

        # --- HITUNG LIABILITAS ---
        liabilitas_data = get_balance_by_type(['Liabilitas'])
        total_liabilitas = Decimal('0')
        liabilitas_items = []
        
        for row in liabilitas_data:
            saldo = Decimal(row['total_credit']) - Decimal(row['total_debit'])
            if saldo != 0:
                total_liabilitas += saldo
                liabilitas_items.append({
                    'nama': row['account_name'],
                    'saldo': saldo
                })

        # --- HITUNG MODAL AKHIR ---
        modal_awal_calc = get_specific_balance('3101', is_credit_normal=True) 
        prive_calc = get_specific_balance('3102', is_credit_normal=False)
        laba_bersih_calc = calculate_net_income_upto_date()
        total_modal_akhir = modal_awal_calc + laba_bersih_calc - prive_calc

        total_liabilitas_modal = total_liabilitas + total_modal_akhir

        # --- BUAT WORKBOOK EXCEL ---
        wb = Workbook()
        ws = wb.active
        ws.title = " Laporan Posisi Keuangan"
        
        # --- HEADER ---
        ws.append([nama_perusahaan])
        ws.append(["Laporan Posisi Keuangan"])
        ws.append([period_text])
        ws.append([])  # Baris kosong
        
        # --- FORMAT HEADER ---
        for row in range(1, 4):
            ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True, size=14 if row == 1 else 12)
            ws.cell(row=row, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')
        
        # --- ASET (KOLOM KIRI) ---
        current_row = 5
        
        # Header Aset
        ws.cell(row=current_row, column=1, value="ASET").font = openpyxl.styles.Font(bold=True, size=12)
        current_row += 1
        
        # Aset Lancar
        ws.cell(row=current_row, column=1, value="Aset Lancar").font = openpyxl.styles.Font(bold=True)
        current_row += 1
        
        for item in aset_lancar_items:
            ws.cell(row=current_row, column=1, value=item['nama'])
            ws.cell(row=current_row, column=2, value=float(item['saldo']))
            current_row += 1
            
        if not aset_lancar_items:
            ws.cell(row=current_row, column=1, value="- 0 -")
            current_row += 1
            
        # Total Aset Lancar
        ws.cell(row=current_row, column=1, value="Total Aset Lancar").font = openpyxl.styles.Font(bold=True)
        ws.cell(row=current_row, column=2, value=float(total_aset_lancar)).font = openpyxl.styles.Font(bold=True)
        current_row += 2
        
        # Aset Tetap
        ws.cell(row=current_row, column=1, value="Aset Tetap").font = openpyxl.styles.Font(bold=True)
        current_row += 1
        
        for item in aset_tetap_items:
            ws.cell(row=current_row, column=1, value=item['nama'])
            ws.cell(row=current_row, column=2, value=float(item['saldo']))
            current_row += 1
            
        if not aset_tetap_items:
            ws.cell(row=current_row, column=1, value="- 0 -")
            current_row += 1
            
        # Total Aset Tetap
        ws.cell(row=current_row, column=1, value="Total Aset Tetap").font = openpyxl.styles.Font(bold=True)
        ws.cell(row=current_row, column=2, value=float(total_aset_tetap)).font = openpyxl.styles.Font(bold=True)
        current_row += 2
        
        # Total Aset
        ws.cell(row=current_row, column=1, value="TOTAL ASET").font = openpyxl.styles.Font(bold=True, size=12)
        ws.cell(row=current_row, column=2, value=float(total_aset)).font = openpyxl.styles.Font(bold=True, size=12)
        
        # --- LIABILITAS & EKUITAS (KOLOM KANAN) ---
        current_row = 5
        
        # Header Liabilitas & Ekuitas
        ws.cell(row=current_row, column=4, value="LIABILITAS & EKUITAS").font = openpyxl.styles.Font(bold=True, size=12)
        current_row += 1
        
        # Liabilitas
        ws.cell(row=current_row, column=4, value="Liabilitas").font = openpyxl.styles.Font(bold=True)
        current_row += 1
        
        for item in liabilitas_items:
            ws.cell(row=current_row, column=4, value=item['nama'])
            ws.cell(row=current_row, column=5, value=float(item['saldo']))
            current_row += 1
            
        if not liabilitas_items:
            ws.cell(row=current_row, column=4, value="- 0 -")
            current_row += 1
            
        # Total Liabilitas
        ws.cell(row=current_row, column=4, value="Total Liabilitas").font = openpyxl.styles.Font(bold=True)
        ws.cell(row=current_row, column=5, value=float(total_liabilitas)).font = openpyxl.styles.Font(bold=True)
        current_row += 2
        
        # Ekuitas
        ws.cell(row=current_row, column=4, value="Ekuitas").font = openpyxl.styles.Font(bold=True)
        current_row += 1
        
        # Modal Akhir
        ws.cell(row=current_row, column=4, value="Modal Akhir")
        ws.cell(row=current_row, column=5, value=float(total_modal_akhir))
        current_row += 1
        
        # Total Ekuitas
        ws.cell(row=current_row, column=4, value="Total Ekuitas").font = openpyxl.styles.Font(bold=True)
        ws.cell(row=current_row, column=5, value=float(total_modal_akhir)).font = openpyxl.styles.Font(bold=True)
        current_row += 2
        
        # Total Liabilitas & Ekuitas
        ws.cell(row=current_row, column=4, value="TOTAL LIABILITAS & EKUITAS").font = openpyxl.styles.Font(bold=True, size=12)
        ws.cell(row=current_row, column=5, value=float(total_liabilitas_modal)).font = openpyxl.styles.Font(bold=True, size=12)
        
        # --- FORMAT ANGKA ---
        for col in [2, 5]:  # Kolom jumlah (B dan E)
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"Rp"#,##0.00;("Rp"#,##0.00)'
        
        # --- ATUR LEBAR KOLOM ---
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10  # Spacer
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
        
        # --- SIMPAN KE BUFFER ---
        mem_file = BytesIO()
        wb.save(mem_file)
        mem_file.seek(0)

        # --- KIRIM SEBAGAI FILE DOWNLOAD ---
        filename = f"Laporan Posisi Keuangan_{end_date.replace('-', '') if end_date else periode_akuntansi}.xlsx"
        return Response(
            mem_file.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        import traceback
        print("Error details:", traceback.format_exc())
        return f"Terjadi kesalahan: {str(e)}", 500

# ==================================
# LAPORAN JURNAL PENUTUP
# ==================================
@app.route("/admin/closing-entries", methods=['GET', 'POST'])
def closing_entries():
    """Menangani perhitungan dan eksekusi Jurnal Penutup."""
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
        
    db = get_db()
    
    # --- Kode Akun Penting ---
    MODAL_ACCOUNT = '3101'
    PRIVE_ACCOUNT = '3102'
    INCOME_SUMMARY_ACCOUNT = '3103' # Ikhtisar L/R
    
    # --- GET Request (Menampilkan Pratinjau) ---
    if request.method == 'GET':
        
        # --- PERBAIKAN DI SINI ---
        # 1. Ambil semua akun Pendapatan
        pendapatan_query = """
        SELECT c.account_code, c.account_name, (SUM(d.credit) - SUM(d.debit)) as balance
        FROM journal_details d
        JOIN chart_of_accounts c ON d.account_code = c.account_code
        WHERE c.account_type = 'Pendapatan'
        GROUP BY c.account_code, c.account_name HAVING balance != 0
        """
        revenues = db.execute(pendapatan_query).fetchall()
        
        # 2. Ambil semua akun Beban
        beban_query = """
        SELECT c.account_code, c.account_name, (SUM(d.debit) - SUM(d.credit)) as balance
        FROM journal_details d
        JOIN chart_of_accounts c ON d.account_code = c.account_code
        WHERE c.account_type = 'Beban'
        GROUP BY c.account_code, c.account_name HAVING balance != 0
        """
        expenses = db.execute(beban_query).fetchall()
        # --- AKHIR PERBAIKAN ---
        
        # 3. Ambil Prive
        prive_query = """
        SELECT (SUM(debit) - SUM(credit)) as balance
        FROM journal_details WHERE account_code = ?
        """
        prive_data = db.execute(prive_query, (PRIVE_ACCOUNT,)).fetchone()
        prive_balance = prive_data['balance'] if prive_data['balance'] else 0.0
        
        # 4. Hitung Laba/Rugi Bersih
        net_income = get_net_income(db)
        
        # Buat Pratinjau
        body = "<h2>Pratinjau Jurnal Penutup</h2>"
        body += "<p>Ini adalah jurnal yang AKAN dibuat jika Anda menekan tombol di bawah. Saldo akan diambil dari data saat ini.</p>"
        
        # ... (Sisa kode HTML pratinjau tidak berubah) ...
        body += "<b>1. Menutup Akun Pendapatan & Beban ke Ikhtisar L/R:</b><br>"
        total_pendapatan = 0
        for r in revenues:
            body += f"<span style='padding-left: 20px;'>D: {r['account_name']}... {format_currency(r['balance'])}</span><br>"
            total_pendapatan += r['balance']
        total_beban = 0
        for e in expenses:
            body += f"<span style='padding-left: 20px;'>C: {e['account_name']}... {format_currency(e['balance'])}</span><br>"
            total_beban += e['balance']
        body += f"<span style='padding-left: 20px;'>D/C: Ikhtisar Laba Rugi... (selisih)</span><br><br>"
        
        body += "<b>2. Menutup Ikhtisar L/R ke Modal:</b><br>"
        if net_income >= 0:
            body += f"<span style='padding-left: 20px;'>D: Ikhtisar Laba Rugi... {format_currency(net_income)}</span><br>"
            body += f"<span style='padding-left: 20px;'>C: Modal... {format_currency(net_income)}</span><br><br>"
        else:
            body += f"<span style='padding-left: 20px;'>D: Modal... {format_currency(abs(net_income))}</span><br>"
            body += f"<span style='padding-left: 20px;'>C: Ikhtisar Laba Rugi... {format_currency(abs(net_income))}</span><br><br>"

        body += "<b>3. Menutup Prive ke Modal:</b><br>"
        if prive_balance > 0:
            body += f"<span style='padding-left: 20px;'>D: Modal... {format_currency(prive_balance)}</span><br>"
            body += f"<span style='padding-left: 20px;'>C: Prive... {format_currency(prive_balance)}</span><br><br>"
        else:
            body += "<span style='padding-left: 20px;'><i>(Tidak ada saldo Prive)</i></span><br><br>"
            
        body += """
        <hr>
        <p style="color: red; font-weight: bold;">PERINGATAN: Aksi ini akan membuat 3-4 entri jurnal baru untuk menutup akun nominal Anda.</p>
        <form action="/admin/closing-entries" method="POST">
            <input type="submit" value="Buat Jurnal Penutup Sekarang" class="btn-red" 
                   onclick="return confirm('Anda yakin ingin membuat Jurnal Penutup?');">
        </form>
        """
        return render_page("Jurnal Penutup", body, sidebar_content=get_admin_sidebar_html())

    # --- POST Request (Membuat Jurnal) ---
    elif request.method == 'POST':
        
        trx_datetime = datetime.now()
        trx_timestamp_str = trx_datetime.strftime('%Y-%m-%d %H:%M:%S')

        # --- PERBAIKAN DI SINI ---
        # 1. Ambil semua akun Pendapatan
        revenues = db.execute(
            "SELECT c.account_code, (SUM(d.credit) - SUM(d.debit)) as balance FROM journal_details d "
            "JOIN chart_of_accounts c ON d.account_code = c.account_code "
            "WHERE c.account_type = 'Pendapatan' GROUP BY c.account_code HAVING balance != 0"
        ).fetchall()
        
        # --- PERBAIKAN DI SINI ---
        # 2. Ambil semua akun Beban
        expenses = db.execute(
            "SELECT c.account_code, (SUM(d.debit) - SUM(d.credit)) as balance FROM journal_details d "
            "JOIN chart_of_accounts c ON d.account_code = c.account_code "
            "WHERE c.account_type = 'Beban' GROUP BY c.account_code HAVING balance != 0"
        ).fetchall()
        # --- AKHIR PERBAIKAN ---
        
        # 3. Ambil Prive
        prive_data = db.execute("SELECT (SUM(debit) - SUM(credit)) as balance FROM journal_details WHERE account_code = ?", (PRIVE_ACCOUNT,)).fetchone()
        prive_balance = prive_data['balance'] if prive_data['balance'] else 0.0
        
        # 4. Hitung Laba/Rugi Bersih
        net_income = get_net_income(db)
        
        with db:
            # --- ENTRI 1: Menutup Pendapatan ---
            j_code_1 = generate_journal_code(db, trx_datetime)
            cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                (j_code_1, trx_timestamp_str, "Jurnal Penutup: Menutup Pendapatan"))
            entry_id_1 = cursor.lastrowid
            
            total_pendapatan = 0.0
            for r in revenues:
                db.execute("INSERT INTO journal_details (entry_id, account_code, debit) VALUES (?, ?, ?)",
                           (entry_id_1, r['account_code'], r['balance']))
                total_pendapatan += r['balance']
            
            if total_pendapatan > 0:
                db.execute("INSERT INTO journal_details (entry_id, account_code, credit) VALUES (?, ?, ?)",
                           (entry_id_1, INCOME_SUMMARY_ACCOUNT, total_pendapatan))

            # --- ENTRI 2: Menutup Beban ---
            j_code_2 = generate_journal_code(db, trx_datetime)
            cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                (j_code_2, trx_timestamp_str, "Jurnal Penutup: Menutup Beban"))
            entry_id_2 = cursor.lastrowid
            
            total_beban = 0.0
            for e in expenses:
                db.execute("INSERT INTO journal_details (entry_id, account_code, credit) VALUES (?, ?, ?)",
                           (entry_id_2, e['account_code'], e['balance']))
                total_beban += e['balance']
            
            if total_beban > 0:
                db.execute("INSERT INTO journal_details (entry_id, account_code, debit) VALUES (?, ?, ?)",
                           (entry_id_2, INCOME_SUMMARY_ACCOUNT, total_beban))

            # --- ENTRI 3: Menutup Ikhtisar L/R ke Modal ---
            if net_income != 0:
                j_code_3 = generate_journal_code(db, trx_datetime)
                cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                    (j_code_3, trx_timestamp_str, "Jurnal Penutup: Menutup Ikhtisar L/R ke Modal"))
                entry_id_3 = cursor.lastrowid
                
                if net_income > 0: # Laba
                    db.execute("INSERT INTO journal_details (entry_id, account_code, debit) VALUES (?, ?, ?)",
                               (entry_id_3, INCOME_SUMMARY_ACCOUNT, net_income))
                    db.execute("INSERT INTO journal_details (entry_id, account_code, credit) VALUES (?, ?, ?)",
                               (entry_id_3, MODAL_ACCOUNT, net_income))
                else: # Rugi
                    db.execute("INSERT INTO journal_details (entry_id, account_code, debit) VALUES (?, ?, ?)",
                               (entry_id_3, MODAL_ACCOUNT, abs(net_income)))
                    db.execute("INSERT INTO journal_details (entry_id, account_code, credit) VALUES (?, ?, ?)",
                               (entry_id_3, INCOME_SUMMARY_ACCOUNT, abs(net_income)))

            # --- ENTRI 4: Menutup Prive ke Modal ---
            if prive_balance > 0:
                j_code_4 = generate_journal_code(db, trx_datetime)
                cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                    (j_code_4, trx_timestamp_str, "Jurnal Penutup: Menutup Prive ke Modal"))
                entry_id_4 = cursor.lastrowid
                
                db.execute("INSERT INTO journal_details (entry_id, account_code, debit) VALUES (?, ?, ?)",
                           (entry_id_4, MODAL_ACCOUNT, prive_balance))
                db.execute("INSERT INTO journal_details (entry_id, account_code, credit) VALUES (?, ?, ?)",
                           (entry_id_4, PRIVE_ACCOUNT, prive_balance))

        return redirect(url_for('general_journal', success="Jurnal Penutup telah berhasil dibuat!"))

    # --- POST Request (Membuat Jurnal) ---
    elif request.method == 'POST':
        
        trx_datetime = datetime.now()
        trx_timestamp_str = trx_datetime.strftime('%Y-%m-%d %H:%M:%S')

        # 1. Ambil semua akun Pendapatan
        revenues = db.execute(
            "SELECT account_code, (SUM(credit) - SUM(debit)) as balance FROM journal_details d "
            "JOIN chart_of_accounts c ON d.account_code = c.account_code "
            "WHERE c.account_type = 'Pendapatan' GROUP BY account_code HAVING balance != 0"
        ).fetchall()
        
        # 2. Ambil semua akun Beban
        expenses = db.execute(
            "SELECT account_code, (SUM(debit) - SUM(credit)) as balance FROM journal_details d "
            "JOIN chart_of_accounts c ON d.account_code = c.account_code "
            "WHERE c.account_type = 'Beban' GROUP BY account_code HAVING balance != 0"
        ).fetchall()
        
        # 3. Ambil Prive
        prive_data = db.execute("SELECT (SUM(debit) - SUM(credit)) as balance FROM journal_details WHERE account_code = ?", (PRIVE_ACCOUNT,)).fetchone()
        prive_balance = prive_data['balance'] if prive_data['balance'] else 0.0
        
        # 4. Hitung Laba/Rugi Bersih
        net_income = get_net_income(db)
        
        with db:
            # --- ENTRI 1: Menutup Pendapatan ---
            j_code_1 = generate_journal_code(db, trx_datetime)
            cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                (j_code_1, trx_timestamp_str, "Jurnal Penutup: Menutup Pendapatan"))
            entry_id_1 = cursor.lastrowid
            
            total_pendapatan = 0.0
            for r in revenues:
                db.execute("INSERT INTO journal_details (entry_id, account_code, debit) VALUES (?, ?, ?)",
                           (entry_id_1, r['account_code'], r['balance']))
                total_pendapatan += r['balance']
            
            if total_pendapatan > 0:
                db.execute("INSERT INTO journal_details (entry_id, account_code, credit) VALUES (?, ?, ?)",
                           (entry_id_1, INCOME_SUMMARY_ACCOUNT, total_pendapatan))

            # --- ENTRI 2: Menutup Beban ---
            j_code_2 = generate_journal_code(db, trx_datetime)
            cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                (j_code_2, trx_timestamp_str, "Jurnal Penutup: Menutup Beban"))
            entry_id_2 = cursor.lastrowid
            
            total_beban = 0.0
            for e in expenses:
                db.execute("INSERT INTO journal_details (entry_id, account_code, credit) VALUES (?, ?, ?)",
                           (entry_id_2, e['account_code'], e['balance']))
                total_beban += e['balance']
            
            if total_beban > 0:
                db.execute("INSERT INTO journal_details (entry_id, account_code, debit) VALUES (?, ?, ?)",
                           (entry_id_2, INCOME_SUMMARY_ACCOUNT, total_beban))

            # --- ENTRI 3: Menutup Ikhtisar L/R ke Modal ---
            if net_income != 0:
                j_code_3 = generate_journal_code(db, trx_datetime)
                cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                    (j_code_3, trx_timestamp_str, "Jurnal Penutup: Menutup Ikhtisar L/R ke Modal"))
                entry_id_3 = cursor.lastrowid
                
                if net_income > 0: # Laba
                    db.execute("INSERT INTO journal_details (entry_id, account_code, debit) VALUES (?, ?, ?)",
                               (entry_id_3, INCOME_SUMMARY_ACCOUNT, net_income))
                    db.execute("INSERT INTO journal_details (entry_id, account_code, credit) VALUES (?, ?, ?)",
                               (entry_id_3, MODAL_ACCOUNT, net_income))
                else: # Rugi
                    db.execute("INSERT INTO journal_details (entry_id, account_code, debit) VALUES (?, ?, ?)",
                               (entry_id_3, MODAL_ACCOUNT, abs(net_income)))
                    db.execute("INSERT INTO journal_details (entry_id, account_code, credit) VALUES (?, ?, ?)",
                               (entry_id_3, INCOME_SUMMARY_ACCOUNT, abs(net_income)))

            # --- ENTRI 4: Menutup Prive ke Modal ---
            if prive_balance > 0:
                j_code_4 = generate_journal_code(db, trx_datetime)
                cursor = db.execute("INSERT INTO journal_entries (journal_code, entry_timestamp, description) VALUES (?, ?, ?)",
                                    (j_code_4, trx_timestamp_str, "Jurnal Penutup: Menutup Prive ke Modal"))
                entry_id_4 = cursor.lastrowid
                
                db.execute("INSERT INTO journal_details (entry_id, account_code, debit) VALUES (?, ?, ?)",
                           (entry_id_4, MODAL_ACCOUNT, prive_balance))
                db.execute("INSERT INTO journal_details (entry_id, account_code, credit) VALUES (?, ?, ?)",
                           (entry_id_4, PRIVE_ACCOUNT, prive_balance))

        return redirect(url_for('general_journal', success="Jurnal Penutup telah berhasil dibuat!"))

@app.route("/process-manual-payment", methods=['POST'])
def process_manual_payment():
    if session.get('role') != 'consumer': return redirect(url_for('index'))
    
    # 1. Ambil Data
    item_id = request.form['item_id']
    qty = request.form['qty']
    amount = request.form['amount']
    username = session.get('username')
    
    # Generate ID Transaksi
    order_id = f"MANUAL-{uuid.uuid4().hex[:6]}"
    
    # 2. Simpan ke Database
    db = get_db()
    with db:
        db.execute(
            """INSERT INTO online_payments 
               (order_id, amount, status, proof_image, item_id, qty, buyer_name) 
               VALUES (?, ?, 'pending', '-', ?, ?, ?)""",
            (order_id, amount, item_id, qty, username)
        )
    
    # 3. Tampilkan Halaman Terima Kasih
    body = f"""
    <style>
        /* Override background body agar sesuai tema gradasi */
        body {{
            background: linear-gradient(135deg, #408080 0%, #D7F7FF 100%) !important;
            min-height: 100vh;
        }}
        
        /* Hilangkan judul h1 bawaan render_page */
        h1 {{ display: none !important; }}

        /* Wrapper khusus untuk menengahkan kartu */
        .thank-wrapper {{
            display: flex;
            justify-content: center;
            align-items: center;
            width: 100%;
            min-height: 80vh; /* Tinggi area konten */
        }}

        /* Kartu Terima Kasih */
        .thank-card {{
            background-color: #63a39c; /* Warna Hijau Tosca */
            width: 450px;
            max-width: 90%;
            padding: 50px 40px;
            border-radius: 25px;
            text-align: center;
            color: white;
            box-shadow: 0 20px 50px rgba(0,0,0,0.2);
            /* Hapus border agar lebih bersih sesuai gambar */
        }}
        
        .thank-title {{
            font-size: 2.2em;
            font-weight: 800;
            margin-bottom: 20px;
            text-transform: uppercase;
            letter-spacing: 1px;
            line-height: 1.1;
        }}
        
        .thank-desc {{
            font-size: 1em;
            line-height: 1.6;
            margin-bottom: 40px;
            font-weight: 500;
            opacity: 0.95;
        }}
        
        .btn-main-menu {{
            background-color: #478ba2; /* Biru laut */
            color: white;
            text-decoration: none;
            padding: 12px 35px;
            border-radius: 10px;
            font-weight: 600;
            font-size: 1em;
            display: inline-block;
            transition: 0.3s;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
            border: 1px solid rgba(255,255,255,0.3);
        }}
        
        .btn-main-menu:hover {{
            background-color: #356d80;
            transform: translateY(-2px);
        }}
    </style>

    <div class="thank-wrapper">
        <div class="thank-card">
            <div class="thank-title">TERIMA KASIH!</div>
            <div class="thank-desc">
                Terima kasih sudah mampir! Semoga setelah<br>
                ini kamu menemukan hal-hal seru lainnya<br>
                yang bikin kamu tetap happy.
            </div>
            
            <a href="/home" class="btn-main-menu">Kembali ke Main Menu</a>
        </div>
    </div>
    """
    
    # Render halaman
    return render_page("", body)

# --- RUTE KHUSUS QRIS MANUAL ---
@app.route("/purchase-manual-qris", methods=['POST'])
def purchase_manual_qris():
    if session.get('role') != 'consumer': return redirect(url_for('index'))
    
    # Ambil data dari form
    item_id = request.form['item_id']
    qty = float(request.form['qty'])
    
    db = get_db()
    item = db.execute("SELECT * FROM inventory_items WHERE id = ?", (item_id,)).fetchone()
    total_price = qty * item['sale_price_per_unit']
    
    body = f"""
    <style>
        /* CSS Halaman */
        body {{
            background: linear-gradient(135deg, #408080 0%, #D7F7FF 100%);
            min-height: 100vh;
            font-family: 'Poppins', sans-serif;
            margin: 0;
            position: relative;
        }}
        
        /* Override Global Form Style (INI SOLUSINYA) */
        /* Kita paksa form di dalam card ini agar transparan */
        .qris-card form {{
            background: transparent !important;
            padding: 0 !important;
            box-shadow: none !important;
            border: none !important;
            margin: 0 !important;
        }}
        
        /* Tombol Dashboard Kapsul */
        .btn-dashboard-pill {{
            position: absolute;
            top: 40px; left: 50px;
            display: inline-flex; align-items: center; gap: 10px;
            padding: 10px 30px;
            border: 2px solid rgba(255, 255, 255, 0.8);
            border-radius: 50px;
            color: white; text-decoration: none;
            font-weight: 600; font-size: 1.2em;
            transition: background 0.3s; z-index: 10;
        }}
        .btn-dashboard-pill:hover {{ background: rgba(255, 255, 255, 0.2); }}
        .icon-back {{
            width: 12px; height: 12px;
            border-top: 3px solid white; border-left: 3px solid white;
            transform: rotate(-45deg);
        }}

        .qris-wrapper {{
            display: flex; justify-content: center; align-items: center;
            min-height: 100vh; padding: 20px;
        }}
        
        /* Kartu Utama */
        .qris-card {{
            background-color: #63a39c; /* Hijau Tosca */
            width: 450px; max-width: 100%;
            padding: 50px 40px;
            border-radius: 25px;
            text-align: center; color: white;
            box-shadow: 0 20px 50px rgba(0,0,0,0.2);
            position: relative;
        }}
        
        .qris-title {{ font-size: 1.8em; font-weight: 700; margin-bottom: 15px; }}
        .qris-total {{ font-size: 1.3em; font-weight: 500; margin-bottom: 25px; }}
        .total-amount {{ font-weight: 900; }}
        .qris-scan-text {{ font-size: 0.95em; margin-bottom: 15px; opacity: 0.9; }}
        .qris-img-container {{
            background: transparent; padding: 10px;
            display: inline-block; border-radius: 15px;
            border: 2px solid rgba(255,255,255,0.4);
            margin-bottom: 20px;
        }}
        .qris-img {{
            width: 180px; height: 180px; object-fit: contain;
            display: block; border-radius: 10px;
        }}
        
        .qris-manual-text {{ font-size: 0.9em; margin-bottom: 40px; font-weight: 500; opacity: 0.9; }}
        
        /* Tombol "Saya Sudah Bayar" */
        .btn-paid-confirm {{
            background-color: #4285F4; /* Biru */
            color: white; border: none;
            padding: 15px 0; width: 100%;
            border-radius: 10px;
            font-weight: 700; font-size: 1.1em;
            cursor: pointer; transition: 0.3s;
            box-shadow: 0 5px 15px rgba(0,0,0,0.2);
        }}
        .btn-paid-confirm:hover {{ background-color: #3367D6; transform: translateY(-2px); }}
        
        h1 {{ display: none !important; }}
    </style>

    <a href="/home" class="btn-dashboard-pill">
        <div class="icon-back"></div> Dashboard
    </a>

    <div class="qris-wrapper">
        <div class="qris-card">
            <div class="qris-title">Selesaikan Pembayaran</div>
            
            <div class="qris-total">
                Total Tagihan: <span class="total-amount">{format_currency(total_price)}</span>
            </div>
            
            <div class="qris-scan-text">Scan QRIS DANA di bawah ini:</div>
            
            <div class="qris-img-container">
                <img src="/static/qris_dana.png" class="qris-img" alt="QRIS Code">
            </div>
            
            <div class="qris-manual-text">Atau transfer manual ke DANA: <b>0822-4191-5050</b></div>
            
            <form action="/process-manual-payment" method="POST">
                <input type="hidden" name="item_id" value="{item_id}">
                <input type="hidden" name="qty" value="{qty}">
                <input type="hidden" name="amount" value="{total_price}">
                
                <button type="submit" class="btn-paid-confirm">Saya sudah bayar</button>
            </form>
        </div>
    </div>
    """
    
    return render_page("", body)
    
@app.route("/admin/verify-payments")
def verify_payments():
    if session.get('role') != 'admin': return redirect(url_for('index'))
    
    db = get_db()
    # Query transaksi pending
    payments = db.execute("SELECT * FROM online_payments WHERE status = 'pending' ORDER BY created_at DESC").fetchall()
    
    # Pesan Notifikasi
    success_msg = request.args.get('success')
    error_msg = request.args.get('error')
    
    alert_html = ""
    if success_msg:
        alert_html = f'<div style="background:#d4edda; color:#155724; padding:10px; border-radius:5px; margin-bottom:15px;">{success_msg}</div>'
    if error_msg:
        alert_html = f'<div style="background:#f8d7da; color:#721c24; padding:10px; border-radius:5px; margin-bottom:15px;">{error_msg}</div>'
    
    rows = ""
    for i, p in enumerate(payments, start=1):
        rows += f"""
        <tr>
            <td style="text-align: center;">{i}</td> 
            <td>{p['created_at']}</td>
            <td>{p['buyer_name']}</td>
            <td style="font-weight: bold; color: green;">{format_currency(p['amount'])}</td>
            <td style="text-align: center;">
                <div style="display: flex; gap: 5px; justify-content: center;">
                    <form action="/admin/approve-payment/{p['order_id']}" method="POST">
                        <input type="submit" value="‚úÖ Terima" class="btn-kotak" 
                            onclick="return confirm('Pastikan saldo sudah masuk. Terima pembayaran ini?');">
                    </form>

                    <form action="/admin/reject-payment/{p['order_id']}" method="POST">
                        <input type="submit" value="‚ùå Tolak" class="btn-kotak-red" 
                            onclick="return confirm('Yakin ingin menolak? Status akan menjadi Gagal di sisi konsumen.');">
                    </form>
                </div>
            </td>
        </tr>
        """
        
    body = f"""
    <style>
        /* --- STYLE TOMBOL HIJAU (TERIMA) --- */
        .btn-kotak {{
            background-color: #28a745 !important; 
            color: white !important;
            padding: 8px 15px !important;
            font-weight: bold !important;
            font-size: 0.85em !important;
            border: none !important;
            border-radius: 4px !important;
            cursor: pointer !important;
            transition: background-color 0.2s !important;
        }}
        .btn-kotak:hover {{ background-color: #218838 !important; }}

        /* --- STYLE TOMBOL MERAH (TOLAK) --- */
        .btn-kotak-red {{
            background-color: #dc3545 !important; 
            color: white !important;
            padding: 8px 15px !important;
            font-weight: bold !important;
            font-size: 0.85em !important;
            border: none !important;
            border-radius: 4px !important;
            cursor: pointer !important;
            transition: background-color 0.2s !important;
        }}
        .btn-kotak-red:hover {{ background-color: #c82333 !important; }}

        /* Tabel Style */
        .table-journal {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
            background: white;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        .table-journal th {{
            background-color: #e9ecef !important;
            color: black !important;
            padding: 12px;
            border: 1px solid #ddd;
            text-align: center;
            font-weight: bold;
        }}
        .table-journal td {{
            padding: 10px;
            border: 1px solid #ddd;
            vertical-align: middle;
        }}
    </style>

    <h3>Verifikasi Pembayaran QRIS</h3>
    
    {alert_html}

    <div style="background-color: #e7f3ff; padding: 15px; border-left: 4px solid #007bff; margin-bottom: 20px;">
        <strong>Panduan Admin:</strong><br>
        Cek aplikasi DANA/E-Wallet Anda. <br>
        - Klik <b>Terima</b> jika uang sudah masuk.<br>
        - Klik <b>Tolak</b> jika bukti palsu atau uang belum masuk.
    </div>
    
    <table class="table-journal"> 
        <thead>
            <tr>
                <th style="width: 50px;">No</th> 
                <th>Tanggal</th>
                <th>Pembeli</th>
                <th>Nominal Masuk</th>
                <th style="width: 250px;">Aksi</th>
            </tr>
        </thead>
        <tbody>
            {rows if rows else '<tr><td colspan="5" style="text-align:center; padding:20px;">Tidak ada pembayaran pending.</td></tr>'}
        </tbody>
    </table>
    """
    return render_page("Verifikasi Pembayaran", body, sidebar_content=get_admin_sidebar_html())

@app.route("/admin/approve-payment/<order_id>", methods=['POST'])
def approve_payment(order_id):
    if session.get('role') != 'admin': return redirect(url_for('index'))
    
    db = get_db()
    trx = db.execute("SELECT * FROM online_payments WHERE order_id = ?", (order_id,)).fetchone()
    
    if trx and trx['status'] == 'pending':
        item = db.execute("SELECT * FROM inventory_items WHERE id = ?", (trx['item_id'],)).fetchone()
        
        # CATAT JURNAL & KURANGI STOK (Pakai fungsi helper yang sudah ada)
        record_sale_transaction(db, item, trx['qty'], trx['buyer_name'], 'Transfer Manual', trx['amount'])
        
        # Update status jadi verified
        with db:
            db.execute("UPDATE online_payments SET status = 'verified' WHERE order_id = ?", (order_id,))
            
        return redirect(url_for('verify_payments', success="Pembayaran dikonfirmasi & Jurnal tercatat!"))
        
    return redirect(url_for('verify_payments', error="Gagal verifikasi"))

@app.route("/admin/reject-payment/<order_id>", methods=['POST'])
def reject_payment(order_id):
    """Menolak pembayaran: Ubah status jadi 'rejected' (Gagal)"""
    if session.get('role') != 'admin': return redirect(url_for('index'))
    
    db = get_db()
    
    # Cek transaksi
    trx = db.execute("SELECT * FROM online_payments WHERE order_id = ?", (order_id,)).fetchone()
    
    if trx and trx['status'] == 'pending':
        # Update status jadi rejected
        with db:
            db.execute("UPDATE online_payments SET status = 'rejected' WHERE order_id = ?", (order_id,))
            
        # Redirect kembali dengan pesan sukses
        return redirect(url_for('verify_payments', error="Pembayaran telah ditolak. Status konsumen berubah menjadi Gagal."))
        
    return redirect(url_for('verify_payments', error="Gagal memproses penolakan."))

# --- Menjalankan Aplikasi ---
if __name__ == '__main__':
    init_db() 
    app.run(debug=True)